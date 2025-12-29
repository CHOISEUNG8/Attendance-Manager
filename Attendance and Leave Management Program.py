"""
근태관리 프로그램
PySide6 + QTableWidget 버전
"""

# --- App identity (UI display) ---
APP_NAME_EN = "WB Attendance Management Program"
APP_VERSION = "v4.1"
APP_TITLE = f"{APP_NAME_EN} | Made by AQMAN | {APP_VERSION}"

# PySide6 import 확인 및 오류 처리
try:
    from PySide6.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, 
                                    QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, 
                                    QHeaderView, QLabel, QLineEdit, QMessageBox, QDialog, 
                                    QComboBox, QDateEdit, QSpinBox, QDialogButtonBox, QTextEdit,
                                    QAbstractItemView, QFileDialog, QFrame, QStatusBar, QStyledItemDelegate, QCheckBox)
    from PySide6.QtCore import Qt, QDate, QTime, Signal, QModelIndex, QRect, QTimer
    from PySide6.QtGui import QColor, QBrush, QFont, QClipboard, QKeyEvent, QIcon, QPainter, QPen
    PYSIDE6_AVAILABLE = True
except ImportError as e:
    PYSIDE6_AVAILABLE = False
    PYSIDE6_ERROR = str(e)

import sqlite3
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import traceback


class DatabaseManager:
    """데이터베이스 관리 클래스"""
    
    def __init__(self, db_path="leave_attendance.db"):
        # exe 파일로 실행 중인지 확인
        import sys
        import os
        import shutil
        
        if getattr(sys, 'frozen', False):
            # exe로 실행 중인 경우
            # 먼저 사용자 AppData 폴더에 저장 시도 (권한 문제 방지)
            try:
                appdata_dir = os.path.join(os.getenv('APPDATA', ''), '근태관리프로그램')
                os.makedirs(appdata_dir, exist_ok=True)
                appdata_db_path = os.path.join(appdata_dir, db_path)
                
                # AppData 폴더에 쓰기 권한이 있는지 테스트
                test_file = os.path.join(appdata_dir, '.test_write')
                try:
                    with open(test_file, 'w') as f:
                        f.write('test')
                    os.remove(test_file)
                    # 쓰기 가능하면 AppData 폴더 사용
                    self.db_path = appdata_db_path
                except (IOError, OSError):
                    # AppData 폴더에 쓰기 불가능하면 EXE 파일 디렉토리 사용
                    exe_dir = os.path.dirname(sys.executable)
                    self.db_path = os.path.join(exe_dir, db_path)
            except Exception:
                # 모든 시도 실패 시 EXE 파일 디렉토리 사용
                exe_dir = os.path.dirname(sys.executable)
                self.db_path = os.path.join(exe_dir, db_path)
            
            # 포함된 데이터베이스 파일이 있으면 복사
            try:
                # PyInstaller의 임시 디렉토리에서 데이터베이스 파일 찾기
                if hasattr(sys, '_MEIPASS'):
                    bundled_db = os.path.join(sys._MEIPASS, db_path)
                    if os.path.exists(bundled_db):
                        # 실행 디렉토리에 데이터베이스가 없거나, 포함된 데이터베이스가 더 최신이면 복사
                        if not os.path.exists(self.db_path):
                            # 데이터베이스 파일이 없으면 포함된 파일 복사
                            try:
                                shutil.copy2(bundled_db, self.db_path)
                            except (IOError, OSError, PermissionError):
                                # 복사 실패 시 AppData 폴더에 복사 시도
                                try:
                                    appdata_dir = os.path.join(os.getenv('APPDATA', ''), '근태관리프로그램')
                                    os.makedirs(appdata_dir, exist_ok=True)
                                    appdata_db_path = os.path.join(appdata_dir, db_path)
                                    shutil.copy2(bundled_db, appdata_db_path)
                                    self.db_path = appdata_db_path
                                except Exception:
                                    pass
                        # 실행 디렉토리에 데이터베이스가 있으면 그대로 사용 (사용자가 수정한 데이터 유지)
            except Exception as e:
                # 오류 발생 시에도 계속 진행 (기존 데이터베이스 사용)
                pass
        else:
            # Python 스크립트로 실행 중인 경우
            self.db_path = db_path
        
        self.init_database()
    
    def get_connection(self):
        """데이터베이스 연결 반환"""
        conn = sqlite3.connect(self.db_path, timeout=30.0)  # 타임아웃 30초 설정
        return conn
    
    def init_database(self):
        """데이터베이스 초기화 및 테이블 생성"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # 직원 정보 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                department TEXT NOT NULL,
                position TEXT NOT NULL,
                name TEXT NOT NULL,
                hire_date DATE NOT NULL,
                display_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(department, position, name, hire_date)
            )
        """)
        
        # display_order 컬럼이 없으면 추가 (기존 데이터베이스 호환성)
        cursor.execute("PRAGMA table_info(employees)")
        columns = [col[1] for col in cursor.fetchall()]
        if 'display_order' not in columns:
            cursor.execute("ALTER TABLE employees ADD COLUMN display_order INTEGER DEFAULT 0")
        
        # is_active 컬럼이 없으면 추가 (중도 퇴사자 숨김 기능)
        if 'is_active' not in columns:
            cursor.execute("ALTER TABLE employees ADD COLUMN is_active INTEGER DEFAULT 1")
        
        # resignation_date 컬럼이 없으면 추가 (퇴사일 저장)
        if 'resignation_date' not in columns:
            cursor.execute("ALTER TABLE employees ADD COLUMN resignation_date DATE")
        
        # phone 컬럼이 없으면 추가 (연락처 저장)
        if 'phone' not in columns:
            cursor.execute("ALTER TABLE employees ADD COLUMN phone TEXT")
        
        # email 컬럼이 없으면 추가 (이메일 저장)
        if 'email' not in columns:
            cursor.execute("ALTER TABLE employees ADD COLUMN email TEXT")
        
        # 연월차 관리 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS leave_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                leave_type TEXT NOT NULL,
                leave_date DATE NOT NULL,
                leave_amount REAL NOT NULL,
                year INTEGER NOT NULL,
                month INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees(id),
                UNIQUE(employee_id, leave_date, leave_type)
            )
        """)

        # leave_records 테이블 컬럼 호환성 검사 (기존 데이터베이스 업그레이드)
        cursor.execute("PRAGMA table_info(leave_records)")
        lr_columns = [col[1] for col in cursor.fetchall()]
        if 'year' not in lr_columns:
            cursor.execute("ALTER TABLE leave_records ADD COLUMN year INTEGER")
        if 'month' not in lr_columns:
            cursor.execute("ALTER TABLE leave_records ADD COLUMN month INTEGER")
        
        # 연월차 소멸 내역 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS leave_expirations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                leave_type TEXT NOT NULL,
                expired_amount REAL NOT NULL,
                expiration_date DATE NOT NULL,
                year INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees(id)
            )
        """)
        
        # 출퇴근 기록 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                work_date DATE NOT NULL,
                arrival_time TIME,
                departure_time TIME,
                early_arrival INTEGER DEFAULT 0,
                late_arrival INTEGER DEFAULT 0,
                late_departure INTEGER DEFAULT 0,
                leave_type TEXT,
                remarks TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees(id),
                UNIQUE(employee_id, work_date)
            )
        """)

        # attendance_records 테이블 컬럼 호환성 검사 (기존 데이터베이스 업그레이드)
        cursor.execute("PRAGMA table_info(attendance_records)")
        ar_columns = [col[1] for col in cursor.fetchall()]
        if 'early_arrival' not in ar_columns:
            cursor.execute("ALTER TABLE attendance_records ADD COLUMN early_arrival INTEGER DEFAULT 0")
        if 'late_arrival' not in ar_columns:
            cursor.execute("ALTER TABLE attendance_records ADD COLUMN late_arrival INTEGER DEFAULT 0")
        if 'late_departure' not in ar_columns:
            cursor.execute("ALTER TABLE attendance_records ADD COLUMN late_departure INTEGER DEFAULT 0")
        if 'leave_type' not in ar_columns:
            cursor.execute("ALTER TABLE attendance_records ADD COLUMN leave_type TEXT")
        if 'remarks' not in ar_columns:
            cursor.execute("ALTER TABLE attendance_records ADD COLUMN remarks TEXT")
        
        # 연월차 관리대장 수동 입력 값 저장 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS leave_manual_values (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                year INTEGER NOT NULL,
                column_index INTEGER NOT NULL,
                manual_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees(id),
                UNIQUE(employee_id, year, column_index)
            )
        """)
        
        # 연도별 잔여수 저장 테이블 (1년 이상 재직인원의 이전 년도 잔여수 저장)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS leave_remaining_by_year (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                year INTEGER NOT NULL,
                remaining_amount REAL NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees(id),
                UNIQUE(employee_id, year)
            )
        """)
        
        conn.commit()
        conn.close()


class LeaveCalculator:
    """연월차 계산 클래스"""
    
    # 1년 이상 재직인원 리스트 (TODAY 기준)
    ONE_YEAR_OR_MORE_EMPLOYEES = [
        '전금희', '박진성', '김미라', '김아름벌', '맹기열', 
        '강지승', '신동인', '두보라', '오세원', '황혜선', '김현경'
    ]
    
    def __init__(self, db_manager):
        self.db = db_manager
    
    def is_one_year_or_more(self, name, hire_date, target_date=None):
        """TODAY 기준으로 1년 이상 재직인원인지 확인"""
        if target_date is None:
            target_date = datetime.now().date()
        
        if isinstance(hire_date, str):
            hire_date = datetime.strptime(hire_date, "%Y-%m-%d").date()
        if isinstance(target_date, str):
            target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
        
        # 명시적으로 1년 이상 재직인원 리스트에 있으면 1년 이상자로 처리
        if name in self.ONE_YEAR_OR_MORE_EMPLOYEES:
            return True
        
        # TODAY 기준으로 입사일로부터 경과 일수 계산
        days_passed = (target_date - hire_date).days
        return days_passed >= 365
    
    def calculate_monthly_leave(self, hire_date, target_date):
        """입사 1년 미만 직원의 월차 계산
        입사일을 기준으로 1개월 만근 시 연차 1개 발생 (최대 11개)
        """
        if isinstance(hire_date, str):
            hire_date = datetime.strptime(hire_date, "%Y-%m-%d").date()
        if isinstance(target_date, str):
            target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
        
        if (target_date - hire_date).days >= 365:
            return 0
        
        # 입사일 기준으로 경과한 개월 수 계산
        months_passed = (target_date.year - hire_date.year) * 12 + (target_date.month - hire_date.month)
        if target_date.day < hire_date.day:
            months_passed -= 1
        
        return min(months_passed, 11)
    
    def calculate_annual_leave(self, hire_date, target_date):
        """연차 계산
        - 입사 1년 경과 시: 입사일 기준으로 연차 15개 부여
        - 근속연수 증가 시 연차 추가:
          * 만3년차: 16일 (예: 2023년 5월 30일 입사 → 2026년 5월 30일에 16일 발생)
          * 만5년차: 17일
          * 만7년차: 18일
          * 만25년차: 25일 (최대)
        """
        if isinstance(hire_date, str):
            hire_date = datetime.strptime(hire_date, "%Y-%m-%d").date()
        if isinstance(target_date, str):
            target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
        
        # 만 근무 연수 계산 (입사일 기준)
        years_passed = (target_date.year - hire_date.year)
        if target_date.month < hire_date.month or (target_date.month == hire_date.month and target_date.day < hire_date.day):
            years_passed -= 1
        
        if years_passed < 1:
            return 0
        
        # 입사 1년 경과 시 연차 15개 부여
        annual_leave = 15
        
        # 근속연수 증가 시 연차 추가
        # 만3년차부터 시작: 만3년(16일), 만5년(17일), 만7년(18일)...
        if years_passed >= 3:  # 만3년 이상
            # 만3년부터 2년마다 1개씩 추가
            # 만3년: +1 (16일), 만5년: +2 (17일), 만7년: +3 (18일)...
            additional_leaves = (years_passed - 3) // 2 + 1
            annual_leave += additional_leaves
        
        # 최대 25일 제한
        return min(annual_leave, 25)
    
    def check_monthly_leave_expiration(self, employee_id, hire_date, target_date):
        """월차 소멸 확인 및 처리 - 1년 미만자만 처리
        입사기념일 기준으로 1년 되는 날 소멸 (1년 이상자는 월차 없음)
        """
        if isinstance(hire_date, str):
            hire_date = datetime.strptime(hire_date, "%Y-%m-%d").date()
        if isinstance(target_date, str):
            target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        # 직원 정보 조회
        cursor.execute("SELECT name FROM employees WHERE id = ?", (employee_id,))
        result = cursor.fetchone()
        if not result:
            conn.close()
            return 0
        
        name = result[0]
        
        # 1년 이상자는 월차 소멸 없음
        if self.is_one_year_or_more(name, hire_date, target_date):
            conn.close()
            return 0
        
        # 1년 미만자: 입사기념일 기준으로 1년 되는 날 소멸
        try:
            one_year_date = datetime(hire_date.year + 1, hire_date.month, hire_date.day).date()
        except ValueError:
            # 2월 29일 같은 경우 처리
            one_year_date = datetime(hire_date.year + 1, hire_date.month, hire_date.day - 1).date()
        
        # 정확히 1년 되는 날에 소멸 처리
        if target_date >= one_year_date:
            # 기존 소멸 기록이 있는지 확인 (중복 방지)
            cursor.execute("""
                SELECT id FROM leave_expirations
                WHERE employee_id = ? AND leave_type = '월차' AND expiration_date = ?
            """, (employee_id, one_year_date))
            existing = cursor.fetchone()
            
            if not existing:
                cursor.execute("""
                    SELECT SUM(leave_amount) as total_used
                    FROM leave_records
                    WHERE employee_id = ? AND leave_type = '월차'
                    AND leave_date < ?
                """, (employee_id, one_year_date))
                result = cursor.fetchone()
                total_used = result[0] if result[0] else 0
                
                # 입사기념일 기준으로 1년 전까지 생성된 월차 수 (최대 11개)
                max_monthly_leave = min(11, self.calculate_monthly_leave(hire_date, one_year_date - timedelta(days=1)))
                expired_amount = max_monthly_leave - total_used
                
                if expired_amount > 0:
                    cursor.execute("""
                        INSERT INTO leave_expirations 
                        (employee_id, leave_type, expired_amount, expiration_date, year)
                        VALUES (?, ?, ?, ?, ?)
                    """, (employee_id, '월차', expired_amount, one_year_date, target_date.year))
                    conn.commit()
                    conn.close()
                    return expired_amount
            
            conn.close()
        return 0
    
    def check_annual_leave_expiration(self, employee_id, target_date):
        """연차 소멸 확인 및 처리
        1년 이상자: 입사기념일 기준으로 연차 생성, 다음년도 입사기념일까지 사용 안하면 소멸
        1년 미만자: 소멸 없음 (1년 경과 후 1년 이상자 로직으로 전환)
        """
        if isinstance(target_date, str):
            target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        # 직원 정보 조회
        cursor.execute("SELECT name, hire_date FROM employees WHERE id = ?", (employee_id,))
        result = cursor.fetchone()
        if not result:
            conn.close()
            return 0
        
        name, hire_date_str = result
        hire_date = datetime.strptime(hire_date_str, "%Y-%m-%d").date()
        
        # 1년 이상자 여부 확인
        if not self.is_one_year_or_more(name, hire_date, target_date):
            # 1년 미만자는 소멸 없음
            conn.close()
            return 0
        
        # 1년 이상자: 입사기념일 기준으로 연차 생성, 다음년도 입사기념일까지 사용 안하면 소멸
        hire_month = hire_date.month
        hire_day = hire_date.day
        
        expired_total = 0
        
        # 입사일 이후의 모든 입사기념일 확인
        for year_offset in range(1, 50):  # 최대 50년까지 확인
            # 입사기념일 계산
            try:
                anniversary_date = datetime(hire_date.year + year_offset, hire_month, hire_day).date()
            except ValueError:
                anniversary_date = datetime(hire_date.year + year_offset, hire_month, hire_day - 1).date()
            
            # 다음년도 입사기념일 (소멸일)
            try:
                next_anniversary_date = datetime(anniversary_date.year + 1, hire_month, hire_day).date()
            except ValueError:
                next_anniversary_date = datetime(anniversary_date.year + 1, hire_month, hire_day - 1).date()
            
            # 소멸일이 지났는지 확인
            if target_date >= next_anniversary_date:
                # 기존 소멸 기록 확인
                cursor.execute("""
                    SELECT id FROM leave_expirations
                    WHERE employee_id = ? AND leave_type = '연차' AND expiration_date = ?
                    """, (employee_id, next_anniversary_date))
                existing = cursor.fetchone()
                
                if not existing:
                    # 입사기념일 기준으로 생성된 연차 계산
                    annual_leave_generated = self.calculate_annual_leave(hire_date, anniversary_date)
                    
                    # 입사기념일부터 다음년도 입사기념일 직전까지 사용한 연차 계산
                    cursor.execute("""
                        SELECT SUM(leave_amount) as total_used
                        FROM leave_records
                        WHERE employee_id = ? AND leave_type = '연차'
                        AND leave_date >= ? AND leave_date < ?
                        """, (employee_id, anniversary_date, next_anniversary_date))
                    result = cursor.fetchone()
                    total_used = result[0] if result[0] else 0
                    
                    expired_amount = annual_leave_generated - total_used
                    
                    if expired_amount > 0:
                        cursor.execute("""
                            INSERT INTO leave_expirations 
                            (employee_id, leave_type, expired_amount, expiration_date, year)
                            VALUES (?, ?, ?, ?, ?)
                            """, (employee_id, '연차', expired_amount, next_anniversary_date, target_date.year))
                        expired_total += expired_amount
            else:
                # 아직 소멸일이 지나지 않았으면 중단
                break
        
        conn.commit()
        conn.close()
        return expired_total


class AttendanceCalculator:
    """출퇴근 계산 클래스"""
    
    def __init__(self, db_manager):
        self.db = db_manager
    
    def process_attendance_record(self, employee_id, work_date, arrival_time, departure_time, leave_type=None, remarks=None, conn=None):
        """출퇴근 기록 처리 및 계산
        
        Args:
            employee_id: 직원 ID
            work_date: 근무일
            arrival_time: 출근 시간
            departure_time: 퇴근 시간
            leave_type: 휴가 유형
            remarks: 비고
            conn: 데이터베이스 연결 (None이면 새로 생성)
        """
        try:
            if isinstance(work_date, str):
                work_date = datetime.strptime(work_date, "%Y-%m-%d").date()
            if isinstance(arrival_time, str):
                # HH:MM:SS 또는 HH:MM 형식 처리
                try:
                    if len(arrival_time) >= 8:
                        arrival_time = datetime.strptime(arrival_time, "%H:%M:%S").time()
                    else:
                        arrival_time = datetime.strptime(arrival_time, "%H:%M").time()
                except:
                    arrival_time = None
            if isinstance(departure_time, str):
                # HH:MM:SS 또는 HH:MM 형식 처리
                try:
                    if len(departure_time) >= 8:
                        departure_time = datetime.strptime(departure_time, "%H:%M:%S").time()
                    else:
                        departure_time = datetime.strptime(departure_time, "%H:%M").time()
                except:
                    departure_time = None
            
            early_arrival = 0
            late_arrival = 0
            late_departure = 0
            
            if arrival_time is not None:
                early_arrival = 1 if arrival_time < datetime.strptime("08:00", "%H:%M").time() else 0
                late_arrival = 1 if arrival_time > datetime.strptime("09:00", "%H:%M").time() else 0
            
            if departure_time is not None:
                late_departure = 1 if departure_time >= datetime.strptime("20:00", "%H:%M").time() else 0
            
            # connection이 제공되지 않으면 새로 생성
            should_close = False
            if conn is None:
                conn = self.db.get_connection()
                should_close = True
            
            cursor = conn.cursor()
            
            try:
                arrival_time_str = arrival_time.strftime("%H:%M:%S") if arrival_time else None
                departure_time_str = departure_time.strftime("%H:%M:%S") if departure_time else None
                
                cursor.execute("""
                    INSERT OR REPLACE INTO attendance_records
                    (employee_id, work_date, arrival_time, departure_time, 
                     early_arrival, late_arrival, late_departure, leave_type, remarks)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (employee_id, work_date, arrival_time_str, departure_time_str, 
                      early_arrival, late_arrival, late_departure, leave_type, remarks))
                
                conn.commit()
            except Exception as e:
                conn.rollback()
                raise Exception(f"데이터베이스 저장 중 오류: {str(e)}")
            finally:
                if should_close:
                    conn.close()
        except Exception as e:
            # 예외를 다시 발생시켜 호출자가 처리할 수 있도록 함
            raise


# GUI 클래스들 - PySide6 + QTableWidget 사용

if not PYSIDE6_AVAILABLE:
    # PySide6가 없을 때 더미 클래스 정의 (NameError 방지)
    class QWidget:
        pass
    class QMainWindow:
        pass
    class QVBoxLayout:
        pass
    class QHBoxLayout:
        pass
    class QPushButton:
        pass
    class QTableWidget:
        pass
    class QTableWidgetItem:
        pass
    class QLabel:
        pass
    class QLineEdit:
        pass
    class QMessageBox:
        pass
    class QDialog:
        pass
    class QComboBox:
        pass
    class QDateEdit:
        pass
    class QDialogButtonBox:
        pass
    class QTextEdit:
        pass
    class QFileDialog:
        pass
    class QAbstractItemView:
        pass
    class QApplication:
        pass
    class QTabWidget:
        pass
    class QStatusBar:
        pass
    class QColor:
        pass
    class Qt:
        pass
    class QDate:
        pass

if PYSIDE6_AVAILABLE:
    class LeaveManagementGUI(QWidget):
        """연월차 관리 GUI"""
        
        def __init__(self, parent, db_manager, leave_calculator, employee_gui=None):
            super().__init__(parent)
            self.db = db_manager
            self.calculator = leave_calculator
            self.employee_gui = employee_gui  # 재직인원 탭 참조
            
            layout = QVBoxLayout(self)
            
            # 버튼 프레임
            button_layout = QHBoxLayout()
            button_layout.addWidget(QPushButton("엑셀 업로드", clicked=self.upload_excel))
            # 연월차 + 출퇴근을 한 파일로 묶어서 다운로드
            button_layout.addWidget(QPushButton("엑셀 다운로드", clicked=self.download_combined_excel))
            button_layout.addWidget(QPushButton("연차 사용 등록", clicked=self.register_leave))
            button_layout.addWidget(QPushButton("소멸 내역 조회", clicked=self.view_expirations))
            button_layout.addWidget(QPushButton("새로고침", clicked=self.refresh_data))
            button_layout.addStretch()
            # 재직인원 수 표시 레이블
            self.employee_count_label = QLabel("재직인원: 0명")
            self.employee_count_label.setStyleSheet("font-weight: bold; color: #0066CC;")
            button_layout.addWidget(self.employee_count_label)
            # TODAY 표시 레이블
            today_str = datetime.now().strftime("%Y-%m-%d")
            self.today_label = QLabel(f"TODAY: {today_str}")
            self.today_label.setStyleSheet("font-weight: bold; color: #FF6600; margin-left: 15px;")
            button_layout.addWidget(self.today_label)
            layout.addLayout(button_layout)
            
            # 조회 기간 선택
            year_layout = QHBoxLayout()
            year_layout.addWidget(QLabel("조회 기간:"))
            
            # 년도 선택 드롭다운
            self.year_combo = QComboBox()
            current_year = datetime.now().year
            max_year = max(current_year, 2028)
            for year in range(max_year, 2019, -1):
                self.year_combo.addItem(str(year), year)
            self.year_combo.setCurrentText(str(current_year))
            self.year_combo.setMaxVisibleItems(20)  # 드롭다운 열었을 때 모든 년도가 보이도록 설정
            self.year_combo.currentIndexChanged.connect(self.refresh_data)  # 년도 변경 시 자동 새로고침
            
            year_layout.addWidget(self.year_combo)
            year_layout.addWidget(QLabel("년"))
            year_layout.addStretch()
            layout.addLayout(year_layout)
            
            # 퇴사자 표시 옵션 (재직인원 탭의 체크박스와 동기화)
            option_layout = QHBoxLayout()
            self.show_inactive_checkbox = QCheckBox("퇴사자 표시")
            if self.employee_gui:
                # 재직인원 탭의 체크박스 상태와 동기화
                self.show_inactive_checkbox.setChecked(self.employee_gui.show_inactive_checkbox.isChecked())
                # 재직인원 탭의 체크박스 상태 변경 시 이 탭의 체크박스도 업데이트 (비동기로 처리)
                def sync_from_employee():
                    from PySide6.QtCore import QTimer
                    self.show_inactive_checkbox.blockSignals(True)
                    self.show_inactive_checkbox.setChecked(self.employee_gui.show_inactive_checkbox.isChecked())
                    self.show_inactive_checkbox.blockSignals(False)
                    # 비동기로 새로고침하여 데이터베이스 충돌 방지
                    QTimer.singleShot(50, lambda: self.refresh_data())
                self.employee_gui.show_inactive_checkbox.stateChanged.connect(sync_from_employee)
                # 이 탭의 체크박스 상태 변경 시 재직인원 탭의 체크박스도 업데이트
                def sync_to_employee():
                    from PySide6.QtCore import QTimer
                    # 재직인원 탭의 체크박스 업데이트 (무한 루프 방지)
                    self.employee_gui.show_inactive_checkbox.blockSignals(True)
                    self.employee_gui.show_inactive_checkbox.setChecked(self.show_inactive_checkbox.isChecked())
                    self.employee_gui.show_inactive_checkbox.blockSignals(False)
                    # 재직인원 탭의 체크박스 변경 핸들러가 다른 탭들도 업데이트하므로 여기서는 refresh만 호출 (비동기)
                    QTimer.singleShot(50, lambda: self.employee_gui.refresh_data() if self.employee_gui else None)
                    # 다른 탭의 체크박스도 동기화 (비동기)
                    if self.employee_gui.leave_gui and self.employee_gui.leave_gui != self:
                        self.employee_gui.leave_gui.show_inactive_checkbox.blockSignals(True)
                        self.employee_gui.leave_gui.show_inactive_checkbox.setChecked(self.show_inactive_checkbox.isChecked())
                        self.employee_gui.leave_gui.show_inactive_checkbox.blockSignals(False)
                        QTimer.singleShot(100, lambda: self.employee_gui.leave_gui.refresh_data() if self.employee_gui and self.employee_gui.leave_gui else None)
                    if self.employee_gui.attendance_gui:
                        self.employee_gui.attendance_gui.show_inactive_checkbox.blockSignals(True)
                        self.employee_gui.attendance_gui.show_inactive_checkbox.setChecked(self.show_inactive_checkbox.isChecked())
                        self.employee_gui.attendance_gui.show_inactive_checkbox.blockSignals(False)
                        QTimer.singleShot(150, lambda: self.employee_gui.attendance_gui.refresh_data() if self.employee_gui and self.employee_gui.attendance_gui else None)
                    # 이 탭도 비동기로 새로고침
                    QTimer.singleShot(0, lambda: self.refresh_data())
                self.show_inactive_checkbox.stateChanged.connect(sync_to_employee)
            else:
                # 재직인원 탭이 없으면 독립적으로 동작
                self.show_inactive_checkbox.setChecked(False)
                self.show_inactive_checkbox.stateChanged.connect(self.refresh_data)
            option_layout.addWidget(self.show_inactive_checkbox)
            option_layout.addStretch()
            layout.addLayout(option_layout)
            
            # 테이블 - EditableTableWidget 사용 (복사/붙여넣기 지원)
            self.table = EditableTableWidget()
            self.table.set_parent_gui(self)  # 부모 GUI 참조 설정
            self.table.setColumnCount(21)
            columns = ["부서", "직급", "이름", "입사일", "2024년 남은연차",
                       "1월", "2월", "3월", "4월", "5월", "6월",
                       "7월", "8월", "9월", "10월", "11월", "12월",
                       "2025년 사용연차", "연차발생수", "잔여수", "소멸내역"]
            self.table.setHorizontalHeaderLabels(columns)
            self.table.horizontalHeader().setStretchLastSection(True)
            self.table.setSelectionBehavior(QAbstractItemView.SelectItems)  # 셀 단위 선택
            self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)  # CTRL/SHIFT로 다중 선택
            # 모든 셀 편집 가능 (더블클릭)
            self.table.setEditTriggers(QAbstractItemView.DoubleClicked)
            
            # 컬럼 넓이 설정
            self.table.setColumnWidth(0, 80)   # 부서
            self.table.setColumnWidth(1, 60)   # 직급
            self.table.setColumnWidth(2, 70)   # 이름
            self.table.setColumnWidth(3, 100)  # 입사일
            self.table.setColumnWidth(4, 120)  # 2024년 남은연차
            # 월 컬럼(5번째부터 16번째까지: 1월~12월) 넓이 작게 설정
            for i in range(5, 17):  # 1월~12월
                self.table.setColumnWidth(i, 50)  # 월별 연차 사용량
            self.table.setColumnWidth(17, 120)  # 2025년 사용연차
            self.table.setColumnWidth(18, 100)  # 연차발생수
            self.table.setColumnWidth(19, 80)   # 잔여수
            # 소멸내역(20번)은 setStretchLastSection(True)로 자동 조정
            
            # 셀 편집 완료 시 이벤트 연결
            self.table.itemChanged.connect(self.on_cell_changed)
            
            # 테이블에 포커스 설정 (키보드 이벤트 처리를 위해)
            self.table.setFocusPolicy(Qt.StrongFocus)
            
            layout.addWidget(self.table)
            
            self._is_refreshing = False  # 데이터 새로고침 중 플래그
            
            self.refresh_data()
        
        def refresh_data(self):
            """데이터 새로고침"""
            self._is_refreshing = True  # 새로고침 시작
            self.table.setRowCount(0)
            
            # 선택된 년도 가져오기
            selected_year = self.year_combo.currentData()
            if selected_year is None:
                selected_year = datetime.now().year
            
            # 조회 기간에 맞춰 헤더 업데이트
            prev_year = selected_year - 1
            prev_year_text = f"{prev_year}년 남은연차"
            current_year_text = f"{selected_year}년 사용연차"
            self.table.setHorizontalHeaderItem(4, QTableWidgetItem(prev_year_text))
            self.table.setHorizontalHeaderItem(17, QTableWidgetItem(current_year_text))
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # 퇴사일을 고려한 조회: 퇴사일이 있으면 해당 년도/월까지 표시
            cursor.execute("""
                SELECT id, department, position, name, hire_date,
                       COALESCE(display_order, 0) as display_order,
                       COALESCE(is_active, 1) as is_active,
                       resignation_date
                FROM employees
                ORDER BY
                    CASE department
                        WHEN '경영지원팀' THEN 1
                        WHEN '영업팀' THEN 2
                        WHEN '글로벌비즈니스팀' THEN 3
                        ELSE 999
                    END,
                    department,
                    CASE position
                        WHEN '이사' THEN 1
                        WHEN '팀장' THEN 2
                        WHEN '파트장' THEN 3
                        WHEN '과장' THEN 4
                        WHEN '대리' THEN 5
                        WHEN '프로' THEN 6
                        ELSE 999
                    END,
                    hire_date ASC
            """)
            all_employees = cursor.fetchall()
            
            # 퇴사자 표시 옵션 확인
            show_inactive = self.show_inactive_checkbox.isChecked()
            
            # 퇴사일을 고려한 필터링: 활성 직원 또는 퇴사일이 선택한 년도 이하인 경우만 표시
            employees = []
            for emp in all_employees:
                emp_id, dept, pos, name, hire_date, display_order, is_active, resignation_date = emp
                
                # resignation_date가 문자열인 경우 date 객체로 변환
                if resignation_date and isinstance(resignation_date, str):
                    try:
                        resignation_date = datetime.strptime(resignation_date, "%Y-%m-%d").date()
                    except:
                        resignation_date = None
                
                if is_active == 1:
                    # 활성 직원은 항상 표시
                    employees.append((emp_id, dept, pos, name, hire_date, display_order))
                elif show_inactive:
                    # 퇴사자 표시 체크박스가 체크되어 있으면 모든 퇴사자 표시
                    employees.append((emp_id, dept, pos, name, hire_date, display_order))
                # 체크박스가 OFF이면 퇴사자는 표시하지 않음 (기존 로직 제거)
            
            current_date = datetime(selected_year, 11, 1).date()
            
            current_department = None
            employee_row_number = 1  # 실제 직원 행 번호 카운터
            for emp_id, dept, pos, name, hire_date, display_order in employees:
                # 부서가 변경되면 구분자 추가
                if current_department != dept:
                    # 구분자 행 추가
                    separator_row = self.table.rowCount()
                    self.table.insertRow(separator_row)
                    
                    # 구분자 행의 행 번호를 공란으로 설정
                    self.table.setVerticalHeaderItem(separator_row, QTableWidgetItem(""))
                    
                    # 구분자 아이템 생성
                    separator_item = QTableWidgetItem(f"━━━ {dept} ━━━")
                    separator_item.setFlags(Qt.NoItemFlags)  # 선택 불가
                    separator_item.setBackground(QColor("#E0E0E0"))
                    separator_item.setFont(QFont("Arial", 10, QFont.Bold))
                    self.table.setItem(separator_row, 0, separator_item)
                    # 모든 컬럼에 걸쳐 병합
                    self.table.setSpan(separator_row, 0, 1, self.table.columnCount())
                    
                    current_department = dept
                hire_date_obj = datetime.strptime(hire_date, "%Y-%m-%d").date()
                
                self.calculator.check_monthly_leave_expiration(emp_id, hire_date_obj, current_date)
                self.calculator.check_annual_leave_expiration(emp_id, current_date)
                
                # 입사일 기준 연차 생성일 계산
                # 예: 2022-08-01 입사 → 2023-08-01에 첫 연차 생성, 2025-08-01에 만 3년 연차 생성
                hire_year = hire_date_obj.year
                hire_month = hire_date_obj.month
                hire_day = hire_date_obj.day
                
                # 선택된 년도의 입사기념일 계산
                try:
                    anniversary_date = datetime(selected_year, hire_month, hire_day).date()
                except ValueError:
                    # 2월 29일 같은 경우 처리
                    anniversary_date = datetime(selected_year, hire_month, hire_day - 1).date()
                
                # 이전 입사기념일 계산 (선택된 년도 이전의 마지막 입사기념일)
                prev_anniversary_year = selected_year - 1
                try:
                    prev_anniversary_date = datetime(prev_anniversary_year, hire_month, hire_day).date()
                except ValueError:
                    prev_anniversary_date = datetime(prev_anniversary_year, hire_month, hire_day - 1).date()
                
                # 입사일부터 입사기념일까지의 일수 계산
                # 입사일과 입사기념일이 같은 경우(같은 년도 입사) 0일이 되므로, 다음 년도 입사기념일까지의 일수로 계산
                if anniversary_date <= hire_date_obj:
                    # 입사기념일이 입사일보다 이전이거나 같으면 다음 년도 입사기념일 사용
                    next_anniversary_date = datetime(selected_year + 1, hire_month, hire_day).date()
                    try:
                        next_anniversary_date = datetime(selected_year + 1, hire_month, hire_day).date()
                    except ValueError:
                        next_anniversary_date = datetime(selected_year + 1, hire_month, hire_day - 1).date()
                    days_from_hire_to_anniversary = (next_anniversary_date - hire_date_obj).days
                else:
                    days_from_hire_to_anniversary = (anniversary_date - hire_date_obj).days
                
                # 이전 년도 남은 연차 계산 (예: 2025년 선택 시 2024년 남은 연차)
                # 입사일이 선택된 년도인 근로자는 해당 사항 없으니 0으로 표기
                # 그 이전 근로자는 입사일에 맞춰서 계산
                if hire_year >= selected_year:
                    # 선택된 년도 입사자는 이전 년도 남은연차가 없음
                    remaining_prev_year = 0.0
                else:
                    # 이전 년도 이전 입사자는 입사일에 맞춰서 계산
                    # 이전 입사기념일부터 이번 입사기념일 직전까지 사용한 연차 계산
                    cursor.execute("""
                        SELECT work_date, leave_type, remarks
                        FROM attendance_records
                        WHERE employee_id = ? AND work_date >= ? AND work_date < ?
                        AND leave_type IN ('연차', '반차', '휴가')
                    """, (emp_id, prev_anniversary_date, anniversary_date))
                    prev_records = cursor.fetchall()
                    
                    used_before_anniversary = 0.0
                    반차_날짜_set_prev = set()
                    
                    for work_date, leave_type, remarks in prev_records:
                        if leave_type == '연차' or leave_type == '휴가':
                            used_before_anniversary += 1.0
                        elif leave_type == '반차':
                            if work_date not in 반차_날짜_set_prev:
                                used_before_anniversary += 0.5
                                반차_날짜_set_prev.add(work_date)
                    
                    # 이전 년도(prev_year) 연말 시점의 잔여수 계산
                    # 이전 년도 연말 시점의 잔여수 = (이전 년도 연차 발생 수 - 이전 년도 사용연차) + (이전 년도 이전의 최종 남은 연차)
                    prev_year = selected_year - 1
                    prev_year_end = datetime(prev_year, 12, 31).date()
                    
                    # 이전 년도 연말 시점까지의 연차 발생 수 계산 (2025년 조회 시와 동일한 로직)
                    # 입사기념일 계산
                    prev_year_anniversary = datetime(prev_year, hire_month, hire_day).date()
                    try:
                        prev_year_anniversary = datetime(prev_year, hire_month, hire_day).date()
                    except ValueError:
                        prev_year_anniversary = datetime(prev_year, hire_month, hire_day - 1).date()
                    
                    if prev_year_anniversary <= hire_date_obj:
                        try:
                            prev_year_anniversary = datetime(prev_year + 1, hire_month, hire_day).date()
                        except ValueError:
                            prev_year_anniversary = datetime(prev_year + 1, hire_month, hire_day - 1).date()
                    
                    days_from_hire_to_prev_year_end = (prev_year_end - hire_date_obj).days
                    
                    if days_from_hire_to_prev_year_end < 0:
                        annual_prev_anniversary = 0
                    elif days_from_hire_to_prev_year_end < 365:
                        # 1년 미만: 입사일부터 이전 년도 연말까지 월차 계산
                        months_passed_prev = (prev_year_end.year - hire_date_obj.year) * 12 + (prev_year_end.month - hire_date_obj.month)
                        if prev_year_end.day < hire_date_obj.day:
                            months_passed_prev -= 1
                        annual_prev_anniversary = min(max(months_passed_prev, 0), 11)
                    elif prev_year_anniversary <= prev_year_end:
                        # 입사 1년 경과 시: 입사기념일 기준으로 연차 계산
                        annual_prev_anniversary = self.calculator.calculate_annual_leave(hire_date_obj, prev_year_anniversary)
                    else:
                        # 입사기념일이 이전 년도 이후: 연도 말일 기준으로 계산
                        annual_prev_anniversary = self.calculator.calculate_annual_leave(hire_date_obj, prev_year_end)
                    
                    # 이전 년도에 사용한 연차 계산 (이전 년도 전체, 입사기념일 기준)
                    # 입사기념일이 이전 년도 내에 있으면 입사기념일 이후만, 아니면 연도 전체
                    if prev_year_anniversary <= prev_year_end:
                        # 입사기념일이 이전 년도 내에 있음
                        if hire_date_obj.year == prev_year:
                            usage_start_date_prev = hire_date_obj
                        else:
                            usage_start_date_prev = prev_year_anniversary
                    else:
                        # 입사기념일이 이전 년도 이후에 있음
                        if hire_date_obj.year == prev_year:
                            usage_start_date_prev = hire_date_obj
                        else:
                            usage_start_date_prev = datetime(prev_year, 1, 1).date()
                    
                    cursor.execute("""
                        SELECT work_date, leave_type, remarks
                        FROM attendance_records
                        WHERE employee_id = ? AND work_date >= ? AND work_date < ?
                        AND (leave_type IN ('연차', '반차', '휴가') 
                             OR remarks IN ('반차_출근', '반차_퇴근'))
                    """, (emp_id, 
                          usage_start_date_prev,
                          datetime(prev_year + 1, 1, 1).date()))
                    prev_year_records = cursor.fetchall()
                    
                    used_prev_year = 0.0
                    반차_날짜_set_prev_year = set()
                    
                    for work_date, leave_type, remarks in prev_year_records:
                        is_half_day = (leave_type == '반차' or remarks == '반차_출근' or remarks == '반차_퇴근')
                        if leave_type == '연차' or leave_type == '휴가':
                            used_prev_year += 1.0
                        elif is_half_day:
                            if work_date not in 반차_날짜_set_prev_year:
                                used_prev_year += 0.5
                                반차_날짜_set_prev_year.add(work_date)
                    
                    # 이전 년도 이전의 최종 남은 연차 계산 (소멸 차감 포함)
                    # 이전 년도 이전의 남은 연차 계산
                    prev_prev_year = prev_year - 1
                    if prev_prev_year >= hire_date_obj.year:
                        # 이전 년도 이전의 연말 시점까지의 연차 발생 수 계산
                        prev_prev_year_end = datetime(prev_prev_year, 12, 31).date()
                        days_from_hire_to_prev_prev_year_end = (prev_prev_year_end - hire_date_obj).days
                        
                        if days_from_hire_to_prev_prev_year_end < 0:
                            annual_prev_prev_anniversary = 0
                        elif days_from_hire_to_prev_prev_year_end < 365:
                            months_passed_prev_prev = (prev_prev_year_end.year - hire_date_obj.year) * 12 + (prev_prev_year_end.month - hire_date_obj.month)
                            if prev_prev_year_end.day < hire_date_obj.day:
                                months_passed_prev_prev -= 1
                            annual_prev_prev_anniversary = min(max(months_passed_prev_prev, 0), 11)
                        else:
                            annual_prev_prev_anniversary = self.calculator.calculate_annual_leave(hire_date_obj, prev_prev_year_end)
                        
                        # 이전 년도 이전에 사용한 연차 계산
                        cursor.execute("""
                            SELECT work_date, leave_type, remarks
                            FROM attendance_records
                            WHERE employee_id = ? AND work_date >= ? AND work_date < ?
                            AND leave_type IN ('연차', '반차', '휴가')
                        """, (emp_id, 
                              datetime(prev_prev_year, 1, 1).date(),
                              datetime(prev_prev_year + 1, 1, 1).date()))
                        prev_prev_year_records = cursor.fetchall()
                        
                        used_prev_prev_year = 0.0
                        반차_날짜_set_prev_prev_year = set()
                        
                        for work_date, leave_type, remarks in prev_prev_year_records:
                            if leave_type == '연차' or leave_type == '휴가':
                                used_prev_prev_year += 1.0
                            elif leave_type == '반차':
                                if work_date not in 반차_날짜_set_prev_prev_year:
                                    used_prev_prev_year += 0.5
                                    반차_날짜_set_prev_prev_year.add(work_date)
                        
                        remaining_prev_prev_year = annual_prev_prev_anniversary - used_prev_prev_year
                    else:
                        remaining_prev_prev_year = 0.0
                    
                    # 이전 년도 연말 시점 이전에 소멸된 연차 조회
                    cursor.execute("""
                        SELECT SUM(expired_amount) as total_expired
                        FROM leave_expirations
                        WHERE employee_id = ? AND leave_type = '연차' AND expiration_date <= ?
                    """, (emp_id, prev_year_end))
                    expired_result = cursor.fetchone()
                    total_expired_prev = expired_result[0] if expired_result[0] else 0
                    
                    # 이전 년도 이전의 최종 남은 연차 (소멸 차감 적용)
                    remaining_prev_prev_year_float = float(remaining_prev_prev_year) if remaining_prev_prev_year is not None else 0.0
                    total_expired_prev_float = float(total_expired_prev) if total_expired_prev is not None else 0.0
                    remaining_prev_prev_year_final = max(0.0, remaining_prev_prev_year_float - total_expired_prev_float)
                    
                    # 이전 년도 연말 시점의 잔여수 계산
                    # 1년 이상 재직인원이고 2026년 조회 시: 저장된 2025년 잔여수 사용
                    if selected_year >= 2026 and self.calculator.is_one_year_or_more(name, hire_date_obj):
                        # 저장된 이전 년도 잔여수 조회
                        cursor.execute("""
                            SELECT remaining_amount FROM leave_remaining_by_year
                            WHERE employee_id = ? AND year = ?
                        """, (emp_id, prev_year))
                        stored_remaining = cursor.fetchone()
                        if stored_remaining:
                            # 저장된 값이 있으면 그대로 사용
                            remaining_prev_year = float(stored_remaining[0])
                        else:
                            # 저장된 값이 없으면 계산
                            remaining_prev_year = (annual_prev_anniversary - used_prev_year) + remaining_prev_prev_year_final
                    else:
                        # 그 외의 경우: 계산된 값 사용
                        remaining_prev_year = (annual_prev_anniversary - used_prev_year) + remaining_prev_prev_year_final
                
                # 월별 사용량 (연차, 반차, 휴가 포함) - 선택된 년도
                # 출퇴근 관리대장의 attendance_records에서 직접 계산하여 반영
                monthly_usage = []
                for month in range(1, 13):
                    month_start = datetime(selected_year, month, 1).date()
                    month_end = datetime(selected_year, month + 1, 1).date() if month < 12 else datetime(selected_year + 1, 1, 1).date()
                    
                    # 입사기념일 필터링 로직:
                    # - 입사기념일이 해당 월보다 이전이면 해당 월 전체 계산
                    # - 입사기념일이 해당 월 내에 있으면 입사기념일 이후만 계산
                    # - 입사기념일이 해당 월보다 이후이면 해당 월 전체 계산 (이전 년도에 생성된 연차 사용 가능)
                    if anniversary_date < month_start:
                        # 입사기념일이 해당 월보다 이전이면 해당 월 전체 계산
                        query_start = month_start
                    elif anniversary_date < month_end:
                        # 입사기념일이 해당 월 내에 있으면 입사기념일 이후만 계산
                        query_start = anniversary_date
                    else:
                        # 입사기념일이 해당 월보다 이후이면 해당 월 전체 계산
                        query_start = month_start
                    
                    # attendance_records에서 work_date를 가져와서 정확히 계산
                    cursor.execute("""
                        SELECT work_date, leave_type, remarks
                        FROM attendance_records
                        WHERE employee_id = ? AND work_date >= ? AND work_date < ?
                        AND (leave_type IN ('연차', '반차', '휴가') 
                             OR remarks IN ('반차_출근', '반차_퇴근'))
                    """, (emp_id, query_start, month_end))
                    month_records = cursor.fetchall()
                    
                    month_leave_amount = 0.0
                    반차_날짜_set = set()  # 반차 중복 방지
                    
                    for work_date, leave_type, remarks in month_records:
                        # remarks가 반차_출근 또는 반차_퇴근인 경우도 반차로 처리
                        is_half_day = (leave_type == '반차' or remarks == '반차_출근' or remarks == '반차_퇴근')
                        
                        if leave_type == '연차' or leave_type == '휴가':
                            month_leave_amount += 1.0
                        elif is_half_day:
                            # 반차는 하루에 0.5만 (같은 날짜에 여러 번 있어도 한 번만)
                            if work_date not in 반차_날짜_set:
                                month_leave_amount += 0.5
                                반차_날짜_set.add(work_date)
                    
                    monthly_usage.append(month_leave_amount)
                
                # 선택된 년도 입사기념일 이후 사용량 (연차, 반차, 휴가 포함)
                # 출퇴근 관리대장의 attendance_records에서 직접 계산
                # 입사기념일이 선택된 년도 내에 있으면 입사기념일 이후만, 아니면 연도 전체
                # 단, 입사일이 선택된 년도 내에 있고 입사기념일이 선택된 년도 내에 있으면
                # 입사일부터 계산 (입사일 이후 발생한 연차 사용 가능)
                if anniversary_date <= datetime(selected_year, 12, 31).date():
                    # 입사기념일이 선택된 년도 내에 있음
                    # 입사일이 선택된 년도 내에 있으면 입사일부터, 아니면 입사기념일부터
                    if hire_date_obj.year == selected_year:
                        # 입사일이 선택된 년도 내에 있으면 입사일부터 계산
                        usage_start_date = hire_date_obj
                    else:
                        # 입사일이 선택된 년도 이전이면 입사기념일부터 계산
                        # 입사기념일 당일부터 포함하여 계산 (>= 대신 > 사용 시 당일 제외됨)
                        usage_start_date = anniversary_date
                else:
                    # 입사기념일이 선택된 년도 이후에 있음 (아직 생성 안 됨)
                    # 입사일이 선택된 년도 내에 있으면 입사일부터, 아니면 연도 시작부터
                    if hire_date_obj.year == selected_year:
                        usage_start_date = hire_date_obj
                    else:
                        usage_start_date = datetime(selected_year, 1, 1).date()
                
                cursor.execute("""
                    SELECT work_date, leave_type, remarks
                    FROM attendance_records
                    WHERE employee_id = ? AND work_date >= ? AND work_date < ?
                    AND (leave_type IN ('연차', '반차', '휴가') 
                         OR remarks IN ('반차_출근', '반차_퇴근'))
                """, (emp_id,
                      usage_start_date,
                      datetime(selected_year + 1, 1, 1).date()))
                year_records = cursor.fetchall()
                
                used_current_year = 0.0
                반차_날짜_set = set()  # 반차 중복 방지
                
                for work_date, leave_type, remarks in year_records:
                    # remarks가 반차_출근 또는 반차_퇴근인 경우도 반차로 처리
                    is_half_day = (leave_type == '반차' or remarks == '반차_출근' or remarks == '반차_퇴근')
                    
                    if leave_type == '연차' or leave_type == '휴가':
                        used_current_year += 1.0
                    elif is_half_day:
                        # 반차는 하루에 0.5만 (같은 날짜에 여러 번 있어도 한 번만)
                        if work_date not in 반차_날짜_set:
                            used_current_year += 0.5
                            반차_날짜_set.add(work_date)
                            # 디버깅: 반차 추가 확인
                            if name == "김미라":
                                print(f"DEBUG {name} (refresh_data): 반차 추가 - work_date={work_date}, used_current_year={used_current_year}")
                
                # 수동 입력된 월별 연차 사용량도 사용연차에 포함
                # 월별 컬럼은 5번(1월)부터 16번(12월)까지
                # 입사기념일 이후의 월만 포함해야 함
                cursor.execute("""
                    SELECT column_index, manual_value
                    FROM leave_manual_values
                    WHERE employee_id = ? AND year = ? AND column_index >= 5 AND column_index <= 16
                """, (emp_id, selected_year))
                manual_monthly_values = cursor.fetchall()
                
                manual_used_current_year = 0.0
                for col_idx, manual_val in manual_monthly_values:
                    if manual_val:
                        try:
                            month_value = float(manual_val)
                            # 컬럼 인덱스 5=1월, 6=2월, ..., 16=12월
                            month = col_idx - 4
                            month_start = datetime(selected_year, month, 1).date()
                            # 월의 마지막 날짜 계산
                            if month == 12:
                                month_end = datetime(selected_year + 1, 1, 1).date()
                            else:
                                month_end = datetime(selected_year, month + 1, 1).date()
                            
                            # 입사기념일 이후의 월만 포함
                            # 입사기념일이 속한 월은 제외하고, 그 다음 월부터 포함
                            if anniversary_date <= datetime(selected_year, 12, 31).date():
                                # 입사기념일이 선택된 년도 내에 있으면 입사기념일이 속한 월을 제외하고 그 다음 월부터 포함
                                # 입사기념일이 해당 월 내에 있으면 해당 월 수동 입력은 사용연차 계산에서 제외 (경고 로그 출력)
                                if anniversary_date >= month_start and anniversary_date < month_end:
                                    try:
                                        print(f"WARNING {name}: 입사기념일이 속한 {month}월의 수동 입력 값({month_value})은 2025년 사용연차 계산에서 제외됩니다. 출퇴근 관리대장 기록만 반영됩니다.")
                                    except Exception:
                                        pass
                                    continue
                                elif month_start > anniversary_date:
                                    # 입사기념일 이후의 월만 포함
                                    manual_used_current_year += month_value
                            else:
                                # 입사기념일이 선택된 년도 이후에 있으면 모든 월 포함
                                manual_used_current_year += month_value
                        except (ValueError, TypeError):
                            pass
                
                # attendance_records에서 계산한 값과 수동 입력 월별 값을 합산
                used_current_year += manual_used_current_year
                
                # 디버깅: 사용연차 계산 결과 확인 (refresh_data)
                if name == "김미라" or name == "전금희" or name == "강지승":
                    print(f"DEBUG {name} (refresh_data): 입사일={hire_date_obj}, 입사기념일={anniversary_date}, usage_start_date={usage_start_date}, year_records 개수={len(year_records)}, 반차_날짜_set={반차_날짜_set}")
                    print(f"DEBUG {name} (refresh_data): attendance_records에서 계산={used_current_year - manual_used_current_year}, 수동 입력 월별={manual_used_current_year}, 합계={used_current_year}")
                    print(f"DEBUG {name} (refresh_data): 조회 기간 = {usage_start_date} ~ {datetime(selected_year + 1, 1, 1).date()}")
                    if manual_monthly_values:
                        print(f"DEBUG {name} (refresh_data): 수동 입력 월별 값 개수={len(manual_monthly_values)}")
                        for col_idx, manual_val in manual_monthly_values:
                            if manual_val:
                                try:
                                    month_value = float(manual_val)
                                    month = col_idx - 4
                                    month_start = datetime(selected_year, month, 1).date()
                                    print(f"DEBUG {name} (refresh_data): 수동 입력 - {month}월: {month_value}, month_start={month_start}, anniversary_date={anniversary_date}")
                                except (ValueError, TypeError):
                                    pass
                    for work_date, leave_type, remarks in year_records:
                        is_half_day_check = (leave_type == '반차' or remarks == '반차_출근' or remarks == '반차_퇴근')
                        print(f"DEBUG {name} (refresh_data): work_date={work_date}, leave_type={leave_type}, remarks={remarks}, is_half_day={is_half_day_check}")
                
                # 연차 발생 수 계산 (입사기념일 기준)
                # - 입사 1년 미만: 입사일 기준으로 1개월 만근 시 연차 1개 발생 (최대 11개)
                # - 입사 1년 경과 시: 입사일 기준으로 연차 15개 부여
                # - 근속연수 증가 시: 만3년차 16일, 만5년차 17일, 만7년차 18일, 만25년차 25일 (최대)
                # 입사일부터 입사기념일까지의 일수 계산
                # 입사일과 입사기념일이 같은 경우(같은 년도 입사) 다음 년도 입사기념일까지의 일수로 계산
                if anniversary_date <= hire_date_obj:
                    # 입사기념일이 입사일보다 이전이거나 같으면 다음 년도 입사기념일 사용
                    try:
                        next_anniversary_date = datetime(selected_year + 1, hire_month, hire_day).date()
                    except ValueError:
                        next_anniversary_date = datetime(selected_year + 1, hire_month, hire_day - 1).date()
                    days_from_hire_to_anniversary = (next_anniversary_date - hire_date_obj).days
                else:
                    days_from_hire_to_anniversary = (anniversary_date - hire_date_obj).days
                days_from_hire_to_year_end = (datetime(selected_year, 12, 31).date() - hire_date_obj).days
                
                # 현재 날짜 기준으로 계산 (실제 근속 개월 수 반영)
                current_date_for_calc = datetime.now().date()
                # 선택된 년도가 현재 년도보다 이후이면 선택된 년도의 말일을 사용
                if selected_year > current_date_for_calc.year:
                    current_date_for_calc = datetime(selected_year, 12, 31).date()
                # 선택된 년도가 현재 년도이면 현재 날짜를 사용
                elif selected_year == current_date_for_calc.year:
                    # 선택된 년도 내에서만 계산
                    pass
                else:
                    # 선택된 년도가 과거이면 해당 년도 말일 사용
                    current_date_for_calc = datetime(selected_year, 12, 31).date()
                
                # 입사일부터 계산 기준일까지의 일수 계산
                # 오늘 날짜를 기준으로 입사일에 맞춰서 계산
                today = datetime.now().date()
                # 선택된 년도가 현재 년도보다 이후이면 선택된 년도의 말일을 사용
                if selected_year > today.year:
                    calc_date = datetime(selected_year, 12, 31).date()
                # 선택된 년도가 현재 년도이면 오늘 날짜를 사용
                elif selected_year == today.year:
                    calc_date = today
                else:
                    # 선택된 년도가 과거이면 해당 년도 말일 사용
                    calc_date = datetime(selected_year, 12, 31).date()
                
                # 입사일부터 계산일까지의 일수
                days_from_hire_to_current = (calc_date - hire_date_obj).days
                
                # TODAY 기준으로 1년 이상/미만 구분
                is_one_year_or_more = self.calculator.is_one_year_or_more(name, hire_date_obj, calc_date)
                
                # 입사 1년 미만 여부 확인
                if not is_one_year_or_more:
                    # 입사 1년 미만: 입사기념일 기준으로 1개월 만근 시 연차 1개 발생
                    # 입사기념일부터 계산 기준일까지의 개월 수 계산
                    # 입사기념일이 지나면 해당 월의 연차가 생성됨
                    # 예: 2025년 10월 13일 입사 → 2025년 11월 13일이 되면(또는 지나면) 연차 1개 생성
                    if calc_date < hire_date_obj:
                        # 입사일이 아직 미래인 경우
                        leave_generated = 0
                    else:
                        # 입사기념일 기준으로 개월 수 계산
                        # 입사일의 월/일을 기준으로 매월 같은 날짜가 지나면 연차 1개 발생
                        months_passed = (calc_date.year - hire_date_obj.year) * 12 + (calc_date.month - hire_date_obj.month)
                        # 입사일의 날짜가 아직 지나지 않았으면 한 달을 뺌
                        if calc_date.day < hire_date_obj.day:
                            months_passed -= 1
                        # 입사일과 같은 날짜이거나 지났으면 해당 월 연차 생성됨
                        # 최대 11개 (1년 전까지)
                        leave_generated = min(max(months_passed, 0), 11)
                elif is_one_year_or_more:
                    # 입사 1년 이상: 입사기념일 기준으로 연차 계산
                    # 입사기념일 기준으로 근속연수 책정하여 연차 생성
                    # 입사기념일이 선택된 년도 내에 있으면 입사기념일 기준으로 계산
                    if anniversary_date <= datetime(selected_year, 12, 31).date():
                        # 입사기념일이 선택된 년도 내에 있음
                        # 2025년 조회 시: 입사기념일 기준으로 연차 생성 (TODAY 확인 없음)
                        # 2026년 이상 조회 시: TODAY 기준으로 입사기념일이 지났을 때만 연차 생성
                        if selected_year <= 2025:
                            # 2025년 이하 조회 시: 입사기념일 기준으로 연차 생성
                            leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, anniversary_date)
                        else:
                            # 2026년 이상 조회 시: TODAY 기준으로 입사기념일이 지났을 때만 연차 생성
                            today = datetime.now().date()
                            if anniversary_date <= today:
                                # 입사기념일이 TODAY 이전이거나 같으면 연차 생성
                                leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, anniversary_date)
                            else:
                                # 입사기념일이 TODAY 이후이면 연차 생성 안 함
                                leave_generated = 0
                    else:
                        # 입사기념일이 선택된 년도 이후에 있음
                        # 이전 입사기념일 기준으로 계산
                        prev_anniversary_year = selected_year - 1
                        if prev_anniversary_year >= hire_date_obj.year:
                            try:
                                prev_anniversary = datetime(prev_anniversary_year, hire_month, hire_day).date()
                            except ValueError:
                                prev_anniversary = datetime(prev_anniversary_year, hire_month, hire_day - 1).date()
                            
                            if prev_anniversary > hire_date_obj:
                                # 이전 입사기념일이 입사일 이후이면 해당 입사기념일 기준으로 계산
                                today = datetime.now().date()
                                if prev_anniversary <= today:
                                    leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, prev_anniversary)
                                else:
                                    leave_generated = 0
                            else:
                                leave_generated = 0
                        else:
                            leave_generated = 0
                else:
                    # 입사기념일이 선택된 년도 이후에 있음
                    # 1년 미만자이거나 1년 이상자이지만 입사기념일이 아직 안 지난 경우
                    if not is_one_year_or_more:
                        # 입사 1년 미만: 입사기념일 기준으로 1개월 만근 시 연차 1개 발생
                        if calc_date < hire_date_obj:
                            leave_generated = 0
                        else:
                            months_passed = (calc_date.year - hire_date_obj.year) * 12 + (calc_date.month - hire_date_obj.month)
                            if calc_date.day < hire_date_obj.day:
                                months_passed -= 1
                            leave_generated = min(max(months_passed, 0), 11)
                    else:
                        # 1년 이상이지만 입사기념일이 선택된 년도 이후: 연도 말일 기준으로 계산
                        # 2026년 이상 조회 시: TODAY 기준으로 입사기념일이 지났는지 확인
                        if selected_year >= 2026:
                            today = datetime.now().date()
                            # 선택된 년도 내의 실제 입사기념일 찾기
                            # 입사일 이후의 첫 번째 입사기념일부터 선택된 년도 말일까지의 입사기념일 중
                            # TODAY 이전이거나 같은 가장 최근 입사기념일 찾기
                            actual_anniversary = None
                            # 입사일 이후의 모든 입사기념일 확인
                            for year_offset in range(1, 50):  # 최대 50년까지 확인
                                check_year = hire_date_obj.year + year_offset
                                if check_year > selected_year:
                                    break
                                try:
                                    check_anniversary = datetime(check_year, hire_month, hire_day).date()
                                except ValueError:
                                    check_anniversary = datetime(check_year, hire_month, hire_day - 1).date()
                                
                                if check_anniversary > hire_date_obj:
                                    if check_anniversary <= today:
                                        # TODAY 이전이거나 같은 입사기념일 중 가장 최근 것
                                        actual_anniversary = check_anniversary
                                    elif check_anniversary > today:
                                        # TODAY 이후의 입사기념일이면 중단
                                        break
                            
                            if actual_anniversary:
                                # 입사기념일이 TODAY 이전이거나 같으면 연차 생성
                                leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, actual_anniversary)
                            else:
                                # 입사기념일을 찾을 수 없거나 아직 지나지 않았으면 연차 생성 안 함
                                leave_generated = 0
                        else:
                            # 2025년 이하 조회 시: 기존 로직 유지
                            leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, datetime(selected_year, 12, 31).date())
                
                # 입사일 기준 매년 입사기념일마다 잔여 연차 소멸 처리
                # 입사일 기준으로 1년 동안 주어진 연차를 사용하지 않을 시 소멸
                # 소멸내역 컬럼에 기재
                current_date = datetime.now().date()
                anniversary_date = hire_date_obj
                max_years = 50  # 최대 50년까지 확인 (무한 루프 방지)
                
                # 입사일 이후의 모든 입사기념일 확인 (매년)
                for year_offset in range(1, max_years + 1):
                    # 다음 입사기념일 계산
                    try:
                        next_anniversary = datetime(hire_date_obj.year + year_offset, hire_date_obj.month, hire_date_obj.day).date()
                    except ValueError:
                        # 2월 29일 같은 경우 처리
                        next_anniversary = datetime(hire_date_obj.year + year_offset, hire_date_obj.month, hire_date_obj.day - 1).date()
                    
                    # 입사기념일이 현재 날짜보다 이후면 중단
                    if next_anniversary > current_date:
                        break
                    
                    # 입사기념일이 지난 경우 소멸 처리
                    # 입사일 기준으로 1년 동안 주어진 연차 - 사용 연차 = 소멸될 연차
                    if current_date >= next_anniversary:
                        # 기존 소멸 기록 확인 (중복 방지)
                        cursor.execute("""
                            SELECT id FROM leave_expirations
                            WHERE employee_id = ? AND leave_type = '연차' AND expiration_date = ?
                        """, (emp_id, next_anniversary))
                        existing_expiration = cursor.fetchone()
                        
                        if not existing_expiration:
                            # 이전 입사기념일 계산 (현재 입사기념일의 1년 전)
                            if year_offset == 1:
                                # 첫 해인 경우 입사일부터 계산
                                period_start = hire_date_obj
                            else:
                                prev_anniversary = datetime(hire_date_obj.year + year_offset - 1, hire_date_obj.month, hire_date_obj.day).date()
                                period_start = prev_anniversary
                            
                            # 입사일(또는 이전 입사기념일)부터 현재 입사기념일 직전까지 발생한 연차 계산
                            leave_generated_period = self.calculator.calculate_annual_leave(hire_date_obj, next_anniversary - timedelta(days=1))
                            
                            # 이전 입사기념일 시점의 연차를 빼야 함 (첫 해가 아닌 경우)
                            if year_offset > 1:
                                leave_generated_prev = self.calculator.calculate_annual_leave(hire_date_obj, period_start - timedelta(days=1))
                                leave_generated_period = leave_generated_period - leave_generated_prev
                            
                            # 해당 기간 동안 사용한 연차 계산 (입사일 또는 이전 입사기념일부터 현재 입사기념일 직전까지)
                            cursor.execute("""
                                SELECT SUM(leave_amount) as total_used
                                FROM leave_records
                                WHERE employee_id = ? AND leave_type = '연차'
                                AND leave_date >= ? AND leave_date < ?
                            """, (emp_id, period_start, next_anniversary))
                            used_result = cursor.fetchone()
                            total_used_period = used_result[0] if used_result[0] else 0
                            
                            # 소멸될 연차 = 입사일 기준으로 1년 동안 주어진 연차 - 사용 연차
                            expired_amount = max(0, leave_generated_period - total_used_period)
                            
                            if expired_amount > 0:
                                cursor.execute("""
                                    INSERT INTO leave_expirations 
                                    (employee_id, leave_type, expired_amount, expiration_date, year)
                                    VALUES (?, ?, ?, ?, ?)
                                """, (emp_id, '연차', expired_amount, next_anniversary, current_date.year))
                                conn.commit()
                                # 소멸 내역에 기록됨
                
                # 잔여 연차 계산 순서
                # 1단계: 이전 년도 남은 연차 - 입사일 기준 1년 소멸 차감 = 이전 년도 최종 남은 연차
                # 2단계: 잔여수 = 연차 발생 수 - 선택된 년도 사용연차 + 이전 년도 최종 남은 연차
                
                # 2026년 이상 조회 시: remaining_prev_year는 이미 2025년 연말 시점의 최종 잔여수이므로
                # 추가 소멸 차감이 필요 없습니다. (이미 2025년 연말 이전의 소멸이 차감된 상태)
                # 2025년 이하 조회 시: 현재 날짜 이전의 소멸 차감 필요
                prev_year = selected_year - 1
                prev_year_end = datetime(prev_year, 12, 31).date()
                
                if selected_year >= 2026:
                    # 2026년 이상 조회 시: 입사기념일 기준 소멸 확인
                    # 1년 이상자는 다음년도 입사기념일까지 사용하지 않으면 소멸
                    remaining_prev_year_float = float(remaining_prev_year) if remaining_prev_year is not None else 0.0
                    
                    # 2026년 입사기념일 계산
                    try:
                        anniversary_2026 = datetime(selected_year, hire_month, hire_day).date()
                    except ValueError:
                        anniversary_2026 = datetime(selected_year, hire_month, hire_day - 1).date()
                    
                    # 2026년 입사기념일이 지났는지 확인
                    current_date = datetime.now().date()
                    if current_date >= anniversary_2026:
                        # 2026년 입사기념일이 지났다면, 2025년 잔여 연차 중 사용하지 않은 부분 소멸
                        # 1년 이상자는 다음년도 입사기념일까지 사용하지 않으면 소멸
                        try:
                            anniversary_2025 = datetime(prev_year, hire_month, hire_day).date()
                        except ValueError:
                            anniversary_2025 = datetime(prev_year, hire_month, hire_day - 1).date()
                        
                        # 2025년 입사기념일부터 2026년 입사기념일 직전까지 사용한 연차
                        cursor.execute("""
                            SELECT SUM(leave_amount) as total_used
                            FROM leave_records
                            WHERE employee_id = ? AND leave_type = '연차'
                            AND leave_date >= ? AND leave_date < ?
                        """, (emp_id, anniversary_2025, anniversary_2026))
                        used_result = cursor.fetchone()
                        total_used_between_anniversaries = float(used_result[0]) if used_result[0] else 0.0
                        
                        # 2025년 잔여 연차 중 2026년 입사기념일까지 사용하지 않은 부분 소멸
                        # 2025년 잔여 연차가 2026년 입사기념일까지 사용되지 않으면 소멸
                        expired_2025_remaining = max(0.0, remaining_prev_year_float - total_used_between_anniversaries)
                        
                        if expired_2025_remaining > 0:
                            # 소멸 기록 확인 (중복 방지)
                            cursor.execute("""
                                SELECT id FROM leave_expirations
                                WHERE employee_id = ? AND leave_type = '연차' AND expiration_date = ?
                            """, (emp_id, anniversary_2026))
                            existing_expiration = cursor.fetchone()
                            
                            if not existing_expiration:
                                cursor.execute("""
                                    INSERT INTO leave_expirations 
                                    (employee_id, leave_type, expired_amount, expiration_date, year)
                                    VALUES (?, ?, ?, ?, ?)
                                """, (emp_id, '연차', expired_2025_remaining, anniversary_2026, selected_year))
                                conn.commit()
                        
                        remaining_prev_year_final = max(0.0, remaining_prev_year_float - expired_2025_remaining)
                        total_expired_float = expired_2025_remaining
                    else:
                        # 2026년 입사기념일이 아직 안 지났으면 소멸 없음
                        remaining_prev_year_final = remaining_prev_year_float
                        total_expired_float = 0.0
                else:
                    # 2025년 이하 조회 시: 현재 날짜 이전의 소멸 차감
                    current_date = datetime.now().date()
                    cursor.execute("""
                        SELECT SUM(expired_amount) as total_expired
                        FROM leave_expirations
                        WHERE employee_id = ? AND leave_type = '연차' AND expiration_date <= ?
                    """, (emp_id, current_date))
                    
                    expired_result = cursor.fetchone()
                    total_expired = expired_result[0] if expired_result[0] else 0
                    
                    # 1단계: 이전 년도 남은 연차 - 입사일 기준 1년 소멸 차감
                    # float로 명시적 변환하여 소수점 계산 지원
                    remaining_prev_year_float = float(remaining_prev_year) if remaining_prev_year is not None else 0.0
                    total_expired_float = float(total_expired) if total_expired is not None else 0.0
                    remaining_prev_year_final = max(0.0, remaining_prev_year_float - total_expired_float)
                
                # 2단계: 잔여수 = 연차 발생 수 - 2025년 사용연차 + 2024년 최종 남은 연차
                # 연차발생수와 사용연차를 명시적으로 float로 변환하여 계산
                leave_generated_float = float(leave_generated) if leave_generated is not None else 0.0
                used_current_year_float = float(used_current_year) if used_current_year is not None else 0.0
                remaining_prev_year_final_float = float(remaining_prev_year_final) if remaining_prev_year_final is not None else 0.0
                
                remaining = (leave_generated_float - used_current_year_float) + remaining_prev_year_final_float
                
                # 잔여수 값이 None이거나 계산되지 않은 경우를 방지
                if remaining is None:
                    remaining = 0.0
                # float 타입으로 명시적 변환
                remaining = float(remaining) if remaining is not None else 0.0
                
                # 1년 이상 재직인원의 경우 해당 년도 잔여수를 저장 (다음 년도 조회 시 사용)
                if is_one_year_or_more:
                    try:
                        cursor.execute("""
                            INSERT OR REPLACE INTO leave_remaining_by_year
                            (employee_id, year, remaining_amount, updated_at)
                            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                        """, (emp_id, selected_year, remaining))
                    except Exception as e:
                        print(f"잔여수 저장 오류 ({name}): {str(e)}")
                
                # 디버깅: 잔여수 계산 확인
                if name == "장지웅" or name == "전금희" or name == "김미라" or name == "김아름" or "김아름" in name:
                    print(f"DEBUG {name} ({selected_year}년) 잔여수: remaining_prev_year={remaining_prev_year}, remaining_prev_year_float={remaining_prev_year_float}")
                    if selected_year >= 2026:
                        print(f"DEBUG {name} ({selected_year}년) 소멸처리: total_expired_float={total_expired_float}, remaining_prev_year_final={remaining_prev_year_final}")
                    else:
                        print(f"DEBUG {name} ({selected_year}년) 소멸처리: total_expired={total_expired}, total_expired_float={total_expired_float}, remaining_prev_year_final={remaining_prev_year_final}")
                    print(f"DEBUG {name} ({selected_year}년) 잔여수: leave_generated={leave_generated_float}, used_current_year={used_current_year_float}, remaining_prev_year_final={remaining_prev_year_final_float}, remaining={remaining}")
                    print(f"DEBUG {name} ({selected_year}년) 계산식: ({leave_generated_float} - {used_current_year_float}) + {remaining_prev_year_final_float} = {remaining}")
                
                # 소멸 내역 조회 및 계산
                # 2025년 이상 조회 시: TODAY 기준으로 입사기념일이 지났을 때 소멸 내역 표기
                today = datetime.now().date()
                expiration_text = ""
                
                if selected_year >= 2025:
                    # 2025년 이상 조회 시: TODAY 기준으로 입사기념일이 지났는지 확인
                    expiration_list = []
                    
                    # 입사일 이후의 모든 입사기념일 확인
                    for year_offset in range(1, 50):  # 최대 50년까지 확인
                        try:
                            check_anniversary = datetime(hire_date_obj.year + year_offset, hire_month, hire_day).date()
                        except ValueError:
                            check_anniversary = datetime(hire_date_obj.year + year_offset, hire_month, hire_day - 1).date()
                        
                        # 입사기념일이 TODAY 이전이거나 같고, 선택된 년도 이내이면 소멸 내역 표기
                        # TODAY 기준으로 입사기념일이 지났을 때만 표기
                        if check_anniversary <= today and check_anniversary.year <= selected_year:
                            # 해당 입사기념일 직전 시점의 잔여수 계산
                            # 입사일부터 해당 입사기념일 직전까지의 연차 발생 수
                            if check_anniversary <= hire_date_obj:
                                # 입사기념일이 입사일보다 이전이면 다음 해 입사기념일 사용
                                continue
                            
                            # 입사일부터 입사기념일 직전까지 발생한 연차 계산
                            if year_offset == 1:
                                # 첫 해: 입사일부터 입사기념일 직전까지
                                period_start = hire_date_obj
                                leave_generated_period = self.calculator.calculate_annual_leave(hire_date_obj, check_anniversary - timedelta(days=1))
                            else:
                                # 이후 해: 이전 입사기념일부터 현재 입사기념일 직전까지
                                prev_anniversary = datetime(hire_date_obj.year + year_offset - 1, hire_month, hire_day).date()
                                try:
                                    prev_anniversary = datetime(hire_date_obj.year + year_offset - 1, hire_month, hire_day).date()
                                except ValueError:
                                    prev_anniversary = datetime(hire_date_obj.year + year_offset - 1, hire_month, hire_day - 1).date()
                                
                                period_start = prev_anniversary
                                leave_generated_period = self.calculator.calculate_annual_leave(hire_date_obj, check_anniversary - timedelta(days=1))
                                leave_generated_prev = self.calculator.calculate_annual_leave(hire_date_obj, prev_anniversary - timedelta(days=1))
                                leave_generated_period = leave_generated_period - leave_generated_prev
                            
                            # 해당 기간 동안 사용한 연차 계산
                            cursor.execute("""
                                SELECT work_date, leave_type, remarks
                                FROM attendance_records
                                WHERE employee_id = ? AND work_date >= ? AND work_date < ?
                                AND (leave_type IN ('연차', '반차', '휴가') 
                                     OR remarks IN ('반차_출근', '반차_퇴근'))
                            """, (emp_id, period_start, check_anniversary))
                            period_records = cursor.fetchall()
                            
                            used_period = 0.0
                            반차_날짜_set_period = set()
                            
                            for work_date, leave_type, remarks in period_records:
                                is_half_day = (leave_type == '반차' or remarks == '반차_출근' or remarks == '반차_퇴근')
                                if leave_type == '연차' or leave_type == '휴가':
                                    used_period += 1.0
                                elif is_half_day:
                                    if work_date not in 반차_날짜_set_period:
                                        used_period += 0.5
                                        반차_날짜_set_period.add(work_date)
                            
                            # 이전 입사기념일 시점의 잔여수 계산 (첫 해가 아닌 경우)
                            remaining_before_period = 0.0
                            if year_offset > 1:
                                # 이전 입사기념일 직전 시점의 잔여수 계산
                                prev_anniversary = datetime(hire_date_obj.year + year_offset - 1, hire_month, hire_day).date()
                                try:
                                    prev_anniversary = datetime(hire_date_obj.year + year_offset - 1, hire_month, hire_day).date()
                                except ValueError:
                                    prev_anniversary = datetime(hire_date_obj.year + year_offset - 1, hire_month, hire_day - 1).date()
                                
                                if prev_anniversary > hire_date_obj:
                                    # 이전 입사기념일 직전까지의 연차 발생 수
                                    if year_offset == 2:
                                        prev_period_start = hire_date_obj
                                        prev_leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, prev_anniversary - timedelta(days=1))
                                    else:
                                        prev_prev_anniversary = datetime(hire_date_obj.year + year_offset - 2, hire_month, hire_day).date()
                                        try:
                                            prev_prev_anniversary = datetime(hire_date_obj.year + year_offset - 2, hire_month, hire_day).date()
                                        except ValueError:
                                            prev_prev_anniversary = datetime(hire_date_obj.year + year_offset - 2, hire_month, hire_day - 1).date()
                                        prev_period_start = prev_prev_anniversary
                                        prev_leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, prev_anniversary - timedelta(days=1))
                                        prev_prev_leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, prev_prev_anniversary - timedelta(days=1))
                                        prev_leave_generated = prev_leave_generated - prev_prev_leave_generated
                                    
                                    # 이전 기간 동안 사용한 연차 계산
                                    cursor.execute("""
                                        SELECT work_date, leave_type, remarks
                                        FROM attendance_records
                                        WHERE employee_id = ? AND work_date >= ? AND work_date < ?
                                        AND (leave_type IN ('연차', '반차', '휴가') 
                                             OR remarks IN ('반차_출근', '반차_퇴근'))
                                    """, (emp_id, prev_period_start, prev_anniversary))
                                    prev_period_records = cursor.fetchall()
                                    
                                    prev_used = 0.0
                                    prev_반차_날짜_set = set()
                                    
                                    for work_date, leave_type, remarks in prev_period_records:
                                        is_half_day = (leave_type == '반차' or remarks == '반차_출근' or remarks == '반차_퇴근')
                                        if leave_type == '연차' or leave_type == '휴가':
                                            prev_used += 1.0
                                        elif is_half_day:
                                            if work_date not in prev_반차_날짜_set:
                                                prev_used += 0.5
                                                prev_반차_날짜_set.add(work_date)
                                    
                                    remaining_before_period = prev_leave_generated - prev_used
                            
                            # 해당 입사기념일 직전 시점의 잔여수 = (발생한 연차 - 사용한 연차) + 이전 잔여수
                            remaining_at_anniversary = (leave_generated_period - used_period) + remaining_before_period
                            
                            # 소멸될 연차 = 해당 입사기념일 직전 시점의 잔여수
                            expired_amount = max(0, remaining_at_anniversary)
                            
                            if expired_amount > 0:
                                if expired_amount.is_integer():
                                    expired_amount_str = str(int(expired_amount))
                                else:
                                    expired_amount_str = f"{expired_amount:.1f}"
                                expiration_list.append(f"연차 {expired_amount_str}개 ({check_anniversary})")
                        
                        # 입사기념일이 TODAY 이후이면 중단
                        if check_anniversary > today:
                            break
                    
                    # 소멸 내역이 있을 때만 표기, 없으면 공란 유지
                    if expiration_list:
                        expiration_text = ", ".join(expiration_list)
                    else:
                        expiration_text = ""  # 소멸 내역이 없으면 공란
                else:
                    # 2025년 이외 조회 시: 기존 로직 유지
                    cursor.execute("""
                        SELECT leave_type, expired_amount, expiration_date
                        FROM leave_expirations
                        WHERE employee_id = ?
                        ORDER BY expiration_date DESC
                    """, (emp_id,))
                    expirations = cursor.fetchall()
                    
                    # 소멸 내역이 있을 때만 표기, 없으면 공란 유지
                    if expirations:
                        expiration_list = []
                        for leave_type, amount, exp_date in expirations:
                            # 소멸된 연차가 0보다 클 때만 표기
                            if amount and float(amount) > 0:
                                expiration_list.append(f"{leave_type} {amount}개 ({exp_date})")
                        if expiration_list:
                            expiration_text = ", ".join(expiration_list)
                        else:
                            expiration_text = ""  # 소멸 내역이 없으면 공란
                    else:
                        expiration_text = ""  # 소멸 내역이 없으면 공란
                
                # 행 추가
                row = self.table.rowCount()
                self.table.insertRow(row)
                
                # 실제 직원 행의 행 번호 설정
                self.table.setVerticalHeaderItem(row, QTableWidgetItem(str(employee_row_number)))
                employee_row_number += 1
                
                # 0.0 값은 빈 문자열로 변환하는 헬퍼 함수
                # col_idx: 컬럼 번호 (잔여수 컬럼 19인 경우 0도 표시)
                def format_value(val, col_idx=None):
                    # None 체크
                    if val is None:
                        return ""
                    if isinstance(val, (int, float)):
                        val_float = float(val)
                        # 잔여수 컬럼(19)인 경우 0도 표시, 음수도 표시
                        if col_idx == 19:
                            if val_float.is_integer():
                                return str(int(val_float))
                            return f"{val_float:.1f}"
                        # 사용연차 컬럼(17)인 경우 소수점 표시 (반차 반영), 음수도 표시
                        if col_idx == 17:
                            if val_float == 0.0:
                                return ""
                            if val_float.is_integer():
                                return str(int(val_float))
                            return f"{val_float:.1f}"
                        # 다른 컬럼은 0이면 빈 문자열
                        if val == 0 or val == 0.0:
                            return ""
                        # float로 변환하여 소수점 처리
                        # 소수점이 0이면 정수로, 아니면 소수점 표시 (예: 14.5)
                        if val_float.is_integer():
                            return str(int(val_float))
                        # 소수점이 있으면 소수점 첫째 자리까지 표시 (예: 14.5, 1.5)
                        return f"{val_float:.1f}"
                    return str(val) if val else ""
                
                # 잔여수 값 확인 및 디버깅
                # print(f"DEBUG 잔여수: {name}, remaining={remaining}, type={type(remaining)}")
                
                # 연차발생수(18번)와 잔여수(19번)의 수동 입력 값이 있으면 먼저 삭제
                # 이렇게 하면 항상 자동 계산 값만 사용됨
                
                # 디버깅: 삭제 전에 수동 입력 값이 있었는지 확인
                if name == "장지웅" or name == "전금희" or name == "김미라" or name == "김아름":
                    cursor.execute("""
                        SELECT column_index, manual_value
                        FROM leave_manual_values
                        WHERE employee_id = ? AND year = ? AND column_index IN (18, 19)
                    """, (emp_id, selected_year))
                    deleted_manual = cursor.fetchall()
                    if deleted_manual:
                        print(f"DEBUG {name}: 삭제 전 수동 입력 값 발견 - {deleted_manual}")
                
                cursor.execute("""
                    DELETE FROM leave_manual_values
                    WHERE employee_id = ? AND year = ? AND column_index IN (18, 19)
                """, (emp_id, selected_year))
                # 중간 commit 제거 - 마지막에 한 번만 commit하여 데이터베이스 잠금 방지
                
                # 저장된 수동 입력 값 조회 (18번, 19번 제외)
                cursor.execute("""
                    SELECT column_index, manual_value
                    FROM leave_manual_values
                    WHERE employee_id = ? AND year = ? AND column_index NOT IN (18, 19)
                """, (emp_id, selected_year))
                manual_values_result = cursor.fetchall()
                manual_values_dict = {col_idx: val for col_idx, val in manual_values_result}
                
                # 디버깅: 19번 컬럼 수동 입력 값이 여전히 있는지 확인 (있으면 안 됨)
                if name == "장지웅" or name == "전금희" or name == "김미라" or name == "김아름" or "김아름" in name:
                    cursor.execute("""
                        SELECT column_index, manual_value
                        FROM leave_manual_values
                        WHERE employee_id = ? AND year = ? AND column_index = 19
                    """, (emp_id, selected_year))
                    remaining_manual = cursor.fetchone()
                    if remaining_manual:
                        print(f"WARNING {name}: 잔여수(19번) 수동 입력 값이 여전히 존재합니다! - {remaining_manual}")
                        # 강제 삭제
                        cursor.execute("""
                            DELETE FROM leave_manual_values
                            WHERE employee_id = ? AND year = ? AND column_index = 19
                        """, (emp_id, selected_year))
                        # 중간 commit 제거 - 마지막에 한 번만 commit하여 데이터베이스 잠금 방지
                        print(f"DEBUG {name}: 잔여수(19번) 수동 입력 값 강제 삭제 완료")
                
                # 기본 계산된 값들
                # 연차발생수와 사용연차를 float로 명시적 변환하여 소수점 표시 지원
                leave_generated_float = float(leave_generated) if leave_generated is not None else 0.0
                used_current_year_float = float(used_current_year) if used_current_year is not None else 0.0
                
                # 디버깅: 사용연차 계산 결과 확인
                if name == "장지웅" or name == "전금희" or name == "김미라" or name == "김아름" or "김아름" in name:
                    print(f"DEBUG {name}: used_current_year 원본={used_current_year}, type={type(used_current_year)}, used_current_year_float={used_current_year_float}, type={type(used_current_year_float)}, is_integer={used_current_year_float.is_integer() if isinstance(used_current_year_float, float) else 'N/A'}")
                
                # 17번 컬럼(2025년 사용연차)은 항상 attendance_records에서 계산한 값을 사용
                # 출퇴근 관리대장에서 입력한 연차가 반영되도록 함
                # 수동 입력 값이 있더라도 무시하고 계산된 값 사용
                if 17 in manual_values_dict:
                    # 수동 입력 값이 있으면 삭제 (출퇴근 관리대장에서 입력한 값이 우선)
                    cursor.execute("""
                        DELETE FROM leave_manual_values
                        WHERE employee_id = ? AND year = ? AND column_index = 17
                    """, (emp_id, selected_year))
                    if name == "장지웅" or name == "전금희" or name == "김미라" or name == "김아름" or "김아름" in name:
                        print(f"DEBUG {name}: 17번 컬럼 수동 입력 값 삭제 - 계산된 값({used_current_year_float}) 사용")
                
                # 잔여수 계산: 연차 발생 수 - 사용연차 + 2024년 최종 남은 연차
                remaining = (leave_generated_float - used_current_year_float) + remaining_prev_year_final_float
                # float 타입으로 명시적 변환
                remaining = float(remaining) if remaining is not None else 0.0
                
                base_values = [dept, pos, name, hire_date, format_value(remaining_prev_year)] + \
                             [format_value(monthly_usage[i]) for i in range(12)] + \
                             [format_value(used_current_year_float, col_idx=17), format_value(leave_generated_float), format_value(remaining, col_idx=19), expiration_text]
                
                # 디버깅: 연차발생수 및 잔여수 계산 결과 확인
                if name == "장지웅" or name == "전금희" or name == "김미라" or name == "김아름" or "김아름" in name:
                    print(f"DEBUG {name}: leave_generated 원본={leave_generated}, type={type(leave_generated)}, leave_generated_float={leave_generated_float}, format_value={format_value(leave_generated_float)}")
                    print(f"DEBUG {name}: used_current_year 원본={used_current_year}, used_current_year_float={used_current_year_float}, format_value(17번)={format_value(used_current_year_float, col_idx=17)}")
                    print(f"DEBUG {name}: remaining_prev_year_final={remaining_prev_year_final}, type={type(remaining_prev_year_final)}")
                    print(f"DEBUG {name}: remaining 계산 전 - leave_generated_float={leave_generated_float}, used_current_year_float={used_current_year_float}, remaining_prev_year_final_float={remaining_prev_year_final_float}")
                    print(f"DEBUG {name}: remaining={remaining}, type={type(remaining)}, is_integer={remaining.is_integer() if isinstance(remaining, float) else 'N/A'}, format_value(19번)={format_value(remaining, col_idx=19)}")
                    print(f"DEBUG {name}: manual_values_dict={manual_values_dict}, 18번 컬럼 수동 입력={manual_values_dict.get(18, '없음')}, 19번 컬럼 수동 입력={manual_values_dict.get(19, '없음')}")
                    print(f"DEBUG {name}: base_values[17]={base_values[17] if len(base_values) > 17 else 'N/A'}, base_values[18]={base_values[18] if len(base_values) > 18 else 'N/A'}, base_values[19]={base_values[19] if len(base_values) > 19 else 'N/A'}")
                
                # 수동 입력 값이 있으면 우선 사용, 없으면 계산된 값 사용
                # 단, 연차발생수(18번 컬럼), 잔여수(19번 컬럼), 사용연차(17번 컬럼)는 항상 자동 계산 값 사용
                values = []
                for col_idx, base_val in enumerate(base_values):
                    # 연차발생수(18번 컬럼), 잔여수(19번 컬럼), 사용연차(17번 컬럼)는 항상 자동 계산 값 사용
                    if col_idx == 17 or col_idx == 18 or col_idx == 19:
                        values.append(base_val)
                    elif col_idx in manual_values_dict:
                        # 저장된 수동 입력 값 사용 (월별 컬럼만)
                        manual_val = manual_values_dict[col_idx]
                        if manual_val is not None:
                            values.append(manual_val)
                            # 디버깅: 수동 입력 값 사용 확인
                            if name == "장지웅" and col_idx == 18:
                                print(f"DEBUG 장지웅: 수동 입력 값 사용 - {manual_val}")
                        else:
                            values.append(base_val)
                    else:
                        # 계산된 값 사용
                        values.append(base_val)
                
                for col, val in enumerate(values):
                    item = QTableWidgetItem(val)
                    # 연차발생수(18번)와 잔여수(19번)는 편집 불가
                    if col == 18 or col == 19:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        # 디버깅: 표시 값 확인
                        if name == "장지웅" or name == "전금희" or name == "김미라" or name == "김아름" or "김아름" in name:
                            if col == 17:  # 사용연차
                                print(f"DEBUG {name}: 사용연차 표시 값 - val='{val}', used_current_year_float={used_current_year_float}, format_value(17번)='{format_value(used_current_year_float, col_idx=17)}'")
                            if col == 19:  # 잔여수
                                print(f"DEBUG {name}: 잔여수 표시 값 - val='{val}', remaining={remaining}, format_value(19번)='{format_value(remaining, col_idx=19)}'")
                    # 사용연차(17번)와 잔여수(19번) 음수 표시 (연차, 반차 모두)
                    if col == 17:  # 사용연차
                        try:
                            val_float = float(used_current_year_float) if used_current_year_float is not None else 0.0
                            if val_float < 0:
                                item.setForeground(QColor("#FF0000"))  # 빨간색으로 표시
                        except:
                            pass
                    elif col == 19:  # 잔여수
                        try:
                            val_float = float(remaining) if remaining is not None else 0.0
                            if val_float < 0:
                                item.setForeground(QColor("#FF0000"))  # 빨간색으로 표시
                        except:
                            pass
                    # 직원 ID와 컬럼 정보 저장 (월별 컬럼만 편집 가능)
                    if col >= 5 and col <= 16:  # 1월~12월 컬럼 (5번째부터 16번째까지)
                        item.setData(Qt.UserRole, {'emp_id': emp_id, 'col': col, 'month': col - 4, 'year': selected_year})
                    else:
                        item.setData(Qt.UserRole, emp_id)  # 직원 ID만 저장
                    self.table.setItem(row, col, item)
            
            # 모든 작업 완료 후 한 번만 commit
            try:
                conn.commit()
            except Exception as e:
                conn.rollback()
                print(f"데이터베이스 커밋 오류: {str(e)}")
            finally:
                conn.close()
            
            # 재직인원 수 계산 (구분자 행 제외)
            employee_count = len(employees)
            if hasattr(self, 'employee_count_label'):
                self.employee_count_label.setText(f"재직인원: {employee_count}명")
            
            self._is_refreshing = False  # 새로고침 완료
        
        def upload_excel(self):
            """엑셀 파일 업로드"""
            file_path, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 선택", "", "Excel files (*.xlsx *.xls);;All files (*.*)")
            if not file_path:
                return
            
            try:
                file_ext = Path(file_path).suffix.lower()
                if file_ext == '.xls':
                    try:
                        df = pd.read_excel(file_path, header=0, engine='xlrd')
                    except:
                        df = pd.read_excel(file_path, header=None, engine='xlrd')
                        if len(df) > 0:
                            first_row = df.iloc[0].astype(str).tolist()
                            if any('부서' in str(cell) or '직급' in str(cell) or '이름' in str(cell) for cell in first_row):
                                df.columns = df.iloc[0]
                                df = df[1:].reset_index(drop=True)
                else:
                    try:
                        df = pd.read_excel(file_path, header=0, engine='openpyxl')
                    except:
                        df = pd.read_excel(file_path, header=None, engine='openpyxl')
                        if len(df) > 0:
                            first_row = df.iloc[0].astype(str).tolist()
                            if any('부서' in str(cell) or '직급' in str(cell) or '이름' in str(cell) for cell in first_row):
                                df.columns = df.iloc[0]
                                df = df[1:].reset_index(drop=True)
                
                conn = self.db.get_connection()
                cursor = conn.cursor()
                
                if df.columns.dtype == 'object':
                    df.columns = df.columns.str.strip()
                
                required_columns = ["부서", "직급", "이름", "입사일"]
                column_mapping = {}
                for req_col in required_columns:
                    if req_col in df.columns:
                        column_mapping[req_col] = req_col
                    else:
                        for df_col in df.columns:
                            if str(df_col).strip().lower() == req_col.lower():
                                column_mapping[req_col] = df_col
                                break
                
                missing_columns = [col for col in required_columns if col not in column_mapping]
                if missing_columns:
                    QMessageBox.critical(self, "오류", f"필수 컬럼이 없습니다: {', '.join(missing_columns)}")
                    conn.close()
                    return
                
                employees_added = 0
                leave_records_added = 0
                
                for idx, row in df.iterrows():
                    try:
                        department = str(row[column_mapping["부서"]]).strip() if pd.notna(row[column_mapping["부서"]]) else ""
                        position = str(row[column_mapping["직급"]]).strip() if pd.notna(row[column_mapping["직급"]]) else ""
                        name = str(row[column_mapping["이름"]]).strip() if pd.notna(row[column_mapping["이름"]]) else ""
                        
                        hire_date_str = row[column_mapping["입사일"]]
                        if pd.isna(hire_date_str):
                            continue
                        
                        if isinstance(hire_date_str, str):
                            try:
                                hire_date = datetime.strptime(hire_date_str, "%Y-%m-%d").date()
                            except:
                                try:
                                    hire_date = pd.to_datetime(hire_date_str).date()
                                except:
                                    continue
                        else:
                            hire_date = pd.to_datetime(hire_date_str).date()
                        
                        cursor.execute("""
                            INSERT OR IGNORE INTO employees (department, position, name, hire_date)
                            VALUES (?, ?, ?, ?)
                        """, (department, position, name, hire_date))
                        
                        cursor.execute("""
                            SELECT id FROM employees
                            WHERE department = ? AND position = ? AND name = ? AND hire_date = ?
                        """, (department, position, name, hire_date))
                        result = cursor.fetchone()
                        
                        if result:
                            employee_id = result[0]
                            employees_added += 1
                            
                            for month in range(1, 13):
                                month_col = f"{month}월"
                                if month_col in df.columns:
                                    leave_amount = row[month_col]
                                    if pd.notna(leave_amount) and float(leave_amount) > 0:
                                        leave_date = datetime(2025, month, 1).date()
                                        hire_date_obj = datetime.strptime(str(hire_date), "%Y-%m-%d").date()
                                        current_date = datetime(2025, 11, 1).date()
                                        leave_type = "월차" if (current_date - hire_date_obj).days < 365 else "연차"
                                        
                                        cursor.execute("""
                                            INSERT OR IGNORE INTO leave_records
                                            (employee_id, leave_type, leave_date, leave_amount, year, month)
                                            VALUES (?, ?, ?, ?, ?, ?)
                                        """, (employee_id, leave_type, leave_date, float(leave_amount), 2025, month))
                                        leave_records_added += 1
                    except Exception as e:
                        print(f"행 {idx+1} 처리 중 오류: {str(e)}")
                        continue
                
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, "성공",
                    f"엑셀 파일이 업로드되었습니다.\n직원: {employees_added}명 추가\n연차 기록: {leave_records_added}건 추가")
                self.refresh_data()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"엑셀 파일 업로드 중 오류 발생: {str(e)}")
        
        def download_excel(self, file_path_override=None, silent=False, open_after=True):
            """엑셀 파일 다운로드 - 테이블에 표시된 데이터를 그대로 다운로드
            (file_path_override가 주어지면 파일 다이얼로그 없이 해당 경로로 저장)
            """
            if file_path_override:
                file_path = file_path_override
            else:
                file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일 저장", "", "Excel files (*.xlsx);;All files (*.*)")
                if not file_path:
                    return
            
            try:
                import os
                # 파일이 이미 존재하고, 엑셀 등에서 열려있으면 PermissionError가 발생할 수 있으므로 선제 처리
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except PermissionError:
                        QMessageBox.warning(
                            self,
                            "경고",
                            f"파일이 다른 프로그램에서 열려있습니다.\n"
                            f"해당 파일을 닫고 다시 시도해주세요.\n\n파일: {file_path}"
                        )
                        return

                # 선택된 년도 가져오기
                selected_year = self.year_combo.currentData()
                if selected_year is None:
                    selected_year = datetime.now().year
                
                prev_year = selected_year - 1
                
                # 테이블에서 직접 데이터 읽기
                data = []
                row_count = self.table.rowCount()
                
                for row_idx in range(row_count):
                    # 부서(팀) 구분자 행은 엑셀 다운로드에서 제외
                    # (테이블에서 "━━━ 부서명 ━━━" 형태로 전체 컬럼 병합된 행)
                    first_item = self.table.item(row_idx, 0)
                    if first_item:
                        txt = (first_item.text() or "").strip()
                        if txt.startswith("━━━") and txt.endswith("━━━"):
                            continue

                    row_data = {}
                    
                    # 부서 (컬럼 0)
                    dept_item = self.table.item(row_idx, 0)
                    row_data["부서"] = dept_item.text() if dept_item else ""
                    
                    # 직급 (컬럼 1)
                    pos_item = self.table.item(row_idx, 1)
                    row_data["직급"] = pos_item.text() if pos_item else ""
                    
                    # 이름 (컬럼 2)
                    name_item = self.table.item(row_idx, 2)
                    row_data["이름"] = name_item.text() if name_item else ""
                    
                    # 입사일 (컬럼 3)
                    hire_date_item = self.table.item(row_idx, 3)
                    row_data["입사일"] = hire_date_item.text() if hire_date_item else ""
                    
                    # 이전 년도 남은연차 (컬럼 4)
                    prev_year_item = self.table.item(row_idx, 4)
                    prev_year_text = prev_year_item.text().strip() if prev_year_item else ""
                    try:
                        row_data[f"{prev_year}년"] = float(prev_year_text) if prev_year_text else 0.0
                    except ValueError:
                        row_data[f"{prev_year}년"] = 0.0
                    
                    # 월별 사용량 (컬럼 5-16: 1월~12월)
                    for month in range(1, 13):
                        month_col = 4 + month  # 5번째부터 16번째까지
                        month_item = self.table.item(row_idx, month_col)
                        month_text = month_item.text().strip() if month_item else ""
                        try:
                            row_data[f"{month}월"] = float(month_text) if month_text else 0.0
                        except ValueError:
                            row_data[f"{month}월"] = 0.0
                    
                    # 선택된 년도 사용연차 (컬럼 17)
                    used_year_item = self.table.item(row_idx, 17)
                    used_year_text = used_year_item.text().strip() if used_year_item else ""
                    try:
                        row_data[f"{selected_year}년"] = float(used_year_text) if used_year_text else 0.0
                    except ValueError:
                        row_data[f"{selected_year}년"] = 0.0
                    
                    # 연차발생수 (컬럼 18)
                    leave_gen_item = self.table.item(row_idx, 18)
                    leave_gen_text = leave_gen_item.text().strip() if leave_gen_item else ""
                    try:
                        row_data["연차발생수"] = float(leave_gen_text) if leave_gen_text else 0.0
                    except ValueError:
                        row_data["연차발생수"] = 0.0
                    
                    # 잔여수 (컬럼 19)
                    remaining_item = self.table.item(row_idx, 19)
                    remaining_text = remaining_item.text().strip() if remaining_item else ""
                    try:
                        row_data["잔여수"] = float(remaining_text) if remaining_text else 0.0
                    except ValueError:
                        row_data["잔여수"] = 0.0
                    
                    # 소멸내역 (컬럼 20)
                    expiration_item = self.table.item(row_idx, 20)
                    row_data["소멸내역"] = expiration_item.text() if expiration_item else ""
                    
                    data.append(row_data)
                
                if not data:
                    QMessageBox.warning(self, "알림", "다운로드할 데이터가 없습니다.")
                    return
                
                # 엑셀 파일 생성
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                    # 시트 이름 생성
                    sheet_name = f"{selected_year}년"
                    # Excel 시트 이름 제한 (31자)
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]
                    
                    df = pd.DataFrame(data)
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    
                    # 헤더 스타일 적용
                    worksheet = writer.sheets[sheet_name]
                    # 눈금선 숨기기
                    worksheet.sheet_view.showGridLines = False

                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    # Excel에서 글자색은 ARGB(8자리)가 안전합니다.
                    header_font = Font(bold=True, color="FF000000")  # 검정색
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # --- 두번째 이미지 스타일: 부서(팀) 셀 세로 병합 + 전체 테두리 + 컬럼 강조 ---
                    max_row = worksheet.max_row
                    max_col = worksheet.max_column

                    # (1) 부서(1열) 동일 구간 병합 (2행부터 데이터 시작)
                    def _merge_dept_blocks():
                        if max_row < 2:
                            return
                        start = 2
                        current = worksheet.cell(row=2, column=1).value
                        for r in range(3, max_row + 2):  # sentinel
                            v = worksheet.cell(row=r, column=1).value if r <= max_row else None
                            if v != current:
                                end = r - 1
                                if current and end >= start:
                                    if end > start:
                                        # 병합 전에 아래 셀 값을 비워야 함.
                                        # 병합 후에는 아래 셀이 MergedCell(읽기 전용)로 바뀌어 value 대입 시 오류 발생함.
                                        for rr in range(start + 1, end + 1):
                                            worksheet.cell(row=rr, column=1).value = ""
                                        worksheet.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
                                    # 가운데 정렬
                                    worksheet.cell(row=start, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                start = r
                                current = v

                    _merge_dept_blocks()

                    # (2) 전체 테두리: 값이 있는 영역(헤더 포함)에는 무조건 얇은 테두리
                    thin_side = Side(style="thin")
                    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                    for r in range(1, max_row + 1):
                        for c in range(1, max_col + 1):
                            cell = worksheet.cell(row=r, column=c)
                            # 데이터가 있는 셀 위주(헤더는 항상)
                            if r == 1 or (cell.value is not None and (not isinstance(cell.value, str) or cell.value.strip() != "")):
                                cell.border = thin_border
                                if r != 1 and c in (1, 2, 3, 4):
                                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    # (3) 컬럼 강조/헤더 표기: 헤더 텍스트로 컬럼 위치를 찾아 적용(연도 변경/컬럼 변화에도 안전)
                    try:
                        # 헤더 매핑(1행)
                        header_to_col = {}
                        for c in range(1, max_col + 1):
                            v = worksheet.cell(row=1, column=c).value
                            if v is None:
                                continue
                            header_to_col[str(v).strip()] = c

                        prev_year_key = f"{prev_year}년"
                        selected_year_key = f"{selected_year}년"

                        # 헤더를 두번째 이미지처럼 변경 (줄바꿈 + wrap)
                        if prev_year_key in header_to_col:
                            c_prev = header_to_col[prev_year_key]
                            worksheet.cell(row=1, column=c_prev).value = f"{prev_year_key}\n남은연차"
                        if selected_year_key in header_to_col:
                            c_sel = header_to_col[selected_year_key]
                            worksheet.cell(row=1, column=c_sel).value = f"{selected_year_key}\n사용연차"

                        # 헤더 줄바꿈 표시를 위해 1행 높이/정렬 설정
                        worksheet.row_dimensions[1].height = 28
                        for cell in worksheet[1]:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        # 컬럼 강조: 선택년도 사용연차(빨강), 잔여수(파랑 + 연녹색 채우기)
                        red_font = Font(color="FFFF0000")
                        blue_font = Font(color="FF0000FF")
                        green_fill = PatternFill(start_color="FFE2F0D9", end_color="FFE2F0D9", fill_type="solid")

                        c_sel = header_to_col.get(selected_year_key)
                        c_rem = header_to_col.get("잔여수")
                        if c_sel:
                            for r in range(2, max_row + 1):
                                worksheet.cell(row=r, column=c_sel).font = red_font
                        if c_rem:
                            for r in range(2, max_row + 1):
                                rem_cell = worksheet.cell(row=r, column=c_rem)
                                rem_cell.font = blue_font
                                rem_cell.fill = green_fill
                    except:
                        pass
                    
                    # 컬럼 너비 설정: "헤더명 기반"으로 적용(연도 변경에도 안전)
                    try:
                        header_to_col = {}
                        for c in range(1, worksheet.max_column + 1):
                            v = worksheet.cell(row=1, column=c).value
                            if v is None:
                                continue
                            header_to_col[str(v).split("\n")[0].strip()] = c  # 줄바꿈 앞부분 기준

                        def _set_width(header_name: str, w: int):
                            col_idx = header_to_col.get(header_name)
                            if col_idx:
                                worksheet.column_dimensions[get_column_letter(col_idx)].width = w

                        _set_width("부서", 12)
                        _set_width("직급", 8)
                        _set_width("이름", 10)
                        _set_width("입사일", 12)
                        _set_width(f"{prev_year}년", 12)

                        for m in range(1, 13):
                            _set_width(f"{m}월", 6)

                        _set_width(f"{selected_year}년", 12)
                        _set_width("연차발생수", 10)
                        _set_width("잔여수", 10)
                        _set_width("소멸내역", 38)
                    except:
                        pass

                    # (4) 전체 셀 가운데 정렬(가로/세로)
                    # - merged cell은 value 변경이 불가하지만, 스타일은 환경에 따라 예외가 날 수 있어 안전하게 처리합니다.
                    try:
                        align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
                        align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        for r in range(1, worksheet.max_row + 1):
                            for c in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=r, column=c)
                                try:
                                    wrap = bool(getattr(getattr(cell, "alignment", None), "wrap_text", False))
                                except Exception:
                                    wrap = False
                                try:
                                    cell.alignment = align_center_wrap if wrap else align_center
                                except Exception:
                                    # MergedCell 등 일부 객체는 환경에 따라 스타일 대입에서 예외가 날 수 있음
                                    pass
                    except Exception:
                        pass
                
                if not silent:
                    QMessageBox.information(self, "성공", f"엑셀 파일이 생성되었습니다.\n{selected_year}년 데이터가 다운로드되었습니다.")
                if open_after:
                    try:
                        import os
                        os.startfile(file_path)
                    except Exception:
                        pass
            except Exception as e:
                if not silent:
                    QMessageBox.critical(self, "오류", f"엑셀 파일 생성 중 오류 발생: {str(e)}")
                else:
                    raise
        
        def register_leave(self):
            """연차 사용 등록"""
            selected = self.table.selectedItems()
            if not selected:
                QMessageBox.warning(self, "경고", "직원을 선택해주세요.")
                return
            
            row = selected[0].row()
            emp_id = self.table.item(row, 0).data(Qt.UserRole)
            
            dialog = QDialog(self)
            dialog.setWindowTitle("연차 사용 등록")
            dialog.setModal(True)
            layout = QVBoxLayout(dialog)
            
            date_layout = QHBoxLayout()
            date_layout.addWidget(QLabel("사용일자:"))
            date_edit = QDateEdit()
            date_edit.setDate(QDate.currentDate())
            date_edit.setCalendarPopup(True)
            date_layout.addWidget(date_edit)
            layout.addLayout(date_layout)
            
            leave_type_layout = QHBoxLayout()
            leave_type_layout.addWidget(QLabel("연차 유형:"))
            leave_type_combo = QComboBox()
            leave_type_combo.addItems(["연차", "월차"])
            leave_type_layout.addWidget(leave_type_combo)
            layout.addLayout(leave_type_layout)
            
            amount_layout = QHBoxLayout()
            amount_layout.addWidget(QLabel("사용량:"))
            amount_combo = QComboBox()
            amount_combo.addItems(["1.0", "0.5"])
            amount_layout.addWidget(amount_combo)
            layout.addLayout(amount_layout)
            
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(lambda: self.save_leave(dialog, emp_id, date_edit.date().toPython(), leave_type_combo.currentText(), float(amount_combo.currentText())))
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)
            
            dialog.exec()

        def download_combined_excel(self):
            """연월차 + 출퇴근을 한 파일에 시트로 묶어서 다운로드"""
            # 출퇴근 GUI 참조가 없으면 기존 방식 사용
            attendance_gui = getattr(self, "attendance_gui", None)
            if attendance_gui is None:
                return self.download_excel()

            file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일 저장(연월차+출퇴근)", "", "Excel files (*.xlsx);;All files (*.*)")
            if not file_path:
                return

            import os
            # 덮어쓰기/잠금 처리
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except PermissionError:
                    QMessageBox.warning(self, "경고", f"파일이 다른 프로그램에서 열려있습니다.\n파일을 닫고 다시 시도해주세요.\n\n파일: {file_path}")
                    return

            root, ext = os.path.splitext(file_path)
            tmp_leave = f"{root}__tmp_leave{ext}"
            tmp_att = f"{root}__tmp_att{ext}"

            try:
                # 통합 다운로드 시 주말 색상 검증/보강에 사용할 출퇴근 연도(출퇴근 탭 기준)
                try:
                    att_year = attendance_gui.year_combo.currentData()
                    if att_year is None:
                        att_year = datetime.now().year
                    att_year = int(att_year)
                except Exception:
                    att_year = datetime.now().year

                # 1) 각각 임시 파일로 조용히 생성
                self.download_excel(file_path_override=tmp_leave, silent=True, open_after=False)
                attendance_gui._download_attendance_excel(file_path_override=tmp_att, silent=True, open_after=False)

                # 2) 임시 파일의 시트를 하나의 워크북으로 병합
                wb_leave = openpyxl.load_workbook(tmp_leave)
                wb_att = openpyxl.load_workbook(tmp_att)

                out_wb = openpyxl.Workbook()
                out_wb.remove(out_wb.active)

                def _clone_sheet(src_ws, dest_title):
                    from copy import copy
                    from openpyxl.utils import column_index_from_string
                    title = dest_title[:31]
                    dst_ws = out_wb.create_sheet(title=title)
                    # sheet view
                    try:
                        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
                        dst_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale
                    except Exception:
                        pass
                    # dimensions
                    for k, dim in src_ws.column_dimensions.items():
                        dst_ws.column_dimensions[k].width = dim.width
                    for r, dim in src_ws.row_dimensions.items():
                        dst_ws.row_dimensions[r].height = dim.height
                    # cells
                    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, min_col=1, max_col=src_ws.max_column):
                        for c in row:
                            # src가 MergedCell일 수 있고, dst는 merge 전이어야 값/스타일 주입 가능
                            col_raw = getattr(c, "col_idx", None)
                            if col_raw is None:
                                col_raw = getattr(c, "column", None)
                            if isinstance(col_raw, str):
                                col_idx = column_index_from_string(col_raw)
                            else:
                                col_idx = int(col_raw)

                            dc = dst_ws.cell(row=c.row, column=col_idx, value=c.value)
                            dc.number_format = c.number_format
                            # StyleProxy(해시 불가) 그대로 대입하면 저장 시 "unhashable type: StyleProxy"가 날 수 있어 복제해서 주입
                            dc.font = copy(c.font)
                            dc.fill = copy(c.fill)
                            dc.border = copy(c.border)
                            dc.alignment = copy(c.alignment)
                            dc.protection = copy(c.protection)
                    # merges (셀 복사 후 적용해야 MergedCell read-only 문제를 피할 수 있음)
                    for m in list(src_ws.merged_cells.ranges):
                        dst_ws.merge_cells(str(m))
                    return dst_ws

                def _apply_border_to_used_range(ws):
                    """통합 파일에서 테두리가 누락되는 경우를 방지하기 위해, 사용 영역 전체에 테두리를 확정 적용"""
                    max_r = ws.max_row or 1
                    max_c = ws.max_column or 1
                    # 테두리 색상을 명시해서 렌더링 안정화
                    thin = Side(style="thin", color="FF000000")
                    thick = Side(style="thick", color="FF000000")
                    cache = {}

                    def _b(l, r, t, b):
                        key = (l.style, r.style, t.style, b.style)
                        if key in cache:
                            return cache[key]
                        bd = Border(left=l, right=r, top=t, bottom=b)
                        cache[key] = bd
                        return bd

                    # 출퇴근 시트는 표의 상단이 2행(헤더 시작)이므로, 상단 외곽선 기준을 2행으로 맞춤
                    # (연월차 시트는 1행이 헤더이므로 1행 기준)
                    top_row = 2 if (ws.title.endswith("월") and ws.title[:-1].isdigit()) else 1

                    for r in range(top_row, max_r + 1):
                        for c in range(1, max_c + 1):
                            cell = ws.cell(row=r, column=c)
                            # 외곽선은 빈 셀이어도 반드시 보여야 하므로, 직사각형 범위 전체에 테두리 적용
                            l = thick if c == 1 else thin
                            rr = thick if c == max_c else thin
                            tt = thick if r == top_row else thin
                            bb = thick if r == max_r else thin
                            cell.border = _b(l, rr, tt, bb)

                def _apply_center_alignment_to_used_range(ws):
                    """통합 파일에서 모든 셀 가로/세로 가운데 정렬을 확정 적용"""
                    max_r = ws.max_row or 1
                    max_c = ws.max_column or 1
                    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    for r in range(1, max_r + 1):
                        for c in range(1, max_c + 1):
                            cell = ws.cell(row=r, column=c)
                            try:
                                wrap = bool(getattr(getattr(cell, "alignment", None), "wrap_text", False))
                            except Exception:
                                wrap = False
                            try:
                                cell.alignment = align_center_wrap if wrap else align_center
                            except Exception:
                                pass

                def _apply_weekend_fill_to_attendance_sheet(ws, year: int):
                    """통합 파일에서 출퇴근 시트(예: 1월)에 토/일 배경색을 확정 적용(시트 복사 과정 누락 방지)"""
                    try:
                        if not (ws.title.endswith("월") and ws.title[:-1].isdigit()):
                            return
                        m = int(ws.title[:-1])
                        from calendar import monthrange as _mr
                        days_in_month = _mr(int(year), m)[1]
                        weekend_days = set()
                        for d in range(1, days_in_month + 1):
                            if datetime(int(year), m, d).weekday() in (5, 6):
                                weekend_days.add(d)
                        # 날짜 컬럼: D(4)=1일 => col = 3 + day
                        fill_weekend = PatternFill(patternType="solid", fgColor="FFD0CECE")
                        max_r = ws.max_row or 1
                        for d in weekend_days:
                            col_idx = 3 + d
                            for r in range(2, max_r + 1):  # 헤더(2행)부터
                                try:
                                    ws.cell(row=r, column=col_idx).fill = fill_weekend
                                except Exception:
                                    pass
                    except Exception:
                        pass

                # 연월차: 시트 이름 고정
                leave_title = "연차집계표-최종"
                for ws in wb_leave.worksheets:
                    out_ws = _clone_sheet(ws, leave_title)
                    _apply_border_to_used_range(out_ws)
                    _apply_center_alignment_to_used_range(out_ws)

                # 출퇴근: 모든 시트 복사 (시트 이름은 월만 보이도록: "1월", "2월" ...)
                for ws in wb_att.worksheets:
                    out_ws = _clone_sheet(ws, ws.title)
                    _apply_border_to_used_range(out_ws)
                    _apply_center_alignment_to_used_range(out_ws)
                    _apply_weekend_fill_to_attendance_sheet(out_ws, att_year)

                out_wb.save(file_path)

                # 일부 Excel 환경에서 스타일이 "기록은 되는데 표시가 안 되는" 현상이 있어,
                # 한 번 다시 로드/저장하여 스타일 테이블을 정규화합니다.
                try:
                    wb_norm = openpyxl.load_workbook(file_path)
                    wb_norm.save(file_path)
                except Exception:
                    pass

                # --- 최종 파일 기준으로 검증 메시지 출력(통합 다운로드에서는 silent 생성이라 기존 검증이 안 보였음) ---
                verify_msg = f"\n\n[저장 경로]\n- {file_path}"
                try:
                    wb_chk = openpyxl.load_workbook(file_path)
                    # 출퇴근 검증 시트 선택: "1월" 우선, 없으면 첫 번째 월 시트, 그것도 없으면 active
                    att_ws = None
                    if "1월" in wb_chk.sheetnames:
                        att_ws = wb_chk["1월"]
                    else:
                        for name in wb_chk.sheetnames:
                            if name.endswith("월") and name[:-1].isdigit():
                                att_ws = wb_chk[name]
                                break
                    if att_ws is None:
                        att_ws = wb_chk.active

                    # 테두리 검증 (A2)
                    b = att_ws.cell(row=2, column=1).border
                    left = getattr(b, "left", None)
                    top = getattr(b, "top", None)
                    left_style = getattr(left, "style", None)
                    top_style = getattr(top, "style", None)
                    verify_msg += (
                        f"\n\n[출퇴근 테두리 검증]\n"
                        f"- 검사 시트: {att_ws.title}\n"
                        f"- A2 left: {left_style}\n"
                        f"- A2 top : {top_style}"
                    )

                    # 토/일 음영(fill) 검증: 해당 월의 첫 토/일 열을 찾아 헤더/데이터 fill 확인
                    try:
                        m = int(att_ws.title[:-1]) if att_ws.title.endswith("월") and att_ws.title[:-1].isdigit() else None
                        if m:
                            from calendar import monthrange as _mr
                            _days = _mr(int(att_year), m)[1]
                            weekend_day = None
                            for d in range(1, _days + 1):
                                if datetime(int(att_year), m, d).weekday() in (5, 6):
                                    weekend_day = d
                                    break
                            if weekend_day is not None:
                                col_idx = 3 + weekend_day  # 날짜 컬럼: D(4)=1일 => col=3+day
                                r_data = 4 if att_ws.max_row >= 4 else 2
                                h_sc = getattr(att_ws.cell(row=2, column=col_idx).fill, "start_color", None)
                                d_sc = getattr(att_ws.cell(row=r_data, column=col_idx).fill, "start_color", None)
                                h_fill = getattr(h_sc, "rgb", None) or getattr(h_sc, "indexed", None)
                                d_fill = getattr(d_sc, "rgb", None) or getattr(d_sc, "indexed", None)
                                verify_msg += (
                                    f"\n\n[출퇴근 토/일 음영 검증]\n"
                                    f"- 기준 연도: {att_year}\n"
                                    f"- 첫 토/일: {weekend_day}일\n"
                                    f"- 헤더(2행) fill: {h_fill}\n"
                                    f"- 데이터({r_data}행) fill: {d_fill}"
                                )
                    except Exception:
                        pass
                except Exception:
                    verify_msg += "\n\n[검증] 실패(파일 재열기 불가)"

                QMessageBox.information(self, "성공", "엑셀 파일이 생성되었습니다.\n(연차집계표-최종 + 출퇴근 월별 시트 포함)" + verify_msg)
                try:
                    os.startfile(file_path)
                except Exception:
                    pass
            except Exception as e:
                QMessageBox.critical(self, "오류", f"엑셀 파일 생성 중 오류 발생: {str(e)}")
            finally:
                for p in (tmp_leave, tmp_att):
                    try:
                        if os.path.exists(p):
                            os.remove(p)
                    except Exception:
                        pass
        
        def save_leave(self, dialog, emp_id, leave_date, leave_type, amount):
            """연차 사용 저장"""
            try:
                conn = self.db.get_connection()
                cursor = conn.cursor()
                
                cursor.execute("""
                    INSERT INTO leave_records
                    (employee_id, leave_type, leave_date, leave_amount, year, month)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (emp_id, leave_type, leave_date, amount, leave_date.year, leave_date.month))
                
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, "성공", "연차 사용이 등록되었습니다.")
                dialog.accept()
                self.refresh_data()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"등록 중 오류 발생: {str(e)}")
    
        def view_expirations(self):
            """소멸 내역 조회"""
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT e.name, le.leave_type, le.expired_amount, le.expiration_date, le.year
                FROM leave_expirations le
                JOIN employees e ON le.employee_id = e.id
                ORDER BY le.expiration_date DESC
            """)
            expirations = cursor.fetchall()
            conn.close()
            
            dialog = QDialog(self)
            dialog.setWindowTitle("소멸 내역 조회")
            dialog.setModal(True)
            dialog.resize(600, 400)
            layout = QVBoxLayout(dialog)
            
            text = QTextEdit()
            text.setReadOnly(True)
            
            if expirations:
                text.append("=== 소멸 내역 ===\n")
                for name, leave_type, amount, exp_date, year in expirations:
                    text.append(f"{name} - {leave_type} {amount}개 소멸 ({exp_date}, {year}년)")
            else:
                text.append("소멸된 연차가 없습니다.")
            
            layout.addWidget(text)
            
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)
            
            dialog.exec()
        
        def on_cell_changed(self, item):
            """셀 편집 완료 시 호출 - 셀 값 수정 및 데이터베이스 저장"""
            if not item:
                return
            
            # 데이터 새로고침 중일 때는 처리하지 않음 (무한 루프 방지)
            if self._is_refreshing:
                return
            
            # item이 유효한지 확인 (삭제된 경우 예외 처리)
            try:
                row = item.row()
                col = item.column()
            except RuntimeError:
                # item이 이미 삭제된 경우
                return
            
            # 입력된 값 가져오기
            new_value = item.text().strip()
            
            # UserRole 데이터 확인
            data = item.data(Qt.UserRole)
            emp_id = None
            if isinstance(data, dict):
                emp_id = data.get('emp_id')
            elif isinstance(data, int):
                emp_id = data
            
            if not emp_id:
                # 편집된 값은 그대로 유지 (데이터베이스 저장 없음)
                return
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # 선택된 년도 가져오기
            selected_year = self.year_combo.currentData()
            if selected_year is None:
                selected_year = datetime.now().year
            
            try:
                # 부서 (컬럼 0)
                if col == 0:
                    if new_value:
                        cursor.execute("""
                            UPDATE employees
                            SET department = ?
                            WHERE id = ?
                        """, (new_value, emp_id))
                        conn.commit()
                        # 수동 입력 값도 저장
                        cursor.execute("""
                            INSERT OR REPLACE INTO leave_manual_values
                            (employee_id, year, column_index, manual_value, updated_at)
                            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                        """, (emp_id, selected_year, col, new_value))
                        conn.commit()
                
                # 직급 (컬럼 1)
                elif col == 1:
                    if new_value:
                        cursor.execute("""
                            UPDATE employees
                            SET position = ?
                            WHERE id = ?
                        """, (new_value, emp_id))
                        conn.commit()
                        # 수동 입력 값도 저장
                        cursor.execute("""
                            INSERT OR REPLACE INTO leave_manual_values
                            (employee_id, year, column_index, manual_value, updated_at)
                            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                        """, (emp_id, selected_year, col, new_value))
                        conn.commit()
                
                # 이름 (컬럼 2)
                elif col == 2:
                    if new_value:
                        cursor.execute("""
                            UPDATE employees
                            SET name = ?
                            WHERE id = ?
                        """, (new_value, emp_id))
                        conn.commit()
                        # 수동 입력 값도 저장
                        cursor.execute("""
                            INSERT OR REPLACE INTO leave_manual_values
                            (employee_id, year, column_index, manual_value, updated_at)
                            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                        """, (emp_id, selected_year, col, new_value))
                        conn.commit()
                
                # 입사일 (컬럼 3)
                elif col == 3:
                    if new_value:
                        try:
                            # 날짜 형식 검증 및 변환
                            hire_date = datetime.strptime(new_value, "%Y-%m-%d").date()
                            cursor.execute("""
                                UPDATE employees
                                SET hire_date = ?
                                WHERE id = ?
                            """, (hire_date, emp_id))
                            conn.commit()
                            # 수동 입력 값도 저장
                            cursor.execute("""
                                INSERT OR REPLACE INTO leave_manual_values
                                (employee_id, year, column_index, manual_value, updated_at)
                                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                            """, (emp_id, selected_year, col, new_value))
                            conn.commit()
                            # 입사일이 변경되면 연차 계산에 영향을 주므로 전체 새로고침
                            self.refresh_data()
                            return
                        except ValueError:
                            QMessageBox.warning(self, "경고", "날짜 형식이 올바르지 않습니다.\n형식: YYYY-MM-DD (예: 2024-01-01)")
                            # 원래 값으로 복원
                            cursor.execute("""
                                SELECT hire_date FROM employees WHERE id = ?
                            """, (emp_id,))
                            result = cursor.fetchone()
                            if result:
                                item.setText(str(result[0]))
                            conn.close()
                            return
                
                # 이전 년도 남은연차 (컬럼 4)
                elif col == 4:
                    # 빈 값이면 수동 입력 값 삭제 (계산된 값으로 복원)
                    if not new_value:
                        cursor.execute("""
                            DELETE FROM leave_manual_values
                            WHERE employee_id = ? AND year = ? AND column_index = ?
                        """, (emp_id, selected_year, col))
                    else:
                        # 수동 입력 값 저장
                        cursor.execute("""
                            INSERT OR REPLACE INTO leave_manual_values
                            (employee_id, year, column_index, manual_value, updated_at)
                            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                        """, (emp_id, selected_year, col, new_value))
                    conn.commit()
                
                # 월별 컬럼 (5번째부터 16번째까지: 1월~12월) - 연차 사용량 저장
                elif col >= 5 and col <= 16:
                    month_data = data if isinstance(data, dict) else {}
                    month = month_data.get('month')
                    
                    if not month:
                        conn.close()
                        return
                    
                    try:
                        # 숫자로 변환
                        leave_amount = float(new_value) if new_value else 0.0
                    except ValueError:
                        QMessageBox.warning(self, "경고", "숫자만 입력 가능합니다.")
                        # 편집된 값 복원을 위해 새로고침하지 않고 원래 값으로 되돌림
                        selected_year = self.year_combo.currentData() or datetime.now().year
                        cursor.execute("""
                            SELECT leave_amount
                            FROM leave_records
                            WHERE employee_id = ? AND year = ? AND month = ?
                        """, (emp_id, selected_year, month))
                        result = cursor.fetchone()
                        old_value = result[0] if result and result[0] else 0.0
                        item.setText(str(old_value) if old_value else "")
                        conn.close()
                        return
                    
                    # 선택된 년도 가져오기
                    selected_year = self.year_combo.currentData()
                    if selected_year is None:
                        selected_year = datetime.now().year
                    
                    # 기존 기록 확인
                    cursor.execute("""
                        SELECT id, leave_amount
                        FROM leave_records
                        WHERE employee_id = ? AND year = ? AND month = ?
                    """, (emp_id, selected_year, month))
                    existing = cursor.fetchone()
                    
                    if leave_amount > 0:
                        if existing:
                            # 기존 기록 업데이트
                            cursor.execute("""
                                UPDATE leave_records
                                SET leave_amount = ?
                                WHERE employee_id = ? AND year = ? AND month = ?
                            """, (leave_amount, emp_id, selected_year, month))
                        else:
                            # 새로 생성 (연차로 저장)
                            cursor.execute("""
                                INSERT INTO leave_records
                                (employee_id, leave_type, leave_date, leave_amount, year, month)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (emp_id, '연차', datetime(selected_year, month, 1).date(), leave_amount, selected_year, month))
                    else:
                        # 0이면 삭제
                        if existing:
                            cursor.execute("""
                                DELETE FROM leave_records
                                WHERE employee_id = ? AND year = ? AND month = ?
                            """, (emp_id, selected_year, month))
                    
                    conn.commit()
                    
                    # 빈 값이면 수동 입력 값 삭제 (계산된 값으로 복원)
                    if not new_value:
                        cursor.execute("""
                            DELETE FROM leave_manual_values
                            WHERE employee_id = ? AND year = ? AND column_index = ?
                        """, (emp_id, selected_year, col))
                    else:
                        # 수동 입력 값도 저장
                        cursor.execute("""
                            INSERT OR REPLACE INTO leave_manual_values
                            (employee_id, year, column_index, manual_value, updated_at)
                            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                        """, (emp_id, selected_year, col, new_value))
                    conn.commit()
                    
                    # 편집된 값은 그대로 유지하고, 요약 정보(2025년 사용연차, 잔여수)만 업데이트
                    # 전체 새로고침 대신 해당 행의 요약 정보만 업데이트
                    self._update_summary_for_row(row)
                
                # 선택된 년도 사용연차 (컬럼 17)
                elif col == 17:
                    # 빈 값이면 수동 입력 값 삭제 (계산된 값으로 복원)
                    if not new_value:
                        cursor.execute("""
                            DELETE FROM leave_manual_values
                            WHERE employee_id = ? AND year = ? AND column_index = ?
                        """, (emp_id, selected_year, col))
                    else:
                        # 수동 입력 값 저장
                        cursor.execute("""
                            INSERT OR REPLACE INTO leave_manual_values
                            (employee_id, year, column_index, manual_value, updated_at)
                            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                        """, (emp_id, selected_year, col, new_value))
                    conn.commit()
                
                # 연차발생수 (컬럼 18) - 자동 계산 값만 사용, 수동 입력 불가
                elif col == 18:
                    # 연차발생수는 자동 계산 값만 사용하므로 수동 입력 값 삭제
                    cursor.execute("""
                        DELETE FROM leave_manual_values
                        WHERE employee_id = ? AND year = ? AND column_index = ?
                    """, (emp_id, selected_year, col))
                    conn.commit()
                    # 자동 계산 값으로 복원
                    self._update_summary_for_row(row)
                    return  # 편집 불가
                
                # 잔여수 (컬럼 19) - 자동 계산 값만 사용, 수동 입력 불가
                elif col == 19:
                    # 잔여수는 자동 계산 값만 사용하므로 수동 입력 값 삭제
                    cursor.execute("""
                        DELETE FROM leave_manual_values
                        WHERE employee_id = ? AND year = ? AND column_index = ?
                    """, (emp_id, selected_year, col))
                    conn.commit()
                    # 자동 계산 값으로 복원
                    self._update_summary_for_row(row)
                    return  # 편집 불가
                
                # 소멸내역 (컬럼 20)
                elif col == 20:
                    # 빈 값이면 수동 입력 값 삭제 (계산된 값으로 복원)
                    if not new_value:
                        cursor.execute("""
                            DELETE FROM leave_manual_values
                            WHERE employee_id = ? AND year = ? AND column_index = ?
                        """, (emp_id, selected_year, col))
                    else:
                        # 수동 입력 값 저장
                        cursor.execute("""
                            INSERT OR REPLACE INTO leave_manual_values
                            (employee_id, year, column_index, manual_value, updated_at)
                            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                        """, (emp_id, selected_year, col, new_value))
                    conn.commit()
                
            except Exception as e:
                QMessageBox.warning(self, "오류", f"데이터 저장 중 오류가 발생했습니다.\n{str(e)}")
                conn.rollback()
            finally:
                conn.close()
        
        def _update_summary_for_row(self, row):
            """특정 행의 요약 정보만 업데이트 (2025년 사용연차, 잔여수)"""
            try:
                # 행이 유효한지 확인
                if row < 0 or row >= self.table.rowCount():
                    return
                
                # 직원 ID 가져오기
                item = self.table.item(row, 0)  # 부서 컬럼에서 emp_id 가져오기
                if not item:
                    return
                
                data = item.data(Qt.UserRole)
                emp_id = data if isinstance(data, int) else None
                if not emp_id:
                    return
                
                conn = self.db.get_connection()
                cursor = conn.cursor()
                
                # 선택된 년도 가져오기
                selected_year = self.year_combo.currentData()
                if selected_year is None:
                    selected_year = datetime.now().year
                
                # 입사일 가져오기
                hire_date_item = self.table.item(row, 3)  # 입사일 컬럼
                if not hire_date_item:
                    conn.close()
                    return
                
                hire_date_str = hire_date_item.text()
                hire_date_obj = datetime.strptime(hire_date_str, "%Y-%m-%d").date()
                
                # 입사기념일 계산
                hire_month = hire_date_obj.month
                hire_day = hire_date_obj.day
                try:
                    anniversary_date = datetime(selected_year, hire_month, hire_day).date()
                except ValueError:
                    anniversary_date = datetime(selected_year, hire_month, hire_day - 1).date()
                
                # attendance_records 테이블에서 직접 사용연차 계산 (refresh_data와 동일한 로직)
                # 입사기념일 이후부터 다음 년도 1월 1일 전까지
                # 단, 입사일이 선택된 년도 내에 있고 입사기념일이 선택된 년도 내에 있으면
                # 입사일부터 계산 (입사일 이후 발생한 연차 사용 가능)
                if anniversary_date <= datetime(selected_year, 12, 31).date():
                    # 입사기념일이 선택된 년도 내에 있음
                    # 입사일이 선택된 년도 내에 있으면 입사일부터, 아니면 입사기념일부터
                    if hire_date_obj.year == selected_year:
                        # 입사일이 선택된 년도 내에 있으면 입사일부터 계산
                        usage_start_date = hire_date_obj
                    else:
                        # 입사일이 선택된 년도 이전이면 입사기념일부터 계산
                        # 입사기념일 당일부터 포함하여 계산 (>= 대신 > 사용 시 당일 제외됨)
                        usage_start_date = anniversary_date
                else:
                    # 입사기념일이 선택된 년도 이후에 있음 (아직 생성 안 됨)
                    # 입사일이 선택된 년도 내에 있으면 입사일부터, 아니면 연도 시작부터
                    if hire_date_obj.year == selected_year:
                        usage_start_date = hire_date_obj
                    else:
                        usage_start_date = datetime(selected_year, 1, 1).date()
                
                cursor.execute("""
                    SELECT work_date, leave_type, remarks
                    FROM attendance_records
                    WHERE employee_id = ? AND work_date >= ? AND work_date < ?
                    AND (leave_type IN ('연차', '반차', '휴가') 
                         OR remarks IN ('반차_출근', '반차_퇴근'))
                """, (emp_id, usage_start_date, datetime(selected_year + 1, 1, 1).date()))
                year_records = cursor.fetchall()
                
                used_current_year = 0.0
                반차_날짜_set = set()  # 반차 중복 방지
                
                for work_date, leave_type, remarks in year_records:
                    # remarks가 반차_출근 또는 반차_퇴근인 경우도 반차로 처리
                    is_half_day = (leave_type == '반차' or remarks == '반차_출근' or remarks == '반차_퇴근')
                    
                    if leave_type == '연차' or leave_type == '휴가':
                        used_current_year += 1.0
                    elif is_half_day:
                        # 반차는 하루에 0.5만 (같은 날짜에 여러 번 있어도 한 번만)
                        if work_date not in 반차_날짜_set:
                            used_current_year += 0.5
                            반차_날짜_set.add(work_date)
                
                # 수동 입력된 월별 연차 사용량도 사용연차에 포함
                # 월별 컬럼은 5번(1월)부터 16번(12월)까지
                # 입사기념일 이후의 월만 포함해야 함
                cursor.execute("""
                    SELECT column_index, manual_value
                    FROM leave_manual_values
                    WHERE employee_id = ? AND year = ? AND column_index >= 5 AND column_index <= 16
                """, (emp_id, selected_year))
                manual_monthly_values = cursor.fetchall()
                
                manual_used_current_year = 0.0
                for col_idx, manual_val in manual_monthly_values:
                    if manual_val:
                        try:
                            month_value = float(manual_val)
                            # 컬럼 인덱스 5=1월, 6=2월, ..., 16=12월
                            month = col_idx - 4
                            month_start = datetime(selected_year, month, 1).date()
                            # 월의 마지막 날짜 계산
                            if month == 12:
                                month_end = datetime(selected_year + 1, 1, 1).date()
                            else:
                                month_end = datetime(selected_year, month + 1, 1).date()
                            
                            # 입사기념일 이후의 월만 포함
                            # 입사기념일이 속한 월은 제외하고, 그 다음 월부터 포함
                            if anniversary_date <= datetime(selected_year, 12, 31).date():
                                # 입사기념일이 선택된 년도 내에 있으면 입사기념일이 속한 월을 제외하고 그 다음 월부터 포함
                                # 입사기념일이 해당 월 내에 있으면 해당 월 제외, 아니면 입사기념일 이후의 월만 포함
                                if anniversary_date >= month_start and anniversary_date < month_end:
                                    # 입사기념일이 해당 월 내에 있으면 해당 월 제외
                                    continue
                                elif month_start > anniversary_date:
                                    # 입사기념일 이후의 월만 포함
                                    manual_used_current_year += month_value
                            else:
                                # 입사기념일이 선택된 년도 이후에 있으면 모든 월 포함
                                manual_used_current_year += month_value
                        except (ValueError, TypeError):
                            pass
                
                # attendance_records에서 계산한 값과 수동 입력 월별 값을 합산
                used_current_year += manual_used_current_year
                
                # 디버깅: 사용연차 계산 결과 확인 (_update_summary_for_row)
                cursor.execute("SELECT name FROM employees WHERE id = ?", (emp_id,))
                emp_name_result = cursor.fetchone()
                emp_name = emp_name_result[0] if emp_name_result else "Unknown"
                if emp_name == "김미라" or emp_name == "전금희" or emp_name == "강지승":
                    print(f"DEBUG {emp_name} (_update_summary_for_row): 입사일={hire_date_obj}, 입사기념일={anniversary_date}, usage_start_date={usage_start_date}, year_records 개수={len(year_records)}, 반차_날짜_set={반차_날짜_set}")
                    print(f"DEBUG {emp_name} (_update_summary_for_row): attendance_records에서 계산={used_current_year - manual_used_current_year}, 수동 입력 월별={manual_used_current_year}, 합계={used_current_year}")
                    print(f"DEBUG {emp_name} (_update_summary_for_row): 조회 기간 = {usage_start_date} ~ {datetime(selected_year + 1, 1, 1).date()}")
                    if manual_monthly_values:
                        print(f"DEBUG {emp_name} (_update_summary_for_row): 수동 입력 월별 값 개수={len(manual_monthly_values)}")
                        for col_idx, manual_val in manual_monthly_values:
                            if manual_val:
                                try:
                                    month_value = float(manual_val)
                                    month = col_idx - 4
                                    month_start = datetime(selected_year, month, 1).date()
                                    print(f"DEBUG {emp_name} (_update_summary_for_row): 수동 입력 - {month}월: {month_value}, month_start={month_start}, anniversary_date={anniversary_date}")
                                except (ValueError, TypeError):
                                    pass
                    for work_date, leave_type, remarks in year_records:
                        is_half_day_check = (leave_type == '반차' or remarks == '반차_출근' or remarks == '반차_퇴근')
                        print(f"DEBUG {emp_name} (_update_summary_for_row): work_date={work_date}, leave_type={leave_type}, remarks={remarks}, is_half_day={is_half_day_check}")
                
                # 17번 컬럼(2025년 사용연차)은 항상 attendance_records에서 계산한 값을 사용
                # 수동 입력 값은 무시 (출퇴근 관리대장에서 입력한 값이 우선)
                
                # 연차 발생 수 계산 (기존 로직과 동일)
                days_from_hire_to_anniversary = (anniversary_date - hire_date_obj).days
                days_from_hire_to_year_end = (datetime(selected_year, 12, 31).date() - hire_date_obj).days
                
                current_date_for_calc = datetime.now().date()
                if selected_year > current_date_for_calc.year:
                    current_date_for_calc = datetime(selected_year, 12, 31).date()
                elif selected_year < current_date_for_calc.year:
                    current_date_for_calc = datetime(selected_year, 12, 31).date()
                
                days_from_hire_to_current = (current_date_for_calc - hire_date_obj).days
                
                if days_from_hire_to_anniversary < 365:
                    if days_from_hire_to_current < 0:
                        leave_generated = 0
                    else:
                        months_passed = (current_date_for_calc.year - hire_date_obj.year) * 12 + (current_date_for_calc.month - hire_date_obj.month)
                        if current_date_for_calc.day < hire_date_obj.day:
                            months_passed -= 1
                        leave_generated = min(max(months_passed, 0), 11)
                elif anniversary_date <= datetime(selected_year, 12, 31).date():
                    leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, anniversary_date)
                else:
                    if days_from_hire_to_year_end < 365:
                        year_end = datetime(selected_year, 12, 31).date()
                        if year_end < hire_date_obj:
                            leave_generated = 0
                        else:
                            months_passed = (year_end.year - hire_date_obj.year) * 12 + (year_end.month - hire_date_obj.month)
                            if year_end.day < hire_date_obj.day:
                                months_passed -= 1
                            leave_generated = min(max(months_passed, 0), 11)
                    else:
                        leave_generated = self.calculator.calculate_annual_leave(hire_date_obj, datetime(selected_year, 12, 31).date())
                
                # 이전 년도 남은 연차 가져오기
                remaining_prev_year_item = self.table.item(row, 4)
                remaining_prev_year = 0.0
                if remaining_prev_year_item:
                    try:
                        text = remaining_prev_year_item.text().strip()
                        remaining_prev_year = float(text) if text else 0.0
                    except:
                        remaining_prev_year = 0.0
                
                # 잔여 수 계산 순서
                # 1단계: 2024년 남은 연차 - 입사일 기준 1년 소멸 차감 = 2024년 최종 남은 연차
                # 2단계: 잔여수 = 연차 발생 수 - 2025년 사용연차 + 2024년 최종 남은 연차
                
                # 현재 날짜 이전에 소멸된 연차 조회 (입사일 기준 1년 소멸)
                current_date = datetime.now().date()
                cursor.execute("""
                    SELECT SUM(expired_amount) as total_expired
                    FROM leave_expirations
                    WHERE employee_id = ? AND leave_type = '연차' AND expiration_date <= ?
                """, (emp_id, current_date))
                expired_result = cursor.fetchone()
                total_expired = expired_result[0] if expired_result[0] else 0
                
                # 1단계: 2024년 남은 연차 - 입사일 기준 1년 소멸 차감
                # float로 명시적 변환하여 소수점 계산 지원
                remaining_prev_year_float = float(remaining_prev_year) if remaining_prev_year is not None else 0.0
                total_expired_float = float(total_expired) if total_expired is not None else 0.0
                remaining_prev_year_final = max(0.0, remaining_prev_year_float - total_expired_float)
                
                # 2단계: 잔여수 = 연차 발생 수 - 2025년 사용연차 + 2024년 최종 남은 연차
                # 연차발생수와 사용연차를 명시적으로 float로 변환하여 계산
                leave_generated_float = float(leave_generated) if leave_generated is not None else 0.0
                used_current_year_float = float(used_current_year) if used_current_year is not None else 0.0
                remaining_prev_year_final_float = float(remaining_prev_year_final) if remaining_prev_year_final is not None else 0.0
                
                remaining = (leave_generated_float - used_current_year_float) + remaining_prev_year_final_float
                
                # 잔여수 값이 None이거나 계산되지 않은 경우를 방지
                if remaining is None:
                    remaining = 0.0
                # float 타입으로 명시적 변환
                remaining = float(remaining) if remaining is not None else 0.0
                
                # 디버깅: 잔여수 계산 확인 (_update_summary_for_row)
                cursor.execute("SELECT name FROM employees WHERE id = ?", (emp_id,))
                emp_name_result = cursor.fetchone()
                emp_name = emp_name_result[0] if emp_name_result else "Unknown"
                if emp_name == "김미라":
                    print(f"DEBUG {emp_name} (_update_summary_for_row) 잔여수 계산: leave_generated={leave_generated_float}, used_current_year={used_current_year_float}, remaining_prev_year_final={remaining_prev_year_final_float}, remaining={remaining}")
                    print(f"DEBUG {emp_name} (_update_summary_for_row) 계산식: ({leave_generated_float} - {used_current_year_float}) + {remaining_prev_year_final_float} = {remaining}")
                
                conn.close()
                
                # 0.0 값은 빈 문자열로 변환하는 헬퍼 함수
                # col_idx: 컬럼 번호 (잔여수 컬럼 19인 경우 0도 표시, 사용연차 컬럼 17인 경우도 소수점 표시)
                def format_value(val, col_idx=None):
                    if isinstance(val, (int, float)):
                        val_float = float(val)
                        # 잔여수 컬럼(19)인 경우 0도 표시
                        if col_idx == 19:
                            if val_float.is_integer():
                                return str(int(val_float))
                            return f"{val_float:.1f}"
                        # 사용연차 컬럼(17)인 경우 소수점 표시 (반차 반영)
                        if col_idx == 17:
                            if val_float == 0.0:
                                return ""
                            if val_float.is_integer():
                                return str(int(val_float))
                            return f"{val_float:.1f}"
                        # 다른 컬럼은 0이면 빈 문자열
                        if val == 0 or val == 0.0:
                            return ""
                        # float로 변환하여 소수점 처리
                        # 소수점이 0이면 정수로, 아니면 소수점 표시 (예: 14.5)
                        if val_float.is_integer():
                            return str(int(val_float))
                        # 소수점이 있으면 소수점 첫째 자리까지 표시 (예: 14.5, 1.5)
                        return f"{val_float:.1f}"
                    return str(val) if val else ""
                
                # 요약 정보 업데이트 (셀 값은 유지)
                # 연차발생수와 사용연차를 float로 명시적 변환하여 소수점 표시 지원
                # (이미 위에서 계산된 float 값 사용)
                if self.table.item(row, 17):  # 선택된 년도 사용연차
                    item_17 = self.table.item(row, 17)
                    item_17.setText(format_value(used_current_year_float, col_idx=17))
                    # 음수 표시 (연차, 반차 모두)
                    try:
                        val_float = float(used_current_year_float) if used_current_year_float is not None else 0.0
                        if val_float < 0:
                            item_17.setForeground(QColor("#FF0000"))  # 빨간색으로 표시
                        else:
                            item_17.setForeground(QColor())  # 기본 색상으로 복원
                    except:
                        pass
                if self.table.item(row, 18):  # 연차발생수
                    self.table.item(row, 18).setText(format_value(leave_generated_float))
                if self.table.item(row, 19):  # 잔여수
                    item_19 = self.table.item(row, 19)
                    item_19.setText(format_value(remaining, col_idx=19))
                    # 음수 표시 (연차, 반차 모두)
                    try:
                        val_float = float(remaining) if remaining is not None else 0.0
                        if val_float < 0:
                            item_19.setForeground(QColor("#FF0000"))  # 빨간색으로 표시
                        else:
                            item_19.setForeground(QColor())  # 기본 색상으로 복원
                    except:
                        pass
                    
            except Exception as e:
                # 오류 발생 시 전체 새로고침
                print(f"요약 정보 업데이트 오류: {str(e)}")
                self.refresh_data()


    class EmployeeManagementGUI(QWidget):
        """재직인원 관리 GUI"""
        
        def __init__(self, parent, db_manager, leave_gui=None, attendance_gui=None):
            super().__init__(parent)
            self.db = db_manager
            self.leave_gui = leave_gui
            self.attendance_gui = attendance_gui
            
            layout = QVBoxLayout(self)
            
            button_layout = QHBoxLayout()
            button_layout.addWidget(QPushButton("직원 추가", clicked=self.add_employee))
            button_layout.addWidget(QPushButton("직원 삭제", clicked=self.delete_employee))
            button_layout.addWidget(QPushButton("퇴사 처리", clicked=self.deactivate_employee))
            button_layout.addWidget(QPushButton("재입사 처리", clicked=self.activate_employee))
            button_layout.addWidget(QPushButton("직원 일괄 등록", clicked=self.bulk_add_employees))
            button_layout.addWidget(QPushButton("새로고침", clicked=self.refresh_data))
            button_layout.addStretch()
            # 재직인원 수 표시 레이블
            self.employee_count_label = QLabel("재직인원: 0명")
            self.employee_count_label.setStyleSheet("font-weight: bold; color: #0066CC;")
            button_layout.addWidget(self.employee_count_label)
            layout.addLayout(button_layout)
            
            # 퇴사자 표시 옵션
            option_layout = QHBoxLayout()
            self.show_inactive_checkbox = QCheckBox("퇴사자 표시")
            self.show_inactive_checkbox.setChecked(False)
            # 체크박스 상태 변경 시 이 탭과 다른 탭들도 새로고침
            def on_checkbox_changed():
                # 다른 탭의 체크박스도 동기화
                if self.leave_gui:
                    self.leave_gui.show_inactive_checkbox.blockSignals(True)
                    self.leave_gui.show_inactive_checkbox.setChecked(self.show_inactive_checkbox.isChecked())
                    self.leave_gui.show_inactive_checkbox.blockSignals(False)
                    self.leave_gui.refresh_data()
                if self.attendance_gui:
                    self.attendance_gui.show_inactive_checkbox.blockSignals(True)
                    self.attendance_gui.show_inactive_checkbox.setChecked(self.show_inactive_checkbox.isChecked())
                    self.attendance_gui.show_inactive_checkbox.blockSignals(False)
                    self.attendance_gui.refresh_data()
                self.refresh_data()
            self.show_inactive_checkbox.stateChanged.connect(on_checkbox_changed)
            option_layout.addWidget(self.show_inactive_checkbox)
            option_layout.addStretch()
            layout.addLayout(option_layout)
            
            self.table = QTableWidget()
            self.table.setColumnCount(7)
            self.table.setHorizontalHeaderLabels(["부서", "직급", "이름", "연락처", "이메일", "입사일", "퇴사일"])
            # 컬럼 넓이 설정
            self.table.setColumnWidth(0, 100)  # 부서
            self.table.setColumnWidth(1, 60)   # 직급
            self.table.setColumnWidth(2, 80)   # 이름
            self.table.setColumnWidth(3, 100)  # 연락처
            self.table.setColumnWidth(4, 200)  # 이메일 (더 넓게)
            self.table.setColumnWidth(5, 100)  # 입사일
            self.table.setColumnWidth(6, 100)  # 퇴사일 (입사일과 같은 넓이)
            self.table.horizontalHeader().setStretchLastSection(False)
            self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
            self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
            # 직급(1), 연락처(3), 이메일(4) 컬럼만 편집 가능하도록 설정
            self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
            self.table.itemChanged.connect(self.on_cell_changed)
            layout.addWidget(self.table)
            
            self.refresh_data()
        
        def refresh_data(self):
            """데이터 새로고침"""
            self.table.setRowCount(0)
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # 퇴사자 표시 옵션 확인
            show_inactive = self.show_inactive_checkbox.isChecked()
            
            # is_active 필터 조건
            if show_inactive:
                # 모든 직원 표시 (활성/비활성 모두)
                where_clause = ""
            else:
                # 활성 직원만 표시
                where_clause = "WHERE COALESCE(is_active, 1) = 1"
            
            cursor.execute(f"""
                SELECT id, department, position, name, hire_date,
                       COALESCE(display_order, 0) as display_order,
                       COALESCE(is_active, 1) as is_active,
                       resignation_date,
                       COALESCE(phone, '') as phone,
                       COALESCE(email, '') as email
                FROM employees
                {where_clause}
                ORDER BY
                    CASE department
                        WHEN '경영지원팀' THEN 1
                        WHEN '영업팀' THEN 2
                        WHEN '글로벌비즈니스팀' THEN 3
                        ELSE 999
                    END,
                    department,
                    CASE position
                        WHEN '이사' THEN 1
                        WHEN '팀장' THEN 2
                        WHEN '파트장' THEN 3
                        WHEN '과장' THEN 4
                        WHEN '대리' THEN 5
                        WHEN '프로' THEN 6
                        ELSE 999
                    END,
                    hire_date ASC
            """)
            employees = cursor.fetchall()
            
            current_department = None
            employee_row_number = 1  # 실제 직원 행 번호 카운터
            for emp_id, dept, pos, name, hire_date, display_order, is_active, resignation_date, phone, email in employees:
                if current_department != dept:
                    # 부서 구분자 추가
                    row = self.table.rowCount()
                    self.table.insertRow(row)
                    
                    # 구분자 행의 행 번호를 공란으로 설정
                    self.table.setVerticalHeaderItem(row, QTableWidgetItem(""))
                    
                    item = QTableWidgetItem(f"━━━ {dept} ━━━")
                    item.setFlags(Qt.NoItemFlags)  # 선택 불가
                    item.setBackground(QColor("#E0E0E0"))
                    item.setTextAlignment(Qt.AlignCenter)  # 가운데 정렬
                    self.table.setItem(row, 0, item)
                    self.table.setSpan(row, 0, 1, 7)  # 7개 컬럼 병합
                    current_department = dept
                
                row = self.table.rowCount()
                self.table.insertRow(row)
                
                # 실제 직원 행의 행 번호 설정
                self.table.setVerticalHeaderItem(row, QTableWidgetItem(str(employee_row_number)))
                employee_row_number += 1
                
                # 퇴사일 표시용 문자열
                resignation_date_str = ""
                if resignation_date:
                    if isinstance(resignation_date, str):
                        resignation_date_str = resignation_date
                    else:
                        resignation_date_str = resignation_date.strftime("%Y-%m-%d")
                
                # 퇴사자인 경우 회색으로 표시
                # 컬럼 순서: 부서(0), 직급(1), 이름(2), 연락처(3), 이메일(4), 입사일(5), 퇴사일(6)
                for col, val in enumerate([dept, pos, name, phone, email, hire_date, resignation_date_str]):
                    item = QTableWidgetItem(str(val) if val else "")
                    item.setData(Qt.UserRole, emp_id)
                    # 모든 열 가운데 정렬
                    item.setTextAlignment(Qt.AlignCenter)
                    if is_active == 0:
                        # 퇴사자는 회색으로 표시
                        item.setForeground(QColor("#808080"))
                    # 직급(1), 연락처(3), 이메일(4) 컬럼만 편집 가능
                    if col not in [1, 3, 4]:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, col, item)
            
            # 재직인원 수 계산 (구분자 행 제외)
            employee_count = len(employees)
            if hasattr(self, 'employee_count_label'):
                self.employee_count_label.setText(f"재직인원: {employee_count}명")
            
            conn.close()
        
        def add_employee(self):
            """직원 추가"""
            dialog = QDialog(self)
            dialog.setWindowTitle("직원 추가")
            dialog.setModal(True)
            layout = QVBoxLayout(dialog)
            
            dept_layout = QHBoxLayout()
            dept_layout.addWidget(QLabel("부서:"))
            dept_entry = QLineEdit()
            dept_layout.addWidget(dept_entry)
            layout.addLayout(dept_layout)
            
            pos_layout = QHBoxLayout()
            pos_layout.addWidget(QLabel("직급:"))
            pos_entry = QLineEdit()
            pos_layout.addWidget(pos_entry)
            layout.addLayout(pos_layout)
            
            name_layout = QHBoxLayout()
            name_layout.addWidget(QLabel("이름:"))
            name_entry = QLineEdit()
            name_layout.addWidget(name_entry)
            layout.addLayout(name_layout)
            
            phone_layout = QHBoxLayout()
            phone_layout.addWidget(QLabel("연락처 (선택사항):"))
            phone_entry = QLineEdit()
            phone_layout.addWidget(phone_entry)
            layout.addLayout(phone_layout)
            
            email_layout = QHBoxLayout()
            email_layout.addWidget(QLabel("이메일 (선택사항):"))
            email_entry = QLineEdit()
            email_layout.addWidget(email_entry)
            layout.addLayout(email_layout)
            
            hire_layout = QHBoxLayout()
            hire_layout.addWidget(QLabel("입사일 (YYYY-MM-DD, 선택사항):"))
            hire_entry = QLineEdit()
            hire_entry.setText(datetime.now().strftime("%Y-%m-%d"))
            hire_layout.addWidget(hire_entry)
            layout.addLayout(hire_layout)
            
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(lambda: self.save_employee(dialog, dept_entry.text(), pos_entry.text(), name_entry.text(), phone_entry.text(), email_entry.text(), hire_entry.text()))
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)
            
            dialog.exec()
        
        def save_employee(self, dialog, department, position, name, phone, email, hire_date_str):
            """직원 저장"""
            try:
                if not all([department, position, name]):
                    QMessageBox.warning(self, "경고", "부서, 직급, 이름은 필수 입력 항목입니다.")
                    return
                
                if hire_date_str.strip():
                    try:
                        hire_date = datetime.strptime(hire_date_str.strip(), "%Y-%m-%d").date()
                    except ValueError:
                        QMessageBox.critical(self, "오류", "입사일 형식이 올바르지 않습니다. (YYYY-MM-DD 형식)")
                        return
                else:
                    hire_date = datetime(1900, 1, 1).date()
                
                conn = self.db.get_connection()
                cursor = conn.cursor()
                
                # 같은 부서의 직원이 있는지 확인
                cursor.execute("""
                    SELECT COUNT(*) FROM employees WHERE department = ?
                """, (department.strip(),))
                dept_count = cursor.fetchone()[0]
                
                if dept_count > 0:
                    # 같은 부서가 있으면 그 부서의 마지막 직원 다음에 추가
                    cursor.execute("""
                        SELECT MAX(COALESCE(display_order, 0))
                        FROM employees
                        WHERE department = ?
                    """, (department.strip(),))
                    max_order_result = cursor.fetchone()
                    max_order = max_order_result[0] if max_order_result[0] is not None else 0
                    new_order = max_order + 1
                else:
                    # 같은 부서가 없으면 맨 위에 추가 (기존 로직)
                    cursor.execute("SELECT MIN(COALESCE(display_order, 0)) FROM employees")
                    min_order_result = cursor.fetchone()
                    new_order = (min_order_result[0] if min_order_result[0] else 0) - 1
                
                # 연락처와 이메일 처리 (빈 문자열이면 None으로 저장)
                phone_value = phone.strip() if phone.strip() else None
                email_value = email.strip() if email.strip() else None
                
                cursor.execute("""
                    INSERT INTO employees (department, position, name, hire_date, display_order, phone, email)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (department.strip(), position.strip(), name.strip(), hire_date, new_order, phone_value, email_value))
                
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, "성공", "직원이 추가되었습니다.")
                dialog.accept()
                self.refresh_data()
                if self.leave_gui:
                    self.leave_gui.refresh_data()
                if self.attendance_gui:
                    self.attendance_gui.refresh_data()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"직원 추가 중 오류 발생: {str(e)}")
        
        def delete_employee(self):
            """직원 삭제"""
            selected = self.table.selectedItems()
            if not selected:
                QMessageBox.warning(self, "경고", "삭제할 직원을 선택해주세요.")
                return
            
            # 선택된 행의 직원 ID 수집 (구분자 행 제외, 중복 제거)
            selected_rows = set()  # 중복된 행 번호 제거를 위한 set
            for item in selected:
                row = item.row()
                selected_rows.add(row)
            
            selected_employees = []
            for row in selected_rows:
                emp_id_item = self.table.item(row, 0)
                if emp_id_item and emp_id_item.flags() != Qt.NoItemFlags:
                    emp_id = emp_id_item.data(Qt.UserRole)
                    if emp_id:
                        dept = self.table.item(row, 0).text()
                        pos = self.table.item(row, 1).text()
                        name = self.table.item(row, 2).text()
                        selected_employees.append({'emp_id': emp_id, 'name': name, 'row': row})
            
            if not selected_employees:
                QMessageBox.warning(self, "경고", "삭제할 직원을 선택해주세요.")
                return
            
            if len(selected_employees) == 1:
                message = f"다음 직원을 삭제하시겠습니까?\n\n{selected_employees[0]['name']}\n\n※ 삭제 시 해당 직원의 모든 연차 기록과 출퇴근 기록도 함께 삭제됩니다."
            else:
                names = [emp['name'] for emp in selected_employees[:5]]
                name_list = "\n".join(f"- {name}" for name in names)
                if len(selected_employees) > 5:
                    name_list += f"\n... 외 {len(selected_employees) - 5}명"
                message = f"다음 {len(selected_employees)}명의 직원을 삭제하시겠습니까?\n\n{name_list}\n\n※ 삭제 시 해당 직원들의 모든 연차 기록과 출퇴근 기록도 함께 삭제됩니다."
            
            reply = QMessageBox.question(self, "직원 삭제 확인", message, QMessageBox.Yes | QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                try:
                    conn = self.db.get_connection()
                    cursor = conn.cursor()
                    
                    for emp in selected_employees:
                        emp_id = emp['emp_id']
                        cursor.execute("DELETE FROM leave_records WHERE employee_id = ?", (emp_id,))
                        cursor.execute("DELETE FROM leave_expirations WHERE employee_id = ?", (emp_id,))
                        cursor.execute("DELETE FROM attendance_records WHERE employee_id = ?", (emp_id,))
                        cursor.execute("DELETE FROM employees WHERE id = ?", (emp_id,))
                    
                    conn.commit()
                    conn.close()
                    
                    QMessageBox.information(self, "성공", f"{len(selected_employees)}명의 직원이 삭제되었습니다.")
                    self.refresh_data()
                    if self.leave_gui:
                        self.leave_gui.refresh_data()
                    if self.attendance_gui:
                        self.attendance_gui.refresh_data()
                except Exception as e:
                    QMessageBox.critical(self, "오류", f"직원 삭제 중 오류 발생: {str(e)}")
        
        def deactivate_employee(self):
            """직원 퇴사 처리 (숨김)"""
            selected = self.table.selectedItems()
            if not selected:
                QMessageBox.warning(self, "경고", "퇴사 처리할 직원을 선택해주세요.")
                return
            
            # 선택된 행의 직원 ID 수집 (구분자 행 제외, 중복 제거)
            selected_rows = set()
            for item in selected:
                row = item.row()
                selected_rows.add(row)
            
            selected_employees = []
            for row in selected_rows:
                emp_id_item = self.table.item(row, 0)
                if emp_id_item and emp_id_item.flags() != Qt.NoItemFlags:
                    emp_id = emp_id_item.data(Qt.UserRole)
                    if emp_id:
                        name = self.table.item(row, 2).text()
                        selected_employees.append({'emp_id': emp_id, 'name': name})
            
            if not selected_employees:
                QMessageBox.warning(self, "경고", "퇴사 처리할 직원을 선택해주세요.")
                return
            
            # 퇴사일 입력 다이얼로그
            from PySide6.QtWidgets import QDialog, QVBoxLayout, QLabel, QDateEdit, QDialogButtonBox
            from PySide6.QtCore import QDate
            
            dialog = QDialog(self)
            dialog.setWindowTitle("퇴사일 입력")
            layout = QVBoxLayout(dialog)
            
            layout.addWidget(QLabel("퇴사일을 선택해주세요:"))
            resignation_date_edit = QDateEdit()
            resignation_date_edit.setCalendarPopup(True)
            resignation_date_edit.setDate(QDate.currentDate())
            resignation_date_edit.setDisplayFormat("yyyy-MM-dd")
            layout.addWidget(resignation_date_edit)
            
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)
            
            if dialog.exec() != QDialog.Accepted:
                return
            
            resignation_date = resignation_date_edit.date().toPython()
            
            if len(selected_employees) == 1:
                message = f"다음 직원을 퇴사 처리하시겠습니까?\n\n{selected_employees[0]['name']}\n퇴사일: {resignation_date}\n\n※ 퇴사 처리 시 퇴사한 월까지는 출퇴근 관리대장과 연월차 관리대장에 표시됩니다.\n데이터는 삭제되지 않으며, 필요 시 재입사 처리로 복구할 수 있습니다."
            else:
                names = [emp['name'] for emp in selected_employees[:5]]
                name_list = "\n".join(f"- {name}" for name in names)
                if len(selected_employees) > 5:
                    name_list += f"\n... 외 {len(selected_employees) - 5}명"
                message = f"다음 {len(selected_employees)}명의 직원을 퇴사 처리하시겠습니까?\n\n{name_list}\n퇴사일: {resignation_date}\n\n※ 퇴사 처리 시 퇴사한 월까지는 출퇴근 관리대장과 연월차 관리대장에 표시됩니다.\n데이터는 삭제되지 않으며, 필요 시 재입사 처리로 복구할 수 있습니다."
            
            reply = QMessageBox.question(self, "퇴사 처리 확인", message, QMessageBox.Yes | QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                try:
                    conn = self.db.get_connection()
                    cursor = conn.cursor()
                    
                    for emp in selected_employees:
                        emp_id = emp['emp_id']
                        cursor.execute("UPDATE employees SET is_active = 0, resignation_date = ? WHERE id = ?", (resignation_date, emp_id))
                    
                    conn.commit()
                    conn.close()
                    
                    QMessageBox.information(self, "성공", f"{len(selected_employees)}명의 직원이 퇴사 처리되었습니다.")
                    self.refresh_data()
                    if self.leave_gui:
                        self.leave_gui.refresh_data()
                    if self.attendance_gui:
                        self.attendance_gui.refresh_data()
                except Exception as e:
                    QMessageBox.critical(self, "오류", f"퇴사 처리 중 오류 발생: {str(e)}")
        
        def activate_employee(self):
            """직원 재입사 처리 (활성화)"""
            selected = self.table.selectedItems()
            if not selected:
                QMessageBox.warning(self, "경고", "재입사 처리할 직원을 선택해주세요.")
                return
            
            # 선택된 행의 직원 ID 수집 (구분자 행 제외, 중복 제거)
            selected_rows = set()
            for item in selected:
                row = item.row()
                selected_rows.add(row)
            
            selected_employees = []
            for row in selected_rows:
                emp_id_item = self.table.item(row, 0)
                if emp_id_item and emp_id_item.flags() != Qt.NoItemFlags:
                    emp_id = emp_id_item.data(Qt.UserRole)
                    if emp_id:
                        name = self.table.item(row, 2).text()
                        selected_employees.append({'emp_id': emp_id, 'name': name})
            
            if not selected_employees:
                QMessageBox.warning(self, "경고", "재입사 처리할 직원을 선택해주세요.")
                return
            
            if len(selected_employees) == 1:
                message = f"다음 직원을 재입사 처리하시겠습니까?\n\n{selected_employees[0]['name']}\n\n※ 재입사 처리 시 출퇴근 관리대장과 연월차 관리대장에 다시 표시됩니다."
            else:
                names = [emp['name'] for emp in selected_employees[:5]]
                name_list = "\n".join(f"- {name}" for name in names)
                if len(selected_employees) > 5:
                    name_list += f"\n... 외 {len(selected_employees) - 5}명"
                message = f"다음 {len(selected_employees)}명의 직원을 재입사 처리하시겠습니까?\n\n{name_list}\n\n※ 재입사 처리 시 출퇴근 관리대장과 연월차 관리대장에 다시 표시됩니다."
            
            reply = QMessageBox.question(self, "재입사 처리 확인", message, QMessageBox.Yes | QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                try:
                    conn = self.db.get_connection()
                    cursor = conn.cursor()
                    
                    for emp in selected_employees:
                        emp_id = emp['emp_id']
                        cursor.execute("UPDATE employees SET is_active = 1, resignation_date = NULL WHERE id = ?", (emp_id,))
                    
                    conn.commit()
                    conn.close()
                    
                    QMessageBox.information(self, "성공", f"{len(selected_employees)}명의 직원이 재입사 처리되었습니다.")
                    self.refresh_data()
                    if self.leave_gui:
                        self.leave_gui.refresh_data()
                    if self.attendance_gui:
                        self.attendance_gui.refresh_data()
                except Exception as e:
                    QMessageBox.critical(self, "오류", f"재입사 처리 중 오류 발생: {str(e)}")
        
        def on_cell_changed(self, item):
            """셀 편집 완료 시 호출 - 직급, 연락처, 이메일 수정"""
            if not item:
                return
            
            try:
                row = item.row()
                col = item.column()
            except RuntimeError:
                return
            
            # 직급(1), 연락처(3), 이메일(4) 컬럼만 처리
            if col not in [1, 3, 4]:
                return
            
            # 구분자 행인지 확인
            if item.flags() == Qt.NoItemFlags:
                return
            
            # 직원 ID 가져오기
            emp_id = item.data(Qt.UserRole)
            if not emp_id:
                return
            
            # 새로운 값
            new_value = item.text().strip()
            
            try:
                conn = self.db.get_connection()
                cursor = conn.cursor()
                
                if col == 1:  # 직급
                    if not new_value:
                        QMessageBox.warning(self, "경고", "직급을 입력해주세요.")
                        self.refresh_data()
                        return
                    cursor.execute("UPDATE employees SET position = ? WHERE id = ?", (new_value, emp_id))
                elif col == 3:  # 연락처
                    cursor.execute("UPDATE employees SET phone = ? WHERE id = ?", (new_value if new_value else None, emp_id))
                elif col == 4:  # 이메일
                    cursor.execute("UPDATE employees SET email = ? WHERE id = ?", (new_value if new_value else None, emp_id))
                
                conn.commit()
                conn.close()
                
                # 직급 수정 시에만 다른 탭도 새로고침
                if col == 1:
                    if self.leave_gui:
                        self.leave_gui.refresh_data()
                    if self.attendance_gui:
                        self.attendance_gui.refresh_data()
                
            except Exception as e:
                QMessageBox.critical(self, "오류", f"데이터 수정 중 오류 발생: {str(e)}")
                self.refresh_data()
        
        def bulk_add_employees(self):
            """직원 일괄 등록"""
            dialog = QDialog(self)
            dialog.setWindowTitle("직원 일괄 등록")
            dialog.setModal(True)
            dialog.resize(600, 400)
            layout = QVBoxLayout(dialog)
            
            layout.addWidget(QLabel("형식: 부서|직급|이름|입사일 (YYYY-MM-DD)\n각 직원은 한 줄에 입력하세요. 입사일은 선택사항입니다."))
            
            text_edit = QTextEdit()
            layout.addWidget(text_edit)
            
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(lambda: self.save_bulk_employees(dialog, text_edit.toPlainText()))
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)
            
            dialog.exec()
        
        def save_bulk_employees(self, dialog, text):
            """일괄 직원 저장"""
            try:
                lines = [line.strip() for line in text.split('\n') if line.strip()]
                if not lines:
                    QMessageBox.warning(self, "경고", "입력된 직원이 없습니다.")
                    return
                
                conn = self.db.get_connection()
                cursor = conn.cursor()
                
                cursor.execute("SELECT MAX(COALESCE(display_order, 0)) FROM employees")
                max_order_result = cursor.fetchone()
                current_order = (max_order_result[0] if max_order_result[0] else 0) + 1
                
                added_count = 0
                for line in lines:
                    parts = line.split('|')
                    if len(parts) >= 3:
                        dept = parts[0].strip()
                        pos = parts[1].strip()
                        name = parts[2].strip()
                        
                        if len(parts) >= 4 and parts[3].strip():
                            try:
                                hire_date = datetime.strptime(parts[3].strip(), "%Y-%m-%d").date()
                            except:
                                hire_date = datetime(1900, 1, 1).date()
                        else:
                            hire_date = datetime(1900, 1, 1).date()
                        
                        try:
                            cursor.execute("""
                                INSERT INTO employees (department, position, name, hire_date, display_order)
                                VALUES (?, ?, ?, ?, ?)
                            """, (dept, pos, name, hire_date, current_order))
                            current_order += 1
                            added_count += 1
                        except:
                            pass
                
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, "성공", f"{added_count}명의 직원이 추가되었습니다.")
                dialog.accept()
                self.refresh_data()
                if self.leave_gui:
                    self.leave_gui.refresh_data()
                if self.attendance_gui:
                    self.attendance_gui.refresh_data()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"일괄 등록 중 오류 발생: {str(e)}")


class DiagonalLineDelegate(QStyledItemDelegate):
    """사선을 그리는 커스텀 델리게이트"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.target_cells = set()  # 사선을 그릴 셀의 (row, col) 튜플 집합
    
    def add_target_cell(self, row, col):
        """사선을 그릴 셀 추가"""
        self.target_cells.add((row, col))
    
    def paint(self, painter, option, index):
        # 기본 그리기
        super().paint(painter, option, index)
        
        # 해당 셀이 사선을 그릴 대상인지 확인
        row = index.row()
        col = index.column()
        if (row, col) in self.target_cells:
            # 사선 그리기
            rect = option.rect
            pen = QPen(QColor("#808080"), 1)  # 회색 사선
            painter.setPen(pen)
            # 왼쪽 위에서 오른쪽 아래로 사선
            painter.drawLine(rect.topLeft(), rect.bottomRight())


class EditableTableWidget(QTableWidget):
    """복사/붙여넣기 기능이 있는 커스텀 테이블 위젯"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_gui = None  # AttendanceManagementGUI 참조
    
    def set_parent_gui(self, parent_gui):
        """부모 GUI 참조 설정"""
        self.parent_gui = parent_gui
    
    def keyPressEvent(self, event: "QKeyEvent"):
        """키보드 이벤트 처리 - Ctrl+C, Ctrl+V, Delete 지원"""
        # Ctrl+C: 복사
        if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_C:
            event.accept()
            self.copy_selected_cells()
            return
        
        # Ctrl+V: 붙여넣기
        if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_V:
            event.accept()
            self.paste_to_selected_cells()
            return
        
        # Delete: 선택된 셀 삭제
        if event.key() == Qt.Key_Delete:
            event.accept()
            self.delete_selected_cells()
            return
        
        # 기본 동작 실행
        super().keyPressEvent(event)
    
    def copy_selected_cells(self):
        """선택된 셀들을 클립보드에 복사"""
        # 선택된 셀 범위 가져오기
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            # 범위가 없으면 선택된 아이템 확인
            selected_items = self.selectedItems()
            if not selected_items:
                return
            
            # 날짜 컬럼만 복사 (3번째부터, 요약 컬럼 전까지)
            summary_start = 34
            valid_items = []
            for item in selected_items:
                col = item.column()
                if col >= 3 and col < summary_start:
                    valid_items.append(item)
            
            if not valid_items:
                return
            
            # 텍스트를 탭과 줄바꿈으로 구분하여 복사 (엑셀 형식)
            texts = []
            current_row = None
            row_texts = []
            
            for item in sorted(valid_items, key=lambda x: (x.row(), x.column())):
                if current_row is None or current_row != item.row():
                    if row_texts:
                        texts.append('\t'.join(row_texts))
                    row_texts = []
                    current_row = item.row()
                row_texts.append(item.text())
            
            if row_texts:
                texts.append('\t'.join(row_texts))
            
            clipboard_text = '\n'.join(texts)
        else:
            # 선택된 범위에서 복사
            summary_start = 34
            all_texts = []
            
            for range_obj in selected_ranges:
                row_texts = []
                for row in range(range_obj.topRow(), range_obj.bottomRow() + 1):
                    cols = []
                    for col in range(range_obj.leftColumn(), range_obj.rightColumn() + 1):
                        if col >= 3 and col < summary_start:
                            item = self.item(row, col)
                            if item:
                                cols.append(item.text())
                            else:
                                cols.append("")
                    if cols:
                        row_texts.append('\t'.join(cols))
                
                if row_texts:
                    all_texts.extend(row_texts)
            
            clipboard_text = '\n'.join(all_texts)
        
        if clipboard_text:
            clipboard = QApplication.clipboard()
            clipboard.setText(clipboard_text)
    
    def paste_to_selected_cells(self):
        """클립보드의 텍스트를 선택된 셀에 붙여넣기 (엑셀 방식 지원)"""
        if not self.parent_gui:
            return
        
        clipboard = QApplication.clipboard()
        clipboard_text = clipboard.text()
        
        if not clipboard_text:
            return
        
        # 클립보드 텍스트를 파싱 (엑셀 형식: 탭으로 열 구분, 줄바꿈으로 행 구분)
        # \r\n, \n, \r 모두 처리
        lines = clipboard_text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        
        # 각 줄을 파싱하여 2차원 배열로 변환
        clipboard_data = []
        for line in lines:
            # 빈 줄은 건너뛰지만, 탭만 있는 줄(빈 행)은 빈 배열로 추가
            if '\t' in line:
                # 탭으로 구분된 값들 (빈 셀은 빈 문자열로 처리)
                values = line.split('\t')
                # 각 값의 앞뒤 공백 제거
                cleaned_values = [v.strip() for v in values]
                clipboard_data.append(cleaned_values)
            elif line.strip():
                # 탭이 없는 단일 값
                clipboard_data.append([line.strip()])
            # 완전히 빈 줄은 무시
        
        if not clipboard_data:
            return
        
        # 클립보드 데이터 크기
        clipboard_rows = len(clipboard_data)
        clipboard_cols = len(clipboard_data[0]) if clipboard_data else 0
        
        summary_start = 34
        
        # 선택된 범위 또는 셀 가져오기
        selected_ranges = self.selectedRanges()
        
        if selected_ranges:
            # 모든 선택된 범위에 대해 붙여넣기 수행
            for range_obj in selected_ranges:
                start_row = range_obj.topRow()
                start_col = range_obj.leftColumn()
                end_row = range_obj.bottomRow()
                end_col = range_obj.rightColumn()
                
                # 선택된 범위 크기 계산
                selected_rows = end_row - start_row + 1
                selected_cols = end_col - start_col + 1
                
                # 클립보드 데이터 크기와 선택된 범위 크기가 정확히 같은 경우
                if clipboard_rows == selected_rows and clipboard_cols == selected_cols:
                    # 정확히 1:1 매핑하여 붙여넣기
                    for row_idx in range(clipboard_rows):
                        for col_idx in range(clipboard_cols):
                            target_row = start_row + row_idx
                            target_col = start_col + col_idx
                            
                            # 날짜 컬럼만 붙여넣기 가능
                            if target_col >= 3 and target_col < summary_start:
                                item = self.item(target_row, target_col)
                                if item:
                                    value = clipboard_data[row_idx][col_idx]
                                    # 빈 값도 처리 (셀 삭제를 위해)
                                    if value:
                                        item.setText(value)
                                    else:
                                        item.setText("")
                else:
                    # 크기가 다른 경우 반복하여 붙여넣기 (엑셀 방식)
                    for row in range(start_row, end_row + 1):
                        for col in range(start_col, end_col + 1):
                            # 날짜 컬럼만 붙여넣기 가능
                            if col >= 3 and col < summary_start:
                                item = self.item(row, col)
                                if item:
                                    # 클립보드 데이터에서 해당 위치의 값을 가져오기 (반복)
                                    clipboard_row_idx = (row - start_row) % clipboard_rows
                                    clipboard_col_idx = (col - start_col) % clipboard_cols
                                    
                                    value = clipboard_data[clipboard_row_idx][clipboard_col_idx]
                                    # 빈 값도 처리 (셀 삭제를 위해)
                                    if value:
                                        item.setText(value)
                                    else:
                                        item.setText("")
        else:
            # 범위가 없으면 선택된 아이템 또는 현재 셀에 붙여넣기
            selected_items = self.selectedItems()
            current_row = self.currentRow()
            current_col = self.currentColumn()
            
            # 선택된 셀이 없고 현재 셀이 날짜 컬럼인 경우, 현재 셀부터 붙여넣기
            if not selected_items and current_row >= 0 and current_col >= 3 and current_col < summary_start:
                # 현재 셀부터 시작하여 클립보드 데이터를 붙여넣기
                start_row = current_row
                start_col = current_col
                
                for row_idx, clipboard_row in enumerate(clipboard_data):
                    target_row = start_row + row_idx
                    
                    # 테이블 범위를 벗어나면 중단
                    if target_row >= self.rowCount():
                        break
                    
                    for col_idx, value in enumerate(clipboard_row):
                        target_col = start_col + col_idx
                        
                        # 날짜 컬럼 범위를 벗어나면 다음 행으로
                        if target_col >= summary_start:
                            break
                        
                        # 날짜 컬럼만 붙여넣기 가능
                        if target_col >= 3:
                            item = self.item(target_row, target_col)
                            if item:
                                # 빈 값도 처리 (셀 삭제를 위해)
                                if value:
                                    # item.setText()만 호출하면 itemChanged 시그널이 자동으로 발생하여 
                                    # on_cell_changed가 호출됨 (직접 호출 불필요)
                                    item.setText(value)
                                else:
                                    item.setText("")
                return
            
            # 선택된 아이템이 있는 경우
            if not selected_items:
                return
            
            # 날짜 컬럼만 붙여넣기 가능
            valid_items = []
            for item in selected_items:
                col = item.column()
                if col >= 3 and col < summary_start:
                    valid_items.append(item)
            
            if not valid_items:
                return
            
            # 선택된 셀에 순서대로 붙여넣기 (클립보드 데이터 반복)
            sorted_items = sorted(valid_items, key=lambda x: (x.row(), x.column()))
            
            # 클립보드 데이터를 1차원 리스트로 변환
            flat_values = []
            for row in clipboard_data:
                flat_values.extend(row)
            
            if not flat_values:
                return
            
            # 선택된 셀에 순서대로 붙여넣기 (반복)
            for i, item in enumerate(sorted_items):
                value_idx = i % len(flat_values)
                value = flat_values[value_idx]
                # 빈 값도 처리 (셀 삭제를 위해)
                if value:
                    # item.setText()만 호출하면 itemChanged 시그널이 자동으로 발생하여 
                    # on_cell_changed가 호출됨 (직접 호출 불필요)
                    item.setText(value)
                else:
                    item.setText("")
    
    def delete_selected_cells(self):
        """선택된 셀들의 내용 삭제"""
        if not self.parent_gui:
            return
        
        summary_start = 34  # 3(기본) + 31(날짜) = 34
        
        # 선택된 범위 또는 셀 가져오기
        selected_ranges = self.selectedRanges()
        
        if selected_ranges:
            # 선택된 범위의 모든 셀 삭제
            for range_obj in selected_ranges:
                for row in range(range_obj.topRow(), range_obj.bottomRow() + 1):
                    for col in range(range_obj.leftColumn(), range_obj.rightColumn() + 1):
                        # 날짜 컬럼만 삭제 가능
                        if col >= 3 and col < summary_start:
                            item = self.item(row, col)
                            if item:
                                # 빈 문자열로 설정하면 on_cell_changed가 호출되어 데이터베이스에서도 삭제됨
                                item.setText("")
        else:
            # 선택된 아이템이 있으면 삭제
            selected_items = self.selectedItems()
            if not selected_items:
                # 선택된 아이템이 없으면 현재 셀 삭제
                current_row = self.currentRow()
                current_col = self.currentColumn()
                if current_row >= 0 and current_col >= 3 and current_col < summary_start:
                    item = self.item(current_row, current_col)
                    if item:
                        item.setText("")
                return
            
            # 선택된 셀들 삭제
            for item in selected_items:
                col = item.column()
                if col >= 3 and col < summary_start:
                    # 빈 문자열로 설정하면 on_cell_changed가 호출되어 데이터베이스에서도 삭제됨
                    item.setText("")


class AttendanceManagementGUI(QWidget):
    """출퇴근 관리 GUI - QTableWidget 사용, 셀별 스타일링 지원"""
    
    @staticmethod
    def is_third_wednesday_17_00(work_date, departure_time):
        """
        매월 셋째주 수요일 17:00 퇴근인지 확인
        Args:
            work_date: 날짜 (date 객체)
            departure_time: 퇴근시간 (time 객체)
        Returns:
            bool: 셋째주 수요일이고 17:00 퇴근이면 True
        """
        # 수요일인지 확인 (weekday() == 2가 수요일)
        if work_date.weekday() != 2:  # 수요일이 아니면 False
            return False
        
        # 셋째주 수요일인지 확인 (15일~21일 사이)
        if not (15 <= work_date.day <= 21):
            return False
        
        # 퇴근시간이 17:00인지 확인
        if departure_time and departure_time.hour == 17 and departure_time.minute == 0:
            return True
        
        return False
    
    def __init__(self, parent, db_manager, attendance_calculator, leave_gui=None, employee_gui=None):
        super().__init__(parent)
        self.db = db_manager
        self.calculator = attendance_calculator
        self.leave_gui = leave_gui  # 연월차 관리대장 참조
        self.employee_gui = employee_gui  # 재직인원 탭 참조
        self._is_refreshing = False  # 데이터 새로고침 중 플래그
        
        layout = QVBoxLayout(self)
        
        button_layout = QHBoxLayout()
        button_layout.addWidget(QPushButton("엑셀 업로드", clicked=self.upload_excel))
        # 연월차 + 출퇴근을 한 파일로 묶어서 다운로드
        button_layout.addWidget(QPushButton("엑셀 다운로드", clicked=self.download_combined_excel))
        button_layout.addWidget(QPushButton("출퇴근 등록", clicked=self.register_attendance))
        button_layout.addWidget(QPushButton("연차/반차 동기화", clicked=self.sync_leave_records))
        button_layout.addWidget(QPushButton("데이터 전체 삭제", clicked=self.delete_all_data))
        self.save_button = QPushButton("저장", clicked=self.save_changes)
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(QPushButton("새로고침", clicked=self.refresh_data))
        button_layout.addStretch()
        # 재직인원 수 표시 레이블
        self.employee_count_label = QLabel("재직인원: 0명")
        self.employee_count_label.setStyleSheet("font-weight: bold; color: #0066CC;")
        button_layout.addWidget(self.employee_count_label)
        layout.addLayout(button_layout)
        
        # Ctrl+S 단축키 설정
        from PySide6.QtGui import QShortcut, QKeySequence
        save_shortcut = QShortcut(QKeySequence("Ctrl+S"), self)
        save_shortcut.activated.connect(self.save_changes)
        
        # 편집된 데이터를 임시 저장할 딕셔너리
        self.pending_changes = {}
        
        month_layout = QHBoxLayout()
        month_layout.addWidget(QLabel("조회 기간:"))
        
        # 년도 선택 드롭다운
        self.year_combo = QComboBox()
        current_year = datetime.now().year
        # 2028년부터 2020년까지 추가
        max_year = max(current_year, 2028)
        for year in range(max_year, 2019, -1):
            self.year_combo.addItem(str(year), year)
        self.year_combo.setCurrentText(str(current_year))
        self.year_combo.setMaxVisibleItems(20)  # 드롭다운 열었을 때 모든 년도가 보이도록 설정
        self.year_combo.currentIndexChanged.connect(self.refresh_data)  # 년도 변경 시 자동 새로고침
        month_layout.addWidget(QLabel("년"))
        month_layout.addWidget(self.year_combo)
        
        # 월 선택 드롭다운
        self.month_combo = QComboBox()
        for month in range(1, 13):
            self.month_combo.addItem(f"{month}월", month)
        current_month = datetime.now().month
        self.month_combo.setCurrentIndex(current_month - 1)  # 현재 월로 설정 (인덱스는 0부터 시작)
        self.month_combo.setMinimumWidth(80)  # 드롭다운 너비 조정
        self.month_combo.setMaxVisibleItems(12)  # 드롭다운 열었을 때 1월부터 12월까지 모두 보이게
        self.month_combo.currentIndexChanged.connect(self.refresh_data)  # 월 변경 시 자동 새로고침
        month_layout.addWidget(self.month_combo)
        month_layout.addStretch()
        layout.addLayout(month_layout)
        
        # 퇴사자 표시 옵션 (재직인원 탭의 체크박스와 동기화)
        option_layout = QHBoxLayout()
        self.show_inactive_checkbox = QCheckBox("퇴사자 표시")
        if self.employee_gui:
            # 재직인원 탭의 체크박스 상태와 동기화
            self.show_inactive_checkbox.setChecked(self.employee_gui.show_inactive_checkbox.isChecked())
            # 재직인원 탭의 체크박스 상태 변경 시 이 탭의 체크박스도 업데이트 (비동기로 처리)
            def sync_from_employee():
                from PySide6.QtCore import QTimer
                self.show_inactive_checkbox.blockSignals(True)
                self.show_inactive_checkbox.setChecked(self.employee_gui.show_inactive_checkbox.isChecked())
                self.show_inactive_checkbox.blockSignals(False)
                # 비동기로 새로고침하여 데이터베이스 충돌 방지
                QTimer.singleShot(50, lambda: self.refresh_data())
            self.employee_gui.show_inactive_checkbox.stateChanged.connect(sync_from_employee)
            # 이 탭의 체크박스 상태 변경 시 재직인원 탭의 체크박스도 업데이트 (비동기로 처리)
            def sync_to_employee():
                from PySide6.QtCore import QTimer
                # 재직인원 탭의 체크박스 업데이트 (무한 루프 방지)
                self.employee_gui.show_inactive_checkbox.blockSignals(True)
                self.employee_gui.show_inactive_checkbox.setChecked(self.show_inactive_checkbox.isChecked())
                self.employee_gui.show_inactive_checkbox.blockSignals(False)
                # 재직인원 탭의 체크박스 변경 핸들러가 다른 탭들도 업데이트하므로 여기서는 refresh만 호출 (비동기)
                QTimer.singleShot(50, lambda: self.employee_gui.refresh_data() if self.employee_gui else None)
                # 다른 탭의 체크박스도 동기화 (비동기)
                if self.employee_gui.leave_gui:
                    self.employee_gui.leave_gui.show_inactive_checkbox.blockSignals(True)
                    self.employee_gui.leave_gui.show_inactive_checkbox.setChecked(self.show_inactive_checkbox.isChecked())
                    self.employee_gui.leave_gui.show_inactive_checkbox.blockSignals(False)
                    QTimer.singleShot(100, lambda: self.employee_gui.leave_gui.refresh_data() if self.employee_gui and self.employee_gui.leave_gui else None)
                if self.employee_gui.attendance_gui and self.employee_gui.attendance_gui != self:
                    self.employee_gui.attendance_gui.show_inactive_checkbox.blockSignals(True)
                    self.employee_gui.attendance_gui.show_inactive_checkbox.setChecked(self.show_inactive_checkbox.isChecked())
                    self.employee_gui.attendance_gui.show_inactive_checkbox.blockSignals(False)
                    QTimer.singleShot(150, lambda: self.employee_gui.attendance_gui.refresh_data() if self.employee_gui and self.employee_gui.attendance_gui else None)
                # 이 탭도 비동기로 새로고침
                QTimer.singleShot(0, lambda: self.refresh_data())
            self.show_inactive_checkbox.stateChanged.connect(sync_to_employee)
        else:
            # 재직인원 탭이 없으면 독립적으로 동작
            self.show_inactive_checkbox.setChecked(False)
            self.show_inactive_checkbox.stateChanged.connect(self.refresh_data)
        option_layout.addWidget(self.show_inactive_checkbox)
        option_layout.addStretch()
        layout.addLayout(option_layout)
        
        # 테이블 - 날짜별 컬럼 구조
        # 날짜 컬럼은 나중에 refresh_data에서 요일 정보와 함께 설정
        date_columns = [str(i) for i in range(1, 32)]
        summary_columns = ["조기출근\n(8시이전)", "지각\n(9시이후)", "야근\n(20시이후)", "연차사용", "평균 출근시간", "평균 퇴근시간"]
        columns = ["직급", "이름", "구분"] + date_columns + summary_columns
        
        self.table = EditableTableWidget()
        self.table.set_parent_gui(self)  # 부모 GUI 참조 설정
        self.table.setColumnCount(len(columns))
        # 헤더는 refresh_data에서 요일 정보와 함께 설정
        self.table.setHorizontalHeaderLabels(columns)
        self.table.horizontalHeader().setStretchLastSection(True)
        # 헤더 높이 조정 (요일 정보 표시를 위해)
        self.table.horizontalHeader().setMinimumHeight(50)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)  # 셀 단위 선택
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)  # CTRL/SHIFT로 다중 선택
        # 날짜 컬럼(3번째부터)은 직접 편집 가능, 나머지는 편집 불가
        # DoubleClicked만 사용하여 여러 셀 선택 가능하도록 함
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked)
        
        # 셀 편집 완료 시 이벤트 연결
        self.table.itemChanged.connect(self.on_cell_changed)
        
        # 컬럼 너비 설정
        # 직급, 이름, 구분 컬럼
        self.table.setColumnWidth(0, 60)  # 직급
        self.table.setColumnWidth(1, 70)  # 이름
        self.table.setColumnWidth(2, 60)  # 구분
        
        # 날짜 컬럼 (1-31일) - 시간 형식에 맞게
        for i in range(3, 34):  # 3번부터 33번까지 (1일~31일)
            self.table.setColumnWidth(i, 45)  # HH:MM 형식에 맞는 너비 (3글자 텍스트 입력 가능)
        
        # 요약 컬럼들
        summary_start = 34  # 3(기본) + 31(날짜) = 34
        self.table.setColumnWidth(summary_start, 100)      # 조기출근(8시이전)
        self.table.setColumnWidth(summary_start + 1, 100)  # 지각(9시이후)
        self.table.setColumnWidth(summary_start + 2, 100)  # 야근(20시이후)
        self.table.setColumnWidth(summary_start + 3, 80)   # 연차사용
        self.table.setColumnWidth(summary_start + 4, 110)  # 평균 출근시간
        self.table.setColumnWidth(summary_start + 5, 110)  # 평균 퇴근시간
        
        # 더블클릭 이벤트
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        
        # 테이블에 포커스 설정 (키보드 이벤트 처리를 위해)
        self.table.setFocusPolicy(Qt.StrongFocus)
        
        # verticalHeader의 행 높이를 조정하여 출근/퇴근 행 병합 효과
        # 출근 행과 퇴근 행이 하나의 번호를 공유하도록 설정
        self.table.verticalHeader().setDefaultSectionSize(25)  # 기본 행 높이 설정
        
        layout.addWidget(self.table)
        
        self.refresh_data()
    
    def save_changes(self):
        """저장 버튼 또는 Ctrl+S로 변경 사항 저장"""
        if not self.pending_changes:
            QMessageBox.information(self, "알림", "저장할 변경 사항이 없습니다.")
            return
        
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            saved_count = 0
            for change_key, change_data in self.pending_changes.items():
                emp_id, work_date, category = change_key
                new_value = change_data['new_value']
                formatted_time = change_data['formatted_time']
                is_time = change_data['is_time']
                is_delete = change_data.get('is_delete', False)
                
                # 삭제 처리 - 출근행/퇴근행 각각 처리
                if is_delete:
                    try:
                        # 기존 기록 확인
                        cursor.execute("""
                            SELECT arrival_time, departure_time, leave_type, remarks
                            FROM attendance_records
                            WHERE employee_id = ? AND work_date = ?
                        """, (emp_id, work_date))
                        existing = cursor.fetchone()
                        
                        if existing:
                            arrival_time_str = existing[0]
                            departure_time_str = existing[1]
                            leave_type = existing[2]
                            remarks = existing[3]
                            
                            # 문자열 시간을 time 객체로 변환하는 헬퍼 함수
                            def parse_time_safe(time_str):
                                if not time_str:
                                    return None
                                if isinstance(time_str, str):
                                    try:
                                        # HH:MM:SS 형식
                                        if len(time_str) >= 8:
                                            return datetime.strptime(time_str, "%H:%M:%S").time()
                                        # HH:MM 형식
                                        elif len(time_str) >= 5:
                                            return datetime.strptime(time_str, "%H:%M").time()
                                    except:
                                        pass
                                return time_str  # 이미 time 객체이거나 변환 실패 시 그대로 반환
                            
                            arrival_time = parse_time_safe(arrival_time_str)
                            departure_time = parse_time_safe(departure_time_str)
                            
                            # category에 따라 해당 시간만 삭제
                            if category == '출근':
                                # 출근행만 삭제: arrival_time을 None으로, departure_time은 유지
                                arrival_time = None
                                
                                # leave_type이 반차/연차/휴가인 경우 leave_records에서도 삭제하고 leave_type도 None으로 설정
                                # remarks가 "{type}_출근"이거나 remarks가 없는 경우(직접 입력한 연차 등)
                                if leave_type in ['반차', '연차', '휴가']:
                                    if not remarks or remarks == f'{leave_type}_출근':
                                        year = self.year_combo.currentData() or datetime.now().year
                                        month = self.month_combo.currentData() or datetime.now().month
                                        cursor.execute("""
                                            DELETE FROM leave_records
                                            WHERE employee_id = ? AND leave_date = ? AND leave_type = ?
                                        """, (emp_id, work_date, leave_type))
                                        # leave_type도 None으로 설정 (출근행 삭제 시)
                                        leave_type = None
                                        remarks = None
                                
                                # departure_time이 없고 leave_type도 없으면 전체 삭제
                                if not departure_time and not leave_type:
                                    cursor.execute("""
                                        DELETE FROM attendance_records
                                        WHERE employee_id = ? AND work_date = ?
                                    """, (emp_id, work_date))
                                    # leave_records에서도 삭제
                                    year = self.year_combo.currentData() or datetime.now().year
                                    month = self.month_combo.currentData() or datetime.now().month
                                    cursor.execute("""
                                        DELETE FROM leave_records
                                        WHERE employee_id = ? AND leave_date = ?
                                    """, (emp_id, work_date))
                                else:
                                    # arrival_time만 None으로 업데이트 (leave_type도 함께 업데이트)
                                    self.calculator.process_attendance_record(
                                        emp_id, work_date, arrival_time, departure_time,
                                        leave_type, remarks, conn
                                    )
                            elif category == '퇴근':
                                # 퇴근행만 삭제: departure_time을 None으로, arrival_time은 유지
                                departure_time = None
                                
                                # leave_type이 반차/연차/휴가인 경우 leave_records에서도 삭제
                                # remarks가 "{type}_퇴근"이거나 remarks가 없는 경우(직접 입력한 연차 등)
                                # 단, 퇴근행만 삭제하는 경우는 leave_type을 유지 (출근행에 반차가 있을 수 있음)
                                if leave_type in ['반차', '연차', '휴가']:
                                    if remarks == f'{leave_type}_퇴근':
                                        # remarks가 "{type}_퇴근"인 경우만 leave_records에서 삭제하고 leave_type도 None으로 설정
                                        year = self.year_combo.currentData() or datetime.now().year
                                        month = self.month_combo.currentData() or datetime.now().month
                                        cursor.execute("""
                                            DELETE FROM leave_records
                                            WHERE employee_id = ? AND leave_date = ? AND leave_type = ?
                                        """, (emp_id, work_date, leave_type))
                                        # leave_type도 None으로 설정 (퇴근행만 삭제하는 경우)
                                        leave_type = None
                                        remarks = None
                                
                                # arrival_time이 없고 leave_type도 없으면 전체 삭제
                                if not arrival_time and not leave_type:
                                    cursor.execute("""
                                        DELETE FROM attendance_records
                                        WHERE employee_id = ? AND work_date = ?
                                    """, (emp_id, work_date))
                                    # leave_records에서도 삭제
                                    year = self.year_combo.currentData() or datetime.now().year
                                    month = self.month_combo.currentData() or datetime.now().month
                                    cursor.execute("""
                                        DELETE FROM leave_records
                                        WHERE employee_id = ? AND leave_date = ?
                                    """, (emp_id, work_date))
                                else:
                                    # departure_time만 None으로 업데이트 (leave_type도 함께 업데이트)
                                    self.calculator.process_attendance_record(
                                        emp_id, work_date, arrival_time, departure_time,
                                        leave_type, remarks, conn
                                    )
                            else:
                                # category가 없으면 전체 삭제
                                cursor.execute("""
                                    DELETE FROM attendance_records
                                    WHERE employee_id = ? AND work_date = ?
                                """, (emp_id, work_date))
                                # leave_records에서도 삭제
                                year = self.year_combo.currentData() or datetime.now().year
                                month = self.month_combo.currentData() or datetime.now().month
                                cursor.execute("""
                                    DELETE FROM leave_records
                                    WHERE employee_id = ? AND leave_date = ?
                                """, (emp_id, work_date))
                        else:
                            # 기존 기록이 없으면 삭제할 것도 없음
                            pass
                        
                        saved_count += 1
                        continue
                    except Exception as e:
                        raise Exception(f"데이터 삭제 중 오류: {str(e)}")
                
                # 기존 기록 확인
                cursor.execute("""
                    SELECT arrival_time, departure_time, leave_type, remarks
                    FROM attendance_records
                    WHERE employee_id = ? AND work_date = ?
                """, (emp_id, work_date))
                existing = cursor.fetchone()
                
                # pending_changes에서 같은 날짜의 다른 카테고리 변경사항 확인
                # 예: 출근 행에 "반차" 입력 시, pending_changes에 퇴근 시간이 있는지 확인
                pending_arrival_time = None
                pending_departure_time = None
                pending_leave_type = None
                pending_remarks = None
                for other_key, other_data in self.pending_changes.items():
                    other_emp_id, other_work_date, other_category = other_key
                    if other_emp_id == emp_id and other_work_date == work_date and other_category != category:
                        # 같은 직원, 같은 날짜, 다른 카테고리
                        if other_data.get('is_time') and not other_data.get('is_delete', False):
                            try:
                                other_formatted_time = other_data.get('formatted_time', '')
                                if ':' in other_formatted_time:
                                    other_time_obj = datetime.strptime(other_formatted_time, "%H:%M").time()
                                    if other_category == '출근':
                                        pending_arrival_time = other_time_obj
                                    elif other_category == '퇴근':
                                        pending_departure_time = other_time_obj
                            except:
                                pass
                        elif not other_data.get('is_time') and not other_data.get('is_delete', False):
                            # 텍스트 입력 (반차, 연차 등)
                            other_new_value = other_data.get('new_value', '').strip()
                            # "연차", "경조사", "예비군", "설날", "추석", "박람회", "출장"을 제외한 나머지는 독립적으로 처리
                            merge_texts = ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']
                            if other_category == '출근' and other_new_value not in merge_texts:
                                pending_leave_type = other_new_value
                                pending_remarks = f"{other_new_value}_출근"
                            elif other_category == '퇴근' and other_new_value not in merge_texts:
                                pending_leave_type = other_new_value
                                pending_remarks = f"{other_new_value}_퇴근"
                
                if is_time:
                    # 시간 형식인 경우
                    try:
                        time_obj = datetime.strptime(formatted_time, "%H:%M").time()
                    except:
                        is_time = False
                
                if is_time:
                    # 시간 형식인 경우
                    if category == '출근':
                        arrival_time = time_obj
                        # pending_changes에서 퇴근 시간 확인 (같은 날짜의 퇴근 행 변경사항)
                        departure_time = pending_departure_time
                        # pending_changes에 없으면 기존 데이터베이스에서 조회
                        if departure_time is None and existing and existing[1]:
                            try:
                                departure_time = datetime.strptime(existing[1], "%H:%M:%S").time()
                            except:
                                pass
                        
                        # 출근 행에 시간을 입력할 때는 퇴근 행의 기존 leave_type과 remarks를 보존
                        # 기존 데이터에서 퇴근 행의 정보 확인
                        leave_type_to_use = None
                        remarks_to_use = None
                        if existing and existing[2] and existing[3]:
                            # 기존 데이터베이스에서 leave_type과 remarks 확인
                            existing_leave_type = existing[2]
                            existing_remarks = existing[3]
                            # remarks가 "{text}_퇴근"인 경우만 유지 (퇴근 행에 텍스트가 있는 경우)
                            merge_texts = ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']
                            if existing_leave_type not in merge_texts and existing_remarks == f'{existing_leave_type}_퇴근':
                                leave_type_to_use = existing_leave_type
                                remarks_to_use = existing_remarks
                            # remarks가 "{text}_출근"인 경우는 무시 (출근 행에만 텍스트가 있는 경우)
                            # 병합 처리 텍스트(연차, 경조사, 예비군 등)는 remarks가 없으면 출퇴근 모두 병합이므로 유지
                            elif existing_leave_type in merge_texts and not existing_remarks:
                                leave_type_to_use = existing_leave_type
                                remarks_to_use = ""
                        elif existing and existing[2]:
                            # remarks가 없는 경우도 확인
                            existing_leave_type = existing[2]
                            merge_texts = ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']
                            # remarks가 없고 leave_type이 병합 처리 텍스트인 경우는 유지
                            if existing_leave_type in merge_texts:
                                leave_type_to_use = existing_leave_type
                                remarks_to_use = ""
                    else:  # 퇴근
                        # pending_changes에서 출근 시간 확인 (같은 날짜의 출근 행 변경사항)
                        arrival_time = pending_arrival_time
                        departure_time = time_obj
                        # pending_changes에 없으면 기존 데이터베이스에서 조회
                        if arrival_time is None and existing and existing[0]:
                            try:
                                arrival_time = datetime.strptime(existing[0], "%H:%M:%S").time()
                            except:
                                pass
                        
                        # 퇴근 행에 시간을 입력할 때는 출근 행의 기존 leave_type과 remarks를 보존
                        # 기존 데이터에서 출근 행의 정보 확인
                        leave_type_to_use = None
                        remarks_to_use = None
                        if existing and existing[2] and existing[3]:
                            # 기존 데이터베이스에서 leave_type과 remarks 확인
                            existing_leave_type = existing[2]
                            existing_remarks = existing[3]
                            merge_texts = ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']
                            # remarks가 "{text}_출근"인 경우만 유지 (출근 행에 텍스트가 있는 경우)
                            if existing_leave_type not in merge_texts and existing_remarks == f'{existing_leave_type}_출근':
                                leave_type_to_use = existing_leave_type
                                remarks_to_use = existing_remarks
                            # remarks가 "{text}_퇴근"인 경우는 무시 (퇴근 행에만 텍스트가 있는 경우)
                            # 병합 처리 텍스트(연차, 경조사, 예비군 등)는 remarks가 없으면 출퇴근 모두 병합이므로 유지
                            elif existing_leave_type in merge_texts and not existing_remarks:
                                leave_type_to_use = existing_leave_type
                                remarks_to_use = ""
                        elif existing and existing[2]:
                            # remarks가 없는 경우도 확인
                            existing_leave_type = existing[2]
                            merge_texts = ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']
                            # remarks가 없고 leave_type이 병합 처리 텍스트인 경우는 유지
                            if existing_leave_type in merge_texts:
                                leave_type_to_use = existing_leave_type
                                remarks_to_use = ""
                    
                    try:
                        self.calculator.process_attendance_record(
                            emp_id, work_date, arrival_time, departure_time,
                            leave_type_to_use, remarks_to_use if remarks_to_use is not None else "", conn
                        )
                        saved_count += 1
                    except Exception as e:
                        raise Exception(f"출퇴근 기록 처리 중 오류: {str(e)}")
                else:
                    # 텍스트 입력
                    input_text = new_value.strip()
                    
                    # "연차", "경조사", "예비군", "설날", "추석", "박람회", "출장"을 제외한 나머지 텍스트는 출근/퇴근을 독립적으로 처리
                    # 독립적으로 처리할 텍스트: 반차, 미팅, 공휴, 민방위, 교육, 휴가 등
                    # 병합 처리할 텍스트: 연차, 경조사, 예비군, 설날, 추석, 박람회, 출장
                    merge_texts = ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']
                    if input_text not in merge_texts:
                        if category == '출근':
                            arrival_time = None
                            # pending_changes에서 퇴근 시간 확인 (같은 날짜의 퇴근 행 변경사항)
                            departure_time = pending_departure_time
                            # pending_changes에 없으면 기존 데이터베이스에서 조회
                            if departure_time is None and existing and existing[1]:
                                try:
                                    departure_time = datetime.strptime(existing[1], "%H:%M:%S").time()
                                except:
                                    pass
                            
                            # 출근 행에 텍스트를 입력할 때는 퇴근 행의 기존 leave_type과 remarks를 보존
                            # 기존 데이터에서 퇴근 행의 정보 확인
                            leave_type_to_use = input_text
                            remarks_to_use = f"{input_text}_출근"
                            if existing and existing[2] and existing[3]:
                                existing_leave_type = existing[2]
                                existing_remarks = existing[3]
                                # 기존에 퇴근 행에 같은 텍스트가 있는 경우 (remarks가 "{text}_퇴근"인 경우)
                                if existing_leave_type == input_text and existing_remarks == f'{input_text}_퇴근':
                                    # 출근 행과 퇴근 행 모두 같은 텍스트이므로 leave_type은 유지
                                    # remarks는 "{text}_출근"으로 설정 (출근 행에 텍스트가 있으므로)
                                    leave_type_to_use = input_text
                                    remarks_to_use = f"{input_text}_출근"
                                # 기존에 병합 처리 텍스트(연차, 경조사, 예비군, 설날, 추석, 박람회, 출장)가 있는 경우 (remarks가 없는 경우)
                                elif existing_leave_type in ['연차', '경조사', '예비군'] and not existing_remarks:
                                    # 출근 행에 독립 텍스트를 입력하면 병합 텍스트는 무시되고 독립 텍스트로 변경
                                    leave_type_to_use = input_text
                                    remarks_to_use = f"{input_text}_출근"
                            
                            try:
                                self.calculator.process_attendance_record(
                                    emp_id, work_date, arrival_time, departure_time,
                                    leave_type_to_use, remarks_to_use, conn
                                )
                                saved_count += 1
                            except Exception as e:
                                raise Exception(f"출퇴근 기록 처리 중 오류: {str(e)}")
                            # leave_records에도 기록 (반차만)
                            if input_text == '반차':
                                year = self.year_combo.currentData() or datetime.now().year
                                month = self.month_combo.currentData() or datetime.now().month
                                cursor.execute("""
                                    INSERT OR REPLACE INTO leave_records
                                    (employee_id, leave_type, leave_date, leave_amount, year, month)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                """, (emp_id, input_text, work_date, 0.5, year, month))
                        else:  # 퇴근
                            # pending_changes에서 출근 시간 확인 (같은 날짜의 출근 행 변경사항)
                            arrival_time = pending_arrival_time
                            departure_time = None
                            # pending_changes에 없으면 기존 데이터베이스에서 조회
                            if arrival_time is None and existing and existing[0]:
                                try:
                                    arrival_time = datetime.strptime(existing[0], "%H:%M:%S").time()
                                except:
                                    pass
                            
                            # 퇴근 행에 텍스트를 입력할 때는 출근 행의 기존 leave_type과 remarks를 보존
                            # 기존 데이터에서 출근 행의 정보 확인
                            leave_type_to_use = input_text
                            remarks_to_use = f"{input_text}_퇴근"
                            if existing and existing[2] and existing[3]:
                                existing_leave_type = existing[2]
                                existing_remarks = existing[3]
                                # 기존에 출근 행에 같은 텍스트가 있는 경우 (remarks가 "{text}_출근"인 경우)
                                if existing_leave_type == input_text and existing_remarks == f'{input_text}_출근':
                                    # 출근 행과 퇴근 행 모두 같은 텍스트이므로 leave_type은 유지
                                    # remarks는 "{text}_퇴근"으로 설정 (퇴근 행에 텍스트가 있으므로)
                                    leave_type_to_use = input_text
                                    remarks_to_use = f"{input_text}_퇴근"
                                # 기존에 병합 처리 텍스트(연차, 경조사, 예비군, 설날, 추석, 박람회, 출장)가 있는 경우 (remarks가 없는 경우)
                                elif existing_leave_type in ['연차', '경조사', '예비군'] and not existing_remarks:
                                    # 퇴근 행에 독립 텍스트를 입력하면 병합 텍스트는 무시되고 독립 텍스트로 변경
                                    leave_type_to_use = input_text
                                    remarks_to_use = f"{input_text}_퇴근"
                            
                            try:
                                self.calculator.process_attendance_record(
                                    emp_id, work_date, arrival_time, departure_time,
                                    leave_type_to_use, remarks_to_use, conn
                                )
                                saved_count += 1
                            except Exception as e:
                                raise Exception(f"출퇴근 기록 처리 중 오류: {str(e)}")
                            # leave_records에도 기록 (반차만)
                            if input_text == '반차':
                                year = self.year_combo.currentData() or datetime.now().year
                                month = self.month_combo.currentData() or datetime.now().month
                                cursor.execute("""
                                    INSERT OR REPLACE INTO leave_records
                                    (employee_id, leave_type, leave_date, leave_amount, year, month)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                """, (emp_id, input_text, work_date, 0.5, year, month))
                    else:
                        # "연차", "경조사", "예비군", "설날", "추석", "박람회", "출장"은 출퇴근 병합 처리
                        try:
                            self.calculator.process_attendance_record(
                                emp_id, work_date, None, None,
                                input_text, "", conn
                            )
                            saved_count += 1
                        except Exception as e:
                            raise Exception(f"출퇴근 기록 처리 중 오류: {str(e)}")
                        # 연차나 휴가인 경우 leave_records에도 기록
                        if input_text in ['연차', '휴가']:
                            year = self.year_combo.currentData() or datetime.now().year
                            month = self.month_combo.currentData() or datetime.now().month
                            leave_amount = 1.0
                            cursor.execute("""
                                INSERT OR REPLACE INTO leave_records
                                (employee_id, leave_type, leave_date, leave_amount, year, month)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (emp_id, input_text, work_date, leave_amount, year, month))
                        elif input_text == '반차':
                            year = self.year_combo.currentData() or datetime.now().year
                            month = self.month_combo.currentData() or datetime.now().month
                            leave_amount = 0.5
                            cursor.execute("""
                                INSERT OR REPLACE INTO leave_records
                                (employee_id, leave_type, leave_date, leave_amount, year, month)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (emp_id, input_text, work_date, leave_amount, year, month))
            
            conn.commit()
            conn.close()
            
            # 저장 완료 후 pending_changes 초기화
            self.pending_changes = {}
            
            # 데이터 새로고침을 비동기로 처리하여 데이터베이스 잠금 해제 대기
            from PySide6.QtCore import QTimer
            QTimer.singleShot(100, lambda: self.refresh_data())
            
            # 연월차 관리대장도 새로고침 (비동기)
            if self.leave_gui:
                QTimer.singleShot(150, lambda: self.leave_gui.refresh_data() if self.leave_gui else None)
            
            QMessageBox.information(self, "저장 완료", f"{saved_count}건의 변경 사항이 저장되었습니다.")
        except Exception as e:
            # 오류 발생 시에도 연결 닫기
            try:
                if 'conn' in locals():
                    conn.close()
            except:
                pass
            QMessageBox.warning(self, "오류", f"데이터 저장 중 오류가 발생했습니다.\n{str(e)}")
    
    def refresh_data(self):
        """데이터 새로고침 - QTableWidget 사용, 빈 셀 음영 처리"""
        try:
            self._is_refreshing = True  # 새로고침 시작
            self.table.setRowCount(0)
            
            # 드롭다운에서 년도와 월 가져오기
            year = self.year_combo.currentData()
            month = self.month_combo.currentData()
            
            if year is None:
                year = datetime.now().year
            if month is None:
                month = datetime.now().month
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            try:
                # 월별 조회
                from calendar import monthrange
                days_in_month = monthrange(year, month)[1]
                
                # 31일 이후의 컬럼 숨기기 (월별 조회 시에만)
                summary_start = 34  # 3(기본) + 31(날짜) = 34
                for day in range(days_in_month + 1, 32):  # days_in_month + 1부터 31까지
                    col = 2 + day  # 3번 컬럼부터 시작하므로 2 + day
                    if col < summary_start:
                        self.table.setColumnHidden(col, True)
                
                # 해당 월에 있는 날짜 컬럼은 보이도록 설정
                for day in range(1, days_in_month + 1):
                    col = 2 + day
                    self.table.setColumnHidden(col, False)
                
                self._refresh_month_data(conn, cursor, year, month, days_in_month, is_year_mode=False)
                conn.close()
            except Exception as e:
                # 데이터베이스 작업 중 예외 발생 시 연결 닫기
                try:
                    conn.close()
                except:
                    pass
                # 예외를 상위로 전파하지 않고 내부에서 처리
                error_msg = str(e)
                # 외부 except 블록에서 처리하도록 예외를 다시 발생시킴
                # 하지만 외부 except에서 잡히므로 프로그램이 종료되지 않음
                raise Exception(f"데이터베이스 작업 중 오류: {error_msg}") from e
        except Exception as e:
            # 예외를 잡아서 처리하고, 상위로 전파하지 않음 (프로그램 종료 방지)
            # 새로고침 중 예외 발생 시 사용자에게 알림 (안전하게 처리)
            try:
                # 위젯이 유효한지 확인 후 메시지 박스 표시
                if self and hasattr(self, 'isVisible'):
                    try:
                        QMessageBox.warning(self, "오류", f"데이터 새로고침 중 오류가 발생했습니다.\n{str(e)}")
                    except:
                        # QMessageBox 호출 실패 시 콘솔에 출력
                        print(f"데이터 새로고침 중 오류 발생: {str(e)}")
                else:
                    print(f"데이터 새로고침 중 오류 발생: {str(e)}")
            except Exception:
                # 모든 예외를 잡아서 프로그램이 종료되지 않도록 함
                print(f"데이터 새로고침 중 오류 발생: {str(e)}")
        finally:
            # 항상 플래그 해제 (예외 발생 여부와 관계없이)
            try:
                self._is_refreshing = False  # 새로고침 완료 (항상 실행)
            except:
                pass
    
    def _refresh_month_data(self, conn, cursor, year, month, days_in_month=None, is_year_mode=False):
        """특정 월의 데이터를 새로고침"""
        if days_in_month is None:
            from calendar import monthrange
            days_in_month = monthrange(year, month)[1]
        
        # 입사일 이전 날짜 셀에 사선을 그리기 위한 델리게이트 (전체 직원 공통)
        diagonal_delegate = DiagonalLineDelegate()
        
        # 날짜 컬럼 헤더에 요일 정보 추가 (월별 조회 시에만)
        if not is_year_mode:
            weekday_names = ['월', '화', '수', '목', '금', '토', '일']
            for day in range(1, days_in_month + 1):
                try:
                    date_obj = datetime(year, month, day).date()
                    weekday = date_obj.weekday()  # 0=월요일, 6=일요일
                    weekday_name = weekday_names[weekday]
                    col = 2 + day  # 3번 컬럼부터 시작 (인덱스는 2부터)
                    
                    # 헤더 아이템 생성 (날짜\n요일 형식)
                    header_item = QTableWidgetItem(f"{day}\n{weekday_name}")
                    header_item.setTextAlignment(Qt.AlignCenter)
                    # 헤더 폰트 설정 (작게 하여 두 줄이 잘 보이도록)
                    header_font = QFont()
                    header_font.setPointSize(8)
                    header_item.setFont(header_font)
                    
                    # 토요일/일요일 색상 설정
                    if weekday == 5:  # 토요일
                        header_item.setForeground(QColor("#0000FF"))  # 파란색
                    elif weekday == 6:  # 일요일
                        header_item.setForeground(QColor("#FF0000"))  # 빨간색
                    
                    # 헤더에 설정
                    self.table.setHorizontalHeaderItem(col, header_item)
                except ValueError:
                    pass  # 유효하지 않은 날짜는 무시
        
        # 년도별 조회인 경우 월별 헤더 추가
        if is_year_mode:
            month_header_row = self.table.rowCount()
            self.table.insertRow(month_header_row)
            month_header = QTableWidgetItem(f"━━━ {year}년 {month}월 ━━━")
            month_header.setFlags(Qt.NoItemFlags)
            month_header.setBackground(QColor("#D0D0D0"))
            month_header.setFont(QFont("Arial", 10, QFont.Bold))
            self.table.setItem(month_header_row, 0, month_header)
            self.table.setSpan(month_header_row, 0, 1, len(self.table.horizontalHeaderLabels()))
        
        cursor.execute("""
            SELECT e.id, e.department, e.position, e.name, e.hire_date,
                   COALESCE(e.display_order, 0) as display_order,
                   COALESCE(e.is_active, 1) as is_active,
                   e.resignation_date
            FROM employees e
            ORDER BY
                CASE e.department
                    WHEN '경영지원팀' THEN 1
                    WHEN '영업팀' THEN 2
                    WHEN '글로벌비즈니스팀' THEN 3
                    ELSE 999
                END,
                e.department,
                CASE e.position
                    WHEN '이사' THEN 1
                    WHEN '팀장' THEN 2
                    WHEN '파트장' THEN 3
                    WHEN '과장' THEN 4
                    WHEN '대리' THEN 5
                    WHEN '프로' THEN 6
                    ELSE 999
                END,
                e.hire_date ASC
        """)
        all_employees = cursor.fetchall()
        
        # 퇴사자 표시 옵션 확인
        show_inactive = self.show_inactive_checkbox.isChecked()
        
        # 퇴사일을 고려한 필터링: 모든 직원을 포함하되, is_active 정보도 함께 저장
        employees = []
        for emp in all_employees:
            emp_id, dept, pos, name, hire_date, display_order, is_active, resignation_date = emp
            
            # resignation_date가 문자열인 경우 date 객체로 변환
            if resignation_date and isinstance(resignation_date, str):
                try:
                    resignation_date = datetime.strptime(resignation_date, "%Y-%m-%d").date()
                except:
                    resignation_date = None
            
            # 모든 직원을 포함하되, is_active와 resignation_date 정보도 함께 저장
            employees.append((emp_id, dept, pos, name, hire_date, display_order, is_active, resignation_date))
        
        current_department = None
        employee_row_number = 1  # 실제 직원 행 번호 카운터 (출근/퇴근 행을 하나로 카운트)
        separator_rows = []  # 구분자 행 추적 (부서, 행 번호)
        for emp_id, dept, pos, name, hire_date_str, display_order, is_active, resignation_date in employees:
            # 입사일 파싱
            hire_date = None
            if hire_date_str:
                try:
                    if isinstance(hire_date_str, str):
                        hire_date = datetime.strptime(hire_date_str, "%Y-%m-%d").date()
                    else:
                        hire_date = hire_date_str
                except:
                    pass
            
            # 퇴사일 파싱
            if resignation_date and isinstance(resignation_date, str):
                try:
                    resignation_date = datetime.strptime(resignation_date, "%Y-%m-%d").date()
                except:
                    resignation_date = None
            # 부서가 변경되면 구분자 추가
            if current_department != dept:
                # 구분자 행 추가 (출근 행만)
                separator_row = self.table.rowCount()
                self.table.insertRow(separator_row)
                
                # 구분자 행 추적
                separator_rows.append({'dept': dept, 'row': separator_row, 'has_visible_employee': False})
                
                # 구분자 행의 행 번호를 공란으로 설정
                self.table.setVerticalHeaderItem(separator_row, QTableWidgetItem(""))
                
                # 구분자 아이템 생성
                separator_item = QTableWidgetItem(f"━━━ {dept} ━━━")
                separator_item.setFlags(Qt.NoItemFlags)  # 선택 불가
                separator_item.setBackground(QColor("#E0E0E0"))
                separator_item.setFont(QFont("Arial", 10, QFont.Bold))
                self.table.setItem(separator_row, 0, separator_item)
                # 모든 컬럼에 걸쳐 병합
                summary_start = 3 + 31  # 3(기본) + 31(날짜)
                self.table.setSpan(separator_row, 0, 1, self.table.columnCount())
                
                current_department = dept
            cursor.execute("""
                SELECT work_date, arrival_time, departure_time,
                       early_arrival, late_arrival, late_departure, leave_type, remarks
                FROM attendance_records
                WHERE employee_id = ? AND strftime('%Y-%m', work_date) = ?
                ORDER BY work_date
            """, (emp_id, f"{year:04d}-{month:02d}"))
            
            records = cursor.fetchall()
            records_dict = {}
            for record in records:
                work_date, arrival, departure, early, late_arr, late_dep, leave_type, remarks = record
                day = int(work_date.split('-')[2])
                records_dict[day] = {
                    'arrival': arrival,
                    'departure': departure,
                    'early': early,
                    'late_arr': late_arr,
                    'late_dep': late_dep,
                    'leave_type': leave_type,
                    'remarks': remarks
                }
            
            # 출근 행 추가
            arrival_row = self.table.rowCount()
            self.table.insertRow(arrival_row)
            
            # 실제 직원 행의 행 번호 설정 (출근 행에만 번호 표시, 가운데 정렬하여 두 행의 중간에 위치)
            header_item_arrival = QTableWidgetItem(str(employee_row_number))
            header_item_arrival.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.table.setVerticalHeaderItem(arrival_row, header_item_arrival)
            
            # 직급, 이름, 구분
            pos_item = QTableWidgetItem(pos)
            pos_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(arrival_row, 0, pos_item)
            name_item = QTableWidgetItem(name)
            name_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(arrival_row, 1, name_item)
            category_item = QTableWidgetItem("출근")
            category_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(arrival_row, 2, category_item)
            
            # 날짜별 데이터 (3번 컬럼부터)
            arrival_times = []
            for day in range(1, days_in_month + 1):
                col = 2 + day  # 3번 컬럼부터 시작
                if day in records_dict:
                    record = records_dict[day]
                    # "연차", "경조사", "예비군", "설날", "추석", "박람회", "출장"을 제외한 나머지는 remarks로 출근/퇴근 구분
                    remarks = record.get('remarks', '')
                    leave_type = record.get('leave_type')
                    merge_texts = ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']
                    
                    # remarks가 "{type}_출근" 형식인 경우 (독립 처리 텍스트)
                    if leave_type and remarks == f'{leave_type}_출근':
                        # 출근 행에 텍스트 표시 → 평균 출근시간에서 제외
                        item = QTableWidgetItem(leave_type)
                        item.setTextAlignment(Qt.AlignCenter)
                        # 연차 또는 경조사 등 출근하지 않은 경우 배경색 처리
                        if leave_type in ['연차', '반차', '경조사', '공휴', '박람회', '예비군', '추석', '설날', '민방위', '출장', '교육']:
                            item.setBackground(QColor("#F8CBAD"))
                    elif leave_type and remarks == f'{leave_type}_퇴근':
                        # 퇴근 행용이므로 출근 행에는 시간 표시 → 평균 출근시간에 포함 (퇴근행에 텍스트가 있으므로)
                        if record['arrival']:
                            arrival_str = str(record['arrival'])[:5] if record['arrival'] else ""
                            item = QTableWidgetItem(arrival_str)
                            item.setTextAlignment(Qt.AlignCenter)
                            try:
                                arrival_time_obj = datetime.strptime(arrival_str, "%H:%M").time()
                                # 토요일(5)과 일요일(6)은 평균 출근시간에서 제외
                                work_date_obj = datetime(year, month, day).date()
                                weekday = work_date_obj.weekday()  # 0=월요일, 5=토요일, 6=일요일
                                if weekday != 5 and weekday != 6:  # 토요일, 일요일이 아닐 때만 추가
                                    arrival_times.append(arrival_time_obj)
                                # 08시 이전 출근 - 초록색
                                if arrival_time_obj < datetime.strptime("08:00", "%H:%M").time():
                                    item.setForeground(QColor("#008000"))
                                # 09시 이후 지각 - 빨간색
                                elif arrival_time_obj > datetime.strptime("09:00", "%H:%M").time():
                                    item.setForeground(QColor("#FF0000"))
                            except:
                                pass
                        else:
                            item = QTableWidgetItem("")  # 빈 셀
                            item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                            item.setTextAlignment(Qt.AlignCenter)
                    elif leave_type and not remarks:
                        # remarks가 없는 경우
                        # "연차", "경조사", "예비군", "설날", "추석", "박람회", "출장"은 출퇴근 병합 처리
                        # 나머지는 remarks로 출근/퇴근을 구분하므로, remarks가 없으면 시간 표시
                        if leave_type in merge_texts:
                            # 병합 처리 텍스트는 출퇴근이 병합되므로 출근 행에도 텍스트 표시
                            item = QTableWidgetItem(leave_type)
                            item.setTextAlignment(Qt.AlignCenter)
                            # 연차 또는 경조사 등 출근하지 않은 경우 배경색 처리
                            if leave_type in ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']:
                                item.setBackground(QColor("#F8CBAD"))
                        else:
                            # 독립 처리 텍스트인데 remarks가 없으면 이전 데이터이거나 잘못된 데이터
                            # 출근 행에는 시간이 있으면 시간 표시, 없으면 빈 셀
                            if record['arrival']:
                                arrival_str = str(record['arrival'])[:5] if record['arrival'] else ""
                                item = QTableWidgetItem(arrival_str)
                                item.setTextAlignment(Qt.AlignCenter)
                                try:
                                    arrival_time_obj = datetime.strptime(arrival_str, "%H:%M").time()
                                    work_date_obj = datetime(year, month, day).date()
                                    weekday = work_date_obj.weekday()
                                    if weekday != 5 and weekday != 6:
                                        arrival_times.append(arrival_time_obj)
                                    if arrival_time_obj < datetime.strptime("08:00", "%H:%M").time():
                                        item.setForeground(QColor("#008000"))
                                    elif arrival_time_obj > datetime.strptime("09:00", "%H:%M").time():
                                        item.setForeground(QColor("#FF0000"))
                                except:
                                    pass
                            else:
                                item = QTableWidgetItem("")  # 빈 셀
                                item.setBackground(QColor("#F0F0F0"))
                                item.setTextAlignment(Qt.AlignCenter)
                    elif record['arrival']:
                        # 일반 시간 입력 → 평균 출근시간에 포함
                        arrival_str = str(record['arrival'])[:5] if record['arrival'] else ""
                        item = QTableWidgetItem(arrival_str)
                        item.setTextAlignment(Qt.AlignCenter)
                        try:
                            arrival_time_obj = datetime.strptime(arrival_str, "%H:%M").time()
                            # 토요일(5)과 일요일(6)은 평균 출근시간에서 제외
                            work_date_obj = datetime(year, month, day).date()
                            weekday = work_date_obj.weekday()  # 0=월요일, 5=토요일, 6=일요일
                            if weekday != 5 and weekday != 6:  # 토요일, 일요일이 아닐 때만 추가
                                arrival_times.append(arrival_time_obj)
                            # 08시 이전 출근 - 초록색
                            if arrival_time_obj < datetime.strptime("08:00", "%H:%M").time():
                                item.setForeground(QColor("#008000"))
                            # 09시 이후 지각 - 빨간색
                            elif arrival_time_obj > datetime.strptime("09:00", "%H:%M").time():
                                item.setForeground(QColor("#FF0000"))
                        except:
                            pass
                    else:
                        item = QTableWidgetItem("")  # 빈 셀
                        item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                        item.setTextAlignment(Qt.AlignCenter)
                else:
                    item = QTableWidgetItem("")  # 빈 셀
                    item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                    item.setTextAlignment(Qt.AlignCenter)
                
                item.setData(Qt.UserRole, {'emp_id': emp_id, 'day': day, 'category': '출근'})
                self.table.setItem(arrival_row, col, item)
            
            # 요약 컬럼 계산
            early_count = sum(1 for r in records_dict.values() if r.get('early'))
            late_arr_count = sum(1 for r in records_dict.values() if r.get('late_arr'))
            # 연차사용 계산: "연차"와 "휴가"는 1.0, "반차"는 0.5만 반영
            # 공휴, 박람회, 출장, 교육, 추석, 설날, 민방위 등은 연차사용에 반영 안 함
            leave_amount = 0.0
            for r in records_dict.values():
                leave_type = r.get('leave_type')
                remarks = r.get('remarks', '')
                if leave_type:
                    if leave_type == '연차' or leave_type == '휴가':
                        # 연차와 휴가는 1.0
                        leave_amount += 1.0
                    elif leave_type == '반차':
                        # 반차는 0.5 (출근 또는 퇴근 중 하나)
                        leave_amount += 0.5
                    # 그 외 휴가 유형(공휴, 박람회, 출장, 교육, 추석, 설날, 민방위 등)은 연차사용에 반영 안 함
            
            avg_arrival = ""
            if arrival_times:
                total_seconds = sum(t.hour * 3600 + t.minute * 60 + t.second for t in arrival_times)
                avg_seconds = total_seconds // len(arrival_times)
                avg_arrival = f"{avg_seconds // 3600:02d}:{(avg_seconds % 3600) // 60:02d}"
            
            # 요약 컬럼 설정
            summary_start = 3 + 31  # 3(기본) + 31(날짜)
            # 조기출근: 0이면 빈 셀로 표시
            if early_count > 0:
                early_item = QTableWidgetItem(str(early_count))
                early_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start, early_item)
            else:
                early_item = QTableWidgetItem("")
                early_item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                early_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start, early_item)
            # 지각: 0이면 빈 셀로 표시
            if late_arr_count > 0:
                late_arr_item = QTableWidgetItem(str(late_arr_count))
                late_arr_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start + 1, late_arr_item)
            else:
                late_arr_item = QTableWidgetItem("")
                late_arr_item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                late_arr_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start + 1, late_arr_item)
            # 출근 행의 야근 컬럼은 해당 없으므로 음영 처리
            late_dep_item_arr = QTableWidgetItem("")
            late_dep_item_arr.setBackground(QColor("#F0F0F0"))  # 음영 처리
            late_dep_item_arr.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(arrival_row, summary_start + 2, late_dep_item_arr)
            # 연차사용: 0이면 빈 셀로 표시
            if leave_amount > 0:
                leave_item = QTableWidgetItem(str(leave_amount))
                leave_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start + 3, leave_item)
            else:
                leave_item = QTableWidgetItem("")
                leave_item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                leave_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start + 3, leave_item)
            # 평균 출근시간: 값이 없으면 빈 셀로 표시
            if avg_arrival:
                avg_arrival_item = QTableWidgetItem(avg_arrival)
                avg_arrival_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start + 4, avg_arrival_item)
            else:
                avg_arrival_item = QTableWidgetItem("")
                avg_arrival_item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                avg_arrival_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start + 4, avg_arrival_item)
            # 출근 행의 평균 퇴근시간 컬럼은 해당 없으므로 음영 처리
            avg_departure_item_arr = QTableWidgetItem("")
            avg_departure_item_arr.setBackground(QColor("#F0F0F0"))  # 음영 처리
            avg_departure_item_arr.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(arrival_row, summary_start + 5, avg_departure_item_arr)
            
            # 퇴근 행 추가
            departure_row = self.table.rowCount()
            self.table.insertRow(departure_row)
            
            # 퇴근 행의 행 번호는 공란으로 설정 (출근 행의 번호가 두 행에 걸쳐 보이도록)
            self.table.setVerticalHeaderItem(departure_row, QTableWidgetItem(""))
            
            dep_pos_item = QTableWidgetItem("")  # 직급 빈칸
            dep_pos_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(departure_row, 0, dep_pos_item)
            dep_name_item = QTableWidgetItem("")  # 이름 빈칸 (시각적 병합)
            dep_name_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(departure_row, 1, dep_name_item)
            dep_category_item = QTableWidgetItem("퇴근")
            dep_category_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(departure_row, 2, dep_category_item)
            
            departure_times = []
            for day in range(1, days_in_month + 1):
                col = 2 + day
                if day in records_dict:
                    record = records_dict[day]
                    # "연차", "경조사", "예비군", "설날", "추석", "박람회", "출장"을 제외한 나머지는 remarks로 출근/퇴근 구분
                    remarks = record.get('remarks', '')
                    leave_type = record.get('leave_type')
                    merge_texts = ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']
                    
                    # remarks가 "{type}_퇴근" 형식인 경우 (독립 처리 텍스트)
                    if leave_type and remarks == f'{leave_type}_퇴근':
                        # 퇴근 행에 텍스트 표시 → 평균 퇴근시간에서 제외
                        item = QTableWidgetItem(leave_type)
                        item.setTextAlignment(Qt.AlignCenter)
                        # 연차 또는 경조사 등 출근하지 않은 경우 배경색 처리
                        if leave_type in ['연차', '반차', '경조사', '공휴', '박람회', '예비군', '추석', '설날', '민방위', '출장', '교육']:
                            item.setBackground(QColor("#F8CBAD"))
                        self.table.setItem(departure_row, col, item)
                    elif leave_type and remarks == f'{leave_type}_출근':
                        # 출근 행용이므로 퇴근 행에는 시간 표시 → 평균 퇴근시간에 포함 (출근행에 텍스트가 있으므로)
                        if record['departure']:
                            departure_str = str(record['departure'])[:5] if record['departure'] else ""
                            item = QTableWidgetItem(departure_str)
                            item.setTextAlignment(Qt.AlignCenter)
                            try:
                                departure_time_obj = datetime.strptime(departure_str, "%H:%M").time()
                                work_date_obj = datetime(year, month, day).date()
                                weekday = work_date_obj.weekday()  # 0=월요일, 5=토요일, 6=일요일
                                # 토요일(5)과 일요일(6)은 평균 퇴근시간에서 제외
                                # 매월 셋째주 수요일 17:00 퇴근도 평균 계산에서 제외
                                if weekday != 5 and weekday != 6:  # 토요일, 일요일이 아닐 때만 체크
                                    if not self.is_third_wednesday_17_00(work_date_obj, departure_time_obj):
                                        departure_times.append(departure_time_obj)
                                # 20시 이후 퇴근 - 파랑색
                                if departure_time_obj >= datetime.strptime("20:00", "%H:%M").time():
                                    item.setForeground(QColor("#0000FF"))
                            except:
                                pass
                        else:
                            item = QTableWidgetItem("")  # 빈 셀
                            item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                            item.setTextAlignment(Qt.AlignCenter)
                        self.table.setItem(departure_row, col, item)
                    elif leave_type and not remarks:
                        # remarks가 없는 경우
                        # "연차", "경조사", "예비군", "설날", "추석", "박람회", "출장"은 출퇴근 병합 처리
                        # 나머지는 remarks로 출근/퇴근을 구분하므로, remarks가 없으면 시간 표시
                        if leave_type in merge_texts:
                            # 병합 처리 텍스트는 출퇴근이 병합되므로 퇴근 행에도 텍스트 표시
                            item = QTableWidgetItem(leave_type)
                            item.setTextAlignment(Qt.AlignCenter)
                            # 연차 또는 경조사 등 출근하지 않은 경우 배경색 처리
                            if leave_type in ['연차', '경조사', '예비군', '설날', '추석', '박람회', '출장']:
                                item.setBackground(QColor("#F8CBAD"))
                            self.table.setItem(departure_row, col, item)
                        else:
                            # 독립 처리 텍스트인데 remarks가 없으면 이전 데이터이거나 잘못된 데이터
                            # 퇴근 행에는 시간이 있으면 시간 표시, 없으면 빈 셀
                            if record['departure']:
                                departure_str = str(record['departure'])[:5] if record['departure'] else ""
                                item = QTableWidgetItem(departure_str)
                                item.setTextAlignment(Qt.AlignCenter)
                                try:
                                    departure_time_obj = datetime.strptime(departure_str, "%H:%M").time()
                                    work_date_obj = datetime(year, month, day).date()
                                    weekday = work_date_obj.weekday()
                                    if weekday != 5 and weekday != 6:
                                        if not self.is_third_wednesday_17_00(work_date_obj, departure_time_obj):
                                            departure_times.append(departure_time_obj)
                                    if departure_time_obj >= datetime.strptime("20:00", "%H:%M").time():
                                        item.setForeground(QColor("#0000FF"))
                                except:
                                    pass
                            else:
                                item = QTableWidgetItem("")  # 빈 셀
                                item.setBackground(QColor("#F0F0F0"))
                                item.setTextAlignment(Qt.AlignCenter)
                        self.table.setItem(departure_row, col, item)
                    elif record['departure']:
                        # 일반 시간 입력 → 평균 퇴근시간에 포함
                        departure_str = str(record['departure'])[:5] if record['departure'] else ""
                        item = QTableWidgetItem(departure_str)
                        item.setTextAlignment(Qt.AlignCenter)
                        try:
                            departure_time_obj = datetime.strptime(departure_str, "%H:%M").time()
                            work_date_obj = datetime(year, month, day).date()
                            weekday = work_date_obj.weekday()  # 0=월요일, 5=토요일, 6=일요일
                            # 토요일(5)과 일요일(6)은 평균 퇴근시간에서 제외
                            # 매월 셋째주 수요일 17:00 퇴근도 평균 계산에서 제외
                            if weekday != 5 and weekday != 6:  # 토요일, 일요일이 아닐 때만 체크
                                if not self.is_third_wednesday_17_00(work_date_obj, departure_time_obj):
                                    departure_times.append(departure_time_obj)
                            # 20시 이후 퇴근 - 파랑색
                            if departure_time_obj >= datetime.strptime("20:00", "%H:%M").time():
                                item.setForeground(QColor("#0000FF"))
                        except:
                            pass
                        self.table.setItem(departure_row, col, item)
                    else:
                        item = QTableWidgetItem("")  # 빈 셀
                        item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                        item.setTextAlignment(Qt.AlignCenter)
                        self.table.setItem(departure_row, col, item)
                else:
                    item = QTableWidgetItem("")  # 빈 셀
                    item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                    item.setTextAlignment(Qt.AlignCenter)
                
                item.setData(Qt.UserRole, {'emp_id': emp_id, 'day': day, 'category': '퇴근'})
                self.table.setItem(departure_row, col, item)
            
            late_dep_count = sum(1 for r in records_dict.values() if r.get('late_dep'))
            avg_departure = ""
            if departure_times:
                total_seconds = sum(t.hour * 3600 + t.minute * 60 + t.second for t in departure_times)
                avg_seconds = total_seconds // len(departure_times)
                avg_departure = f"{avg_seconds // 3600:02d}:{(avg_seconds % 3600) // 60:02d}"
            
            # 퇴근 행의 조기출근 컬럼은 해당 없으므로 음영 처리
            early_item_dep = QTableWidgetItem("")
            early_item_dep.setBackground(QColor("#F0F0F0"))  # 음영 처리
            early_item_dep.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(departure_row, summary_start, early_item_dep)
            
            # 퇴근 행의 지각 컬럼은 해당 없으므로 음영 처리
            late_arr_item_dep = QTableWidgetItem("")
            late_arr_item_dep.setBackground(QColor("#F0F0F0"))  # 음영 처리
            late_arr_item_dep.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(departure_row, summary_start + 1, late_arr_item_dep)
            
            # 야근: 0이면 빈 셀로 표시
            if late_dep_count > 0:
                late_dep_item = QTableWidgetItem(str(late_dep_count))
                late_dep_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(departure_row, summary_start + 2, late_dep_item)
            else:
                late_dep_item = QTableWidgetItem("")
                late_dep_item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                late_dep_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(departure_row, summary_start + 2, late_dep_item)
            
            # 퇴근 행의 연차사용 컬럼은 해당 없으므로 음영 처리
            leave_item_dep = QTableWidgetItem("")
            leave_item_dep.setBackground(QColor("#F0F0F0"))  # 음영 처리
            leave_item_dep.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(departure_row, summary_start + 3, leave_item_dep)
            
            # 퇴근 행의 평균 출근시간 컬럼은 해당 없으므로 음영 처리
            avg_arrival_item_dep = QTableWidgetItem("")
            avg_arrival_item_dep.setBackground(QColor("#F0F0F0"))  # 음영 처리
            avg_arrival_item_dep.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(departure_row, summary_start + 4, avg_arrival_item_dep)
            
            # 평균 퇴근시간: 값이 없으면 빈 셀로 표시
            if avg_departure:
                avg_departure_item = QTableWidgetItem(avg_departure)
                avg_departure_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(departure_row, summary_start + 5, avg_departure_item)
            else:
                avg_departure_item = QTableWidgetItem("")
                avg_departure_item.setBackground(QColor("#F0F0F0"))  # 음영 처리
                avg_departure_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(departure_row, summary_start + 5, avg_departure_item)
            
            # 출근 행과 퇴근 행을 날짜 컬럼을 제외한 컬럼에서 병합
            # 직급, 이름 컬럼 병합
            self.table.setSpan(arrival_row, 0, 2, 1)  # 직급 병합 (2행, 1열)
            self.table.setSpan(arrival_row, 1, 2, 1)  # 이름 병합 (2행, 1열)
            
            # 요약 컬럼들 병합
            # 조기출근: 출근 행의 값 사용
            self.table.setSpan(arrival_row, summary_start, 2, 1)
            
            # 지각: 출근 행의 값 사용
            self.table.setSpan(arrival_row, summary_start + 1, 2, 1)
            
            # 야근: 퇴근 행의 값 사용 - 출근 행의 셀을 퇴근 행의 값으로 덮어쓰고 병합
            if late_dep_count > 0:
                late_dep_merged = QTableWidgetItem(str(late_dep_count))
                late_dep_merged.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start + 2, late_dep_merged)
            else:
                late_dep_merged = QTableWidgetItem("")
                late_dep_merged.setBackground(QColor("#F0F0F0"))
                late_dep_merged.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(arrival_row, summary_start + 2, late_dep_merged)
            self.table.setSpan(arrival_row, summary_start + 2, 2, 1)
            
            # 연차사용: 출근 행의 값 사용
            self.table.setSpan(arrival_row, summary_start + 3, 2, 1)
            
            # 평균 출근시간: 출근 행의 값 사용
            self.table.setSpan(arrival_row, summary_start + 4, 2, 1)
            
            # 평균 퇴근시간: 퇴근 행의 값 사용 - 출근 행의 셀을 퇴근 행의 값으로 덮어쓰고 병합
            if avg_departure:
                self.table.setItem(arrival_row, summary_start + 5, QTableWidgetItem(avg_departure))
            else:
                avg_dep_merged = QTableWidgetItem("")
                avg_dep_merged.setBackground(QColor("#F0F0F0"))
                self.table.setItem(arrival_row, summary_start + 5, avg_dep_merged)
            self.table.setSpan(arrival_row, summary_start + 5, 2, 1)
            
            # 특정 텍스트(연차, 박람회, 예비군, 경조사, 추석, 설날, 공휴)가 있는 날짜 컬럼에서 출근행과 퇴근행 병합
            merge_texts = ['연차', '박람회', '예비군', '경조사', '추석', '설날', '공휴']
            for day in range(1, days_in_month + 1):
                col = 2 + day  # 3번 컬럼부터 시작
                arrival_item = self.table.item(arrival_row, col)
                departure_item = self.table.item(departure_row, col)
                
                # 이미 병합된 셀인지 확인 (rowSpan이 2 이상이면 이미 병합됨)
                if self.table.rowSpan(arrival_row, col) >= 2:
                    continue
                
                # 출근행과 퇴근행 모두에 셀이 있는 경우
                if arrival_item and departure_item:
                    arrival_text = arrival_item.text().strip()
                    departure_text = departure_item.text().strip()
                    
                    # 출근행과 퇴근행의 텍스트가 같고, 병합 대상 텍스트인 경우
                    if arrival_text == departure_text and arrival_text in merge_texts:
                        # 퇴근행의 셀을 먼저 제거 (병합 전에 제거해야 함)
                        self.table.setItem(departure_row, col, None)
                        # 출근행과 퇴근행 병합 (2행, 1열)
                        self.table.setSpan(arrival_row, col, 2, 1)
                elif day in records_dict:
                    record = records_dict[day]
                    leave_type = record.get('leave_type')
                    # leave_type이 병합 대상 텍스트 중 하나인 경우
                    if leave_type in merge_texts:
                        # 출근행의 셀을 가져와서 병합
                        if arrival_item:
                            # 퇴근행의 셀을 먼저 제거 (병합 전에 제거해야 함)
                            if departure_item:
                                self.table.setItem(departure_row, col, None)
                            # 출근행과 퇴근행 병합 (2행, 1열)
                            self.table.setSpan(arrival_row, col, 2, 1)
            
            # 입사일 이전 날짜에 대해 병합 및 사선 처리
            if hire_date:
                for day in range(1, days_in_month + 1):
                    try:
                        current_date = datetime(year, month, day).date()
                        # 입사일 이전 날짜인 경우만 대각선 처리 (입사일 당일은 제외)
                        # 예: 11월 25일 입사 → 11월 1일~24일만 대각선, 12월 1일 이후는 대각선 없음
                        if current_date < hire_date:
                            col = 2 + day  # 3번 컬럼부터 시작
                            
                            # 이미 병합된 셀인지 확인 (rowSpan이 2 이상이면 이미 병합됨)
                            if self.table.rowSpan(arrival_row, col) >= 2:
                                # 이미 병합된 경우 배경색만 설정
                                arrival_item = self.table.item(arrival_row, col)
                                if arrival_item:
                                    arrival_item.setBackground(QColor("#E8E8E8"))
                                continue
                            
                            # 출근행의 셀 확인
                            arrival_item = self.table.item(arrival_row, col)
                            if not arrival_item:
                                # 셀이 없으면 빈 셀 생성
                                arrival_item = QTableWidgetItem("")
                                arrival_item.setBackground(QColor("#E8E8E8"))
                                self.table.setItem(arrival_row, col, arrival_item)
                            else:
                                # 배경색 설정 (연한 회색)
                                arrival_item.setBackground(QColor("#E8E8E8"))
                            
                            # 출근행과 퇴근행 병합 (2행, 1열)
                            self.table.setSpan(arrival_row, col, 2, 1)
                            
                            # 퇴근행의 셀 제거
                            departure_item = self.table.item(departure_row, col)
                            if departure_item:
                                self.table.setItem(departure_row, col, None)
                            
                            # 사선을 그릴 셀 추가
                            diagonal_delegate.add_target_cell(arrival_row, col)
                    except ValueError:
                        pass  # 유효하지 않은 날짜는 무시
            
            # 퇴사일 이후 날짜에 대해 병합 및 사선 처리
            if resignation_date:
                for day in range(1, days_in_month + 1):
                    try:
                        current_date = datetime(year, month, day).date()
                        # 퇴사일 이후 날짜인 경우만 대각선 처리 (퇴사일 당일은 제외)
                        # 예: 12월 8일 퇴사 → 12월 9일부터 대각선
                        if current_date > resignation_date:
                            col = 2 + day  # 3번 컬럼부터 시작
                            
                            # 이미 병합된 셀인지 확인 (rowSpan이 2 이상이면 이미 병합됨)
                            if self.table.rowSpan(arrival_row, col) >= 2:
                                # 이미 병합된 경우 배경색만 설정
                                arrival_item = self.table.item(arrival_row, col)
                                if arrival_item:
                                    arrival_item.setBackground(QColor("#E8E8E8"))
                                else:
                                    # 셀이 없으면 빈 셀 생성
                                    arrival_item = QTableWidgetItem("")
                                    arrival_item.setBackground(QColor("#E8E8E8"))
                                    self.table.setItem(arrival_row, col, arrival_item)
                                # 사선을 그릴 셀 추가
                                diagonal_delegate.add_target_cell(arrival_row, col)
                                continue
                            
                            # 출근행의 셀 확인
                            arrival_item = self.table.item(arrival_row, col)
                            if not arrival_item:
                                # 셀이 없으면 빈 셀 생성
                                arrival_item = QTableWidgetItem("")
                                arrival_item.setBackground(QColor("#E8E8E8"))
                                self.table.setItem(arrival_row, col, arrival_item)
                            else:
                                # 배경색 설정 (연한 회색)
                                arrival_item.setBackground(QColor("#E8E8E8"))
                            
                            # 출근행과 퇴근행 병합 (2행, 1열)
                            self.table.setSpan(arrival_row, col, 2, 1)
                            
                            # 퇴근행의 셀 제거
                            departure_item = self.table.item(departure_row, col)
                            if departure_item:
                                self.table.setItem(departure_row, col, None)
                            
                            # 사선을 그릴 셀 추가
                            diagonal_delegate.add_target_cell(arrival_row, col)
                    except ValueError:
                        pass  # 유효하지 않은 날짜는 무시
            
            # 퇴사자이고 체크박스가 OFF일 경우 출근행과 퇴근행 숨김 처리
            if is_active != 1 and not show_inactive:
                # 모든 병합 해제 (숨기기 전에 병합을 해제해야 완전히 숨김 처리됨)
                summary_start = 3 + 31  # 3(기본) + 31(날짜)
                
                # 직급, 이름 컬럼 병합 해제
                self.table.setSpan(arrival_row, 0, 1, 1)
                self.table.setSpan(arrival_row, 1, 1, 1)
                
                # 요약 컬럼들 병합 해제
                for col_idx in range(summary_start, summary_start + 6):
                    if self.table.rowSpan(arrival_row, col_idx) >= 2:
                        self.table.setSpan(arrival_row, col_idx, 1, 1)
                
                # 날짜 컬럼들의 병합 해제
                for day in range(1, days_in_month + 1):
                    col = 2 + day
                    if self.table.rowSpan(arrival_row, col) >= 2:
                        self.table.setSpan(arrival_row, col, 1, 1)
                
                # 행 숨김 처리
                self.table.setRowHidden(arrival_row, True)
                self.table.setRowHidden(departure_row, True)
            else:
                # 표시되는 직원이 있으면 해당 부서의 구분자 행도 표시
                for sep_info in separator_rows:
                    if sep_info['dept'] == dept:
                        sep_info['has_visible_employee'] = True
                        break
            
            # 다음 직원을 위한 행 번호 증가
            employee_row_number += 1
        
        # 모든 구분자 행 처리: 표시되는 직원이 없는 부서의 구분자 행 숨김
        for sep_info in separator_rows:
            if not sep_info['has_visible_employee']:
                # 구분자 행의 병합 해제 (숨기기 전에 병합을 해제해야 완전히 숨김 처리됨)
                sep_row = sep_info['row']
                # 구분자 행은 모든 컬럼에 걸쳐 병합되어 있으므로 병합 해제
                if self.table.columnSpan(sep_row, 0) > 1:
                    self.table.setSpan(sep_row, 0, 1, 1)
                # 행 숨김 처리
                self.table.setRowHidden(sep_row, True)
        
        # 재직인원 수 계산 (구분자 행 제외, 활성 직원만 카운트)
        active_employee_count = sum(1 for emp in employees if emp[6] == 1)  # emp[6]은 is_active
        if hasattr(self, 'employee_count_label'):
            self.employee_count_label.setText(f"재직인원: {active_employee_count}명")
        
        # 모든 직원 처리 완료 후 델리게이트 적용
        if diagonal_delegate.target_cells:
            self.table.setItemDelegate(diagonal_delegate)
    
    def on_cell_double_clicked(self, row, col):
        """셀 더블클릭 이벤트 - 시간 편집 (여러 셀 선택 지원)"""
        # 날짜 컬럼만 편집 가능 (3번째 컬럼부터, 즉 col >= 3)
        if col < 3:
            return
        
        # 더블클릭한 셀 가져오기
        item = self.table.item(row, col)
        if not item:
            return
        
        # 선택된 모든 셀 가져오기
        selected_items = self.table.selectedItems()
        
        # 선택된 셀이 없으면 현재 셀만 처리
        if not selected_items:
            selected_items = [item]
        
        # 유효한 셀만 필터링 (UserRole 데이터가 있는 셀만)
        valid_cells = []
        for selected_item in selected_items:
            if not selected_item:
                continue
            data = selected_item.data(Qt.UserRole)
            if not data or not isinstance(data, dict):
                # UserRole 데이터가 없는 경우도 처리 (날짜 컬럼이 아닌 경우)
                continue
            emp_id = data.get('emp_id')
            day = data.get('day')
            category = data.get('category')
            if not all([emp_id, day, category]):
                continue
            valid_cells.append({
                'item': selected_item,
                'emp_id': emp_id,
                'day': day,
                'category': category
            })
        
        if not valid_cells:
            # 유효한 셀이 없으면 메시지 표시 (디버깅용)
            # print(f"더블클릭: row={row}, col={col}, valid_cells={len(valid_cells)}")
            return
        
        # 여러 셀 선택 시 일괄 편집
        if len(valid_cells) > 1:
            # 첫 번째 셀의 현재 값 표시 (참고용)
            first_value = valid_cells[0]['item'].text()
            self.edit_multiple_cells_dialog(valid_cells, first_value)
        else:
            # 단일 셀 편집
            cell = valid_cells[0]
            year = self.year_combo.currentData()
            month = self.month_combo.currentData()
            if year is None:
                year = datetime.now().year
            if month is None:
                month = datetime.now().month
            
            try:
                work_date = datetime(year, month, cell['day']).date()
            except ValueError:
                QMessageBox.critical(self, "오류", "유효하지 않은 날짜입니다.")
                return
            
            self.edit_time_dialog(cell['emp_id'], work_date, cell['category'], cell['item'].text())
    
    def on_cell_changed(self, item):
        """셀 편집 완료 시 호출 - 직접 편집된 값 저장"""
        # 최상위 예외 처리 - 모든 예외를 잡아서 프로그램이 종료되지 않도록 함
        try:
            if not item:
                return
            
            # 데이터 새로고침 중일 때는 처리하지 않음 (무한 루프 방지)
            if self._is_refreshing:
                return
            
            # item이 유효한지 확인 (삭제된 경우 예외 처리)
            try:
                row = item.row()
                col = item.column()
            except RuntimeError:
                # item이 이미 삭제된 경우
                return
            
            # 날짜 컬럼만 편집 가능 (3번째 컬럼부터, 요약 컬럼 전까지)
            summary_start = 34  # 3(기본) + 31(날짜) = 34
            if col < 3 or col >= summary_start:
                return
            
            # UserRole 데이터 확인
            data = item.data(Qt.UserRole)
            if not data or not isinstance(data, dict):
                return
            
            emp_id = data.get('emp_id')
            day = data.get('day')
            category = data.get('category')
            
            if not all([emp_id, day, category]):
                return
            
            # emp_id가 유효한지 확인 (퇴사자 포함 - is_active와 관계없이 확인)
            conn_check = self.db.get_connection()
            cursor_check = conn_check.cursor()
            try:
                cursor_check.execute("SELECT id FROM employees WHERE id = ?", (emp_id,))
                if not cursor_check.fetchone():
                    conn_check.close()
                    QMessageBox.warning(self, "오류", "유효하지 않은 직원 ID입니다.")
                    return
            except Exception as e:
                conn_check.close()
                QMessageBox.warning(self, "오류", f"직원 정보 확인 중 오류 발생: {str(e)}")
                return
            finally:
                conn_check.close()
            
            # 년도와 월 가져오기
            year = self.year_combo.currentData()
            month = self.month_combo.currentData()
            if year is None:
                year = datetime.now().year
            if month is None:
                month = datetime.now().month
            
            try:
                work_date = datetime(year, month, day).date()
            except ValueError:
                return
            
            # 입력된 값 가져오기
            new_value = item.text().strip()
            input_text = None  # 연월차 관리대장 업데이트를 위해 변수 초기화
            
            # 빈 값 처리 - pending_changes에 저장 (저장 버튼에서 처리)
            if not new_value:
                # 빈 값이면 기록 삭제를 pending_changes에 저장
                change_key = (emp_id, work_date, category)
                self.pending_changes[change_key] = {
                    'new_value': '',
                    'formatted_time': '',
                    'is_time': False,
                    'item': item,
                    'is_delete': True  # 삭제 플래그
                }
                
                return
        
            # 시간 형식 변환 (예: 1000 -> 10:00)
            def format_time_input(time_str):
                time_str = time_str.strip()
                if ':' in time_str:
                    return time_str
                if time_str.isdigit():
                    if len(time_str) == 4:
                        return f"{time_str[:2]}:{time_str[2:]}"
                    elif len(time_str) == 3:
                        return f"0{time_str[0]}:{time_str[1:]}"
                return time_str
            
            formatted_time = format_time_input(new_value)
            
            # 시간 형식인지 확인 (':' 포함 또는 숫자만)
            is_time = ':' in formatted_time or (formatted_time.isdigit() and len(formatted_time) >= 3)
            
            # 시간 형식인 경우 색상 업데이트만 수행
            if is_time:
                try:
                    # 시간 형식 변환 시도
                    if ':' in formatted_time:
                        time_obj = datetime.strptime(formatted_time, "%H:%M").time()
                    elif formatted_time.isdigit() and len(formatted_time) == 4:
                        time_obj = datetime.strptime(f"{formatted_time[:2]}:{formatted_time[2:]}", "%H:%M").time()
                    else:
                        is_time = False
                    
                    if is_time:
                        # 색상 업데이트만 수행
                        if category == '출근':
                            # 08시 이전 출근 - 초록색
                            if time_obj < datetime.strptime("08:00", "%H:%M").time():
                                item.setForeground(QColor("#008000"))
                            # 09시 이후 지각 - 빨간색
                            elif time_obj > datetime.strptime("09:00", "%H:%M").time():
                                item.setForeground(QColor("#FF0000"))
                            else:
                                # 08시~09시 사이 또는 정확히 09시 - 검정색 (기본 색상)
                                item.setForeground(QColor("#000000"))
                        else:  # 퇴근
                            # 20시 이후 야근 - 파랑색
                            if time_obj >= datetime.strptime("20:00", "%H:%M").time():
                                item.setForeground(QColor("#0000FF"))
                            else:
                                # 20시 이전 - 검정색 (기본 색상)
                                item.setForeground(QColor("#000000"))
                except:
                    is_time = False
            
            # 모든 변경사항을 pending_changes에 저장 (저장 버튼에서 일괄 처리)
            change_key = (emp_id, work_date, category)
            self.pending_changes[change_key] = {
                'new_value': new_value,
                'formatted_time': formatted_time,
                'is_time': is_time,
                'item': item
            }
            
            # 텍스트 입력 시 색상을 검정색으로 원복 (시간이 아닌 경우)
            if not is_time and (category == '출근' or category == '퇴근'):
                item.setForeground(QColor("#000000"))
            
            # 즉시 저장하지 않고 pending_changes에만 저장 (저장 버튼에서 일괄 처리)
        except Exception as e:
            # 최상위 예외 처리 - 예상치 못한 오류 발생 시
            import traceback
            print(f"셀 편집 처리 중 오류 발생: {str(e)}")
            print(traceback.format_exc())
    
    def sync_leave_records(self):
        """출퇴근 관리대장의 연차/반차 기록을 leave_records에 동기화"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        try:
            # attendance_records에서 연차/반차 기록 조회
            cursor.execute("""
                SELECT ar.employee_id, ar.work_date, ar.leave_type, ar.remarks,
                       e.name
                FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.id
                WHERE ar.leave_type IN ('연차', '반차', '휴가')
                ORDER BY ar.work_date
            """)
            attendance_records = cursor.fetchall()
            
            synced_count = 0
            updated_count = 0
            skipped_count = 0
            
            # 반차 중복 방지를 위한 집합 (하루에 한 번만 기록)
            반차_processed = set()
            
            for emp_id, work_date, leave_type, remarks, emp_name in attendance_records:
                if isinstance(work_date, str):
                    work_date = datetime.strptime(work_date, "%Y-%m-%d").date()
                
                year = work_date.year
                month = work_date.month
                
                # 반차의 경우 중복 방지
                if leave_type == '반차':
                    반차_key = (emp_id, work_date)
                    if 반차_key in 반차_processed:
                        skipped_count += 1
                        continue
                    반차_processed.add(반차_key)
                    leave_amount = 0.5
                elif leave_type in ['연차', '휴가']:
                    leave_amount = 1.0
                else:
                    continue
                
                # 기존 leave_records 확인
                cursor.execute("""
                    SELECT id FROM leave_records
                    WHERE employee_id = ? AND leave_date = ? AND leave_type = ?
                """, (emp_id, work_date, leave_type))
                existing = cursor.fetchone()
                
                if existing:
                    # 기존 기록이 있으면 업데이트 (leave_amount가 다를 수 있음)
                    # 반차의 경우 항상 0.5로 강제 업데이트
                    if leave_type == '반차':
                        leave_amount = 0.5
                    cursor.execute("""
                        UPDATE leave_records
                        SET leave_amount = ?, year = ?, month = ?
                        WHERE employee_id = ? AND leave_date = ? AND leave_type = ?
                    """, (leave_amount, year, month, emp_id, work_date, leave_type))
                    updated_count += 1
                else:
                    # 새로 추가
                    cursor.execute("""
                        INSERT INTO leave_records
                        (employee_id, leave_type, leave_date, leave_amount, year, month)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (emp_id, leave_type, work_date, leave_amount, year, month))
                    synced_count += 1
            
            conn.commit()
            
            # 결과 메시지
            total_processed = synced_count + updated_count + skipped_count
            message = f"동기화 완료\n\n"
            message += f"처리된 기록: {total_processed}건\n"
            message += f"새로 추가: {synced_count}건\n"
            message += f"업데이트: {updated_count}건\n"
            if skipped_count > 0:
                message += f"건너뜀 (중복): {skipped_count}건"
            
            QMessageBox.information(self, "동기화 완료", message)
            
            # 연월차 관리대장도 새로고침
            if self.leave_gui:
                self.leave_gui.refresh_data()
            
        except Exception as e:
            QMessageBox.critical(self, "오류", f"동기화 중 오류 발생: {str(e)}")
        finally:
            conn.close()
    
    def edit_time_dialog(self, emp_id, work_date, category, current_value):
        """시간 편집 다이얼로그"""
        dialog = QDialog(self)
        dialog.setWindowTitle("시간 수정")
        dialog.setModal(True)
        dialog.resize(400, 250)
        layout = QVBoxLayout(dialog)
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM employees WHERE id = ?", (emp_id,))
        name = cursor.fetchone()[0]
        conn.close()
    
        layout.addWidget(QLabel(f"직원: {name}"))
        layout.addWidget(QLabel(f"날짜: {work_date.strftime('%Y년 %m월 %d일')}"))
        layout.addWidget(QLabel(f"구분: {category}"))
        
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel("시간:"))
        time_entry = QLineEdit()
        if current_value and ':' in current_value:
            time_entry.setText(current_value[:5])
        elif current_value:
            time_entry.setText(current_value)
        time_layout.addWidget(time_entry)
        layout.addLayout(time_layout)
        
        layout.addWidget(QLabel("시간 형식: HH:MM (예: 09:00) 또는 휴가 유형 (예: 연차, 반차, 공휴)"))
        
        def format_time_input(time_str):
            """시간 입력 형식 변환 (예: 1000 -> 10:00)"""
            time_str = time_str.strip()
            if ':' in time_str:
                return time_str
            if time_str.isdigit():
                if len(time_str) == 4:
                    return f"{time_str[:2]}:{time_str[2:]}"
                elif len(time_str) == 3:
                    return f"0{time_str[0]}:{time_str[1:]}"
                elif len(time_str) == 2:
                    return f"00:{time_str}"
                elif len(time_str) == 1:
                    return f"00:0{time_str}"
            return time_str
        
        def save_time():
            try:
                time_str = time_entry.text().strip()
                
                if not time_str:
                    # 빈 값이면 기록 삭제
                    conn = self.db.get_connection()
                    cursor = conn.cursor()
                    cursor.execute("""
                        DELETE FROM attendance_records
                        WHERE employee_id = ? AND work_date = ?
                    """, (emp_id, work_date))
                    conn.commit()
                    conn.close()
                    
                    QMessageBox.information(self, "성공", "기록이 삭제되었습니다.")
                    dialog.accept()
                    self.refresh_data()
                    return
                
                time_str = format_time_input(time_str)
                
                arrival_time = None
                departure_time = None
                leave_type = None
                
                if ':' in time_str and len(time_str) <= 5:
                    try:
                        time_obj = datetime.strptime(time_str, "%H:%M").time()
                        if category == "출근":
                            arrival_time = time_obj
                            # 기존 퇴근 시간 유지
                            conn = self.db.get_connection()
                            cursor = conn.cursor()
                            cursor.execute("""
                                SELECT departure_time FROM attendance_records
                                WHERE employee_id = ? AND work_date = ?
                            """, (emp_id, work_date))
                            existing = cursor.fetchone()
                            if existing and existing[0]:
                                try:
                                    departure_time = datetime.strptime(existing[0], "%H:%M:%S").time()
                                except:
                                    pass
                            conn.close()
                        else:
                            departure_time = time_obj
                            # 기존 출근 시간 유지
                            conn = self.db.get_connection()
                            cursor = conn.cursor()
                            cursor.execute("""
                                SELECT arrival_time FROM attendance_records
                                WHERE employee_id = ? AND work_date = ?
                            """, (emp_id, work_date))
                            existing = cursor.fetchone()
                            if existing and existing[0]:
                                try:
                                    arrival_time = datetime.strptime(existing[0], "%H:%M:%S").time()
                                except:
                                    pass
                            conn.close()
                    except ValueError:
                        QMessageBox.critical(self, "오류", "시간 형식이 올바르지 않습니다. (HH:MM 형식)")
                        return
                elif time_str in ['연차', '반차', '반반차', '공휴', '박람회', '민방위', '출장', '교육', '추석', '설날']:
                    leave_type = time_str
                else:
                    QMessageBox.critical(self, "오류", "시간 형식(HH:MM) 또는 휴가 유형을 입력해주세요.")
                    return
                
                self.calculator.process_attendance_record(
                    emp_id, work_date, arrival_time, departure_time,
                    leave_type, ""
                )
                
                QMessageBox.information(self, "성공", "시간이 수정되었습니다.")
                dialog.accept()
                self.refresh_data()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"수정 중 오류 발생: {str(e)}")
        
            time_entry.returnPressed.connect(save_time)
            
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(save_time)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)
            
            dialog.exec()
    
    def edit_multiple_cells_dialog(self, valid_cells, current_value=""):
        """여러 셀 일괄 편집 다이얼로그"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"여러 셀 일괄 수정 ({len(valid_cells)}개 셀)")
        dialog.setModal(True)
        dialog.resize(450, 280)
        layout = QVBoxLayout(dialog)
        
        # 선택된 셀 정보 표시
        info_text = f"선택된 셀: {len(valid_cells)}개\n\n"
        info_text += "모든 선택된 셀에 동일한 값이 적용됩니다.\n"
        info_text += "시간 형식: HH:MM (예: 09:00) 또는 휴가 유형 (예: 연차, 반차, 공휴)"
        layout.addWidget(QLabel(info_text))
        
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel("입력할 값:"))
        time_entry = QLineEdit()
        if current_value:
            time_entry.setText(current_value)
        time_entry.setPlaceholderText("예: 09:00, 연차, 반차, 공휴")
        time_layout.addWidget(time_entry)
        layout.addLayout(time_layout)
        
        def format_time_input(time_str):
            """시간 입력 형식 변환 (예: 1000 -> 10:00)"""
            time_str = time_str.strip()
            if ':' in time_str:
                return time_str
            if time_str.isdigit():
                if len(time_str) == 4:
                    return f"{time_str[:2]}:{time_str[2:]}"
                elif len(time_str) == 3:
                    return f"0{time_str[0]}:{time_str[1:]}"
            return time_str
        
        def save_time():
            try:
                time_str = time_entry.text().strip()
                
                if not time_str:
                    # 빈 값이면 모든 선택된 셀의 기록 삭제
                    reply = QMessageBox.question(
                        dialog, "삭제 확인",
                        f"선택된 {len(valid_cells)}개 셀의 데이터를 모두 삭제하시겠습니까?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No
                    )
                    if reply != QMessageBox.Yes:
                        return
                    
                    conn = self.db.get_connection()
                    cursor = conn.cursor()
                    deleted_count = 0
                    base_year = self.year_combo.currentData()
                    base_month = self.month_combo.currentData()
                    if base_year is None:
                        base_year = datetime.now().year
                    if base_month is None:
                        base_month = datetime.now().month
                    
                    for cell in valid_cells:
                        year = base_year
                        month = base_month
                        # 셀의 날짜 정보를 사용할 수 없으므로 현재 선택된 월 사용
                        
                        try:
                            work_date = datetime(year, month, cell['day']).date()
                        except ValueError:
                            continue
                        
                        cursor.execute("""
                            DELETE FROM attendance_records
                            WHERE employee_id = ? AND work_date = ?
                        """, (cell['emp_id'], work_date))
                        deleted_count += cursor.rowcount
                    
                    conn.commit()
                    conn.close()
                    
                    if deleted_count > 0:
                        QMessageBox.information(dialog, "완료", f"{deleted_count}개 셀의 데이터가 삭제되었습니다.")
                    dialog.accept()
                    self.refresh_data()
                    return
                
                # 시간 형식 변환
                formatted_time = format_time_input(time_str)
                
                # 시간 또는 휴가 유형인지 확인
                is_time = ':' in formatted_time
                is_leave_type = formatted_time in ['연차', '반차', '반반차', '공휴', '박람회', '민방위', '출장', '교육', '추석', '설날']
                
                if not is_time and not is_leave_type:
                    QMessageBox.warning(dialog, "경고", "올바른 시간 형식(HH:MM) 또는 휴가 유형을 입력해주세요.")
                    return
                
                # 모든 선택된 셀에 적용
                conn = self.db.get_connection()
                cursor = conn.cursor()
                updated_count = 0
                base_year = self.year_combo.currentData()
                base_month = self.month_combo.currentData()
                if base_year is None:
                    base_year = datetime.now().year
                if base_month is None or base_month == 0:
                    base_month = datetime.now().month
                
                for cell in valid_cells:
                    year = base_year
                    month = base_month
                    # 셀의 날짜 정보를 사용할 수 없으므로 현재 선택된 월 사용
                    
                    try:
                        work_date = datetime(year, month, cell['day']).date()
                    except ValueError:
                        continue
                    
                    # 기존 기록 확인
                    cursor.execute("""
                        SELECT arrival_time, departure_time, leave_type
                        FROM attendance_records
                        WHERE employee_id = ? AND work_date = ?
                    """, (cell['emp_id'], work_date))
                    existing = cursor.fetchone()
                    
                    if is_time:
                        # 시간 형식인 경우
                        try:
                            time_obj = datetime.strptime(formatted_time, "%H:%M").time()
                        except:
                            continue
                        
                        if cell['category'] == '출근':
                            arrival_time = time_obj
                            departure_time = None
                            if existing and existing[1]:
                                try:
                                    departure_time = datetime.strptime(existing[1], "%H:%M:%S").time()
                                except:
                                    pass
                        else:  # 퇴근
                            arrival_time = None
                            departure_time = time_obj
                            if existing and existing[0]:
                                try:
                                    arrival_time = datetime.strptime(existing[0], "%H:%M:%S").time()
                                except:
                                    pass
                        
                        self.calculator.process_attendance_record(
                            cell['emp_id'], work_date, arrival_time, departure_time,
                            None, ""
                        )
                    else:
                        # 휴가 유형인 경우
                        self.calculator.process_attendance_record(
                            cell['emp_id'], work_date, None, None,
                            formatted_time, ""
                        )
                    
                    updated_count += 1
                
                conn.commit()
                conn.close()
                
                QMessageBox.information(dialog, "완료", f"{updated_count}개 셀이 수정되었습니다.")
                dialog.accept()
                self.refresh_data()
                
            except Exception as e:
                QMessageBox.critical(dialog, "오류", f"수정 중 오류 발생: {str(e)}")
        
        # Enter 키로 저장
        time_entry.returnPressed.connect(save_time)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(save_time)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        time_entry.setFocus()
        dialog.exec()
    
    def upload_excel(self):
        """엑셀 파일 업로드 - 개선된 로직"""
        file_path, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 선택", "", "Excel files (*.xlsx *.xls);;All files (*.*)")
        if not file_path:
            return
        
        try:
            file_ext = Path(file_path).suffix.lower()
            if file_ext == '.xls':
                try:
                    df = pd.read_excel(file_path, header=None, engine='xlrd')
                except Exception as e:
                    QMessageBox.critical(self, "오류", f".xls 파일 읽기 오류: {str(e)}\n\nxlrd 라이브러리가 필요합니다: pip install xlrd>=2.0.1")
                    return
            else:
                df = pd.read_excel(file_path, header=None, engine='openpyxl')
            
            # 헤더 찾기
            name_col_idx = None
            category_col_idx = None
            date_cols = {}
            
            for idx, row in df.iterrows():
                for col_idx, cell in enumerate(row):
                    cell_str = str(cell).strip() if pd.notna(cell) else ""
                    if '이름' in cell_str or 'name' in cell_str.lower():
                        name_col_idx = col_idx
                    if '구분' in cell_str or 'category' in cell_str.lower():
                        category_col_idx = col_idx
                    try:
                        day = int(cell_str)
                        if 1 <= day <= 31:
                            date_cols[day] = col_idx
                    except:
                        pass
                
                if name_col_idx is not None and len(date_cols) > 0:
                    break
            
            if name_col_idx is None:
                QMessageBox.critical(self, "오류", "엑셀 파일에서 '이름' 컬럼을 찾을 수 없습니다.")
                return
            
            if not date_cols:
                QMessageBox.critical(self, "오류", "엑셀 파일에서 날짜 컬럼(1~31)을 찾을 수 없습니다.")
                return
            
            year = self.year_combo.currentData()
            month = self.month_combo.currentData()
            if year is None:
                year = datetime.now().year
            if month is None:
                month = datetime.now().month
            
            from calendar import monthrange
            days_in_month = monthrange(year, month)[1]
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT id, name FROM employees")
            employees = cursor.fetchall()
            employee_dict = {name: emp_id for emp_id, name in employees}
            # 전금희(지문) 매칭 처리
            employee_dict['전금희(지문)'] = employee_dict.get('전금희', None)
            if '전금희(지문)' in employee_dict and employee_dict['전금희(지문)'] is None:
                # 전금희가 없으면 전금희(지문)도 None으로 설정
                del employee_dict['전금희(지문)']
            
            records_added = 0
            current_name = None
            
            def format_time_input(time_value):
                """시간 형식 정규화: 다양한 입력 형식 -> HH:MM"""
                # None이나 빈 값 처리
                if time_value is None or (isinstance(time_value, str) and time_value.strip() == ''):
                    return None
                
                # datetime.time 객체인 경우 (직접)
                from datetime import time as dt_time
                if isinstance(time_value, dt_time):
                    return time_value.strftime("%H:%M")
                
                # datetime 객체인 경우
                if isinstance(time_value, datetime):
                    return time_value.time().strftime("%H:%M")
                
                # pandas Timestamp 객체인 경우
                if hasattr(time_value, 'time'):
                    try:
                        time_obj = time_value.time()
                        if isinstance(time_obj, dt_time):
                            return time_obj.strftime("%H:%M")
                        elif isinstance(time_obj, datetime):
                            return time_obj.time().strftime("%H:%M")
                    except:
                        pass
                
                # pandas의 datetime64 타입 처리
                try:
                    if isinstance(time_value, pd.Timestamp):
                        return time_value.to_pydatetime().time().strftime("%H:%M")
                except:
                    pass
                
                # 문자열로 변환
                time_str = str(time_value).strip()
                
                # 이미 HH:MM 형식인 경우
                if ':' in time_str:
                    parts = time_str.split(':')
                    if len(parts) >= 2:
                        try:
                            hour = int(parts[0])
                            minute = int(parts[1])
                            if 0 <= hour <= 23 and 0 <= minute <= 59:
                                return f"{hour:02d}:{minute:02d}"
                        except:
                            pass
                        # HH:MM:SS 형식
                        if len(parts) >= 3:
                            try:
                                hour = int(parts[0])
                                minute = int(parts[1])
                                if 0 <= hour <= 23 and 0 <= minute <= 59:
                                    return f"{hour:02d}:{minute:02d}"
                            except:
                                pass
                
                # 숫자 형식 (예: 1837 -> 18:37)
                if time_str.isdigit():
                    if len(time_str) == 4:
                        try:
                            hour = int(time_str[:2])
                            minute = int(time_str[2:])
                            if 0 <= hour <= 23 and 0 <= minute <= 59:
                                return f"{hour:02d}:{minute:02d}"
                        except:
                            pass
                    elif len(time_str) == 3:
                        try:
                            hour = int(time_str[0])
                            minute = int(time_str[1:])
                            if 0 <= hour <= 23 and 0 <= minute <= 59:
                                return f"0{hour}:{minute:02d}"
                        except:
                            pass
                
                return time_str
            
            def parse_time(time_value):
                """시간 값을 time 객체로 변환 (다양한 형식 지원)"""
                if time_value is None:
                    return None
                
                # datetime.time 객체인 경우 (직접)
                from datetime import time as dt_time
                if isinstance(time_value, dt_time):
                    return time_value
                
                # datetime 객체인 경우
                if isinstance(time_value, datetime):
                    return time_value.time()
                
                # pandas Timestamp 객체인 경우
                if hasattr(time_value, 'time'):
                    try:
                        time_obj = time_value.time()
                        if isinstance(time_obj, dt_time):
                            return time_obj
                        elif isinstance(time_obj, datetime):
                            return time_obj.time()
                    except:
                        pass
                
                # pandas의 datetime64 타입 처리
                try:
                    if isinstance(time_value, pd.Timestamp):
                        return time_value.to_pydatetime().time()
                except:
                    pass
                
                # 문자열로 변환
                time_str = format_time_input(time_value)
                if time_str is None:
                    return None
                
                # 여러 형식 시도
                for fmt in ["%H:%M", "%H:%M:%S"]:
                    try:
                        return datetime.strptime(time_str, fmt).time()
                    except:
                        continue
                
                return None
            
            for idx, row in df.iterrows():
                try:
                    name_cell = str(row[name_col_idx]).strip() if pd.notna(row[name_col_idx]) else ""
                    
                    category_cell = ""
                    if category_col_idx is not None:
                        category_cell = str(row[category_col_idx]).strip() if pd.notna(row[category_col_idx]) else ""
                    
                    is_arrival = (category_cell == '출근')
                    is_departure = (category_cell == '퇴근')
                    
                    if not is_arrival and not is_departure:
                        continue
                    
                    if name_cell and name_cell != 'nan' and name_cell not in ['출근', '퇴근']:
                        current_name = name_cell
                    
                    if current_name is None or current_name not in employee_dict:
                        continue
                    
                    emp_id = employee_dict[current_name]
                    
                    # 출근 행을 찾으면, 다음 행(퇴근 행)도 함께 확인
                    next_row = None
                    next_row_category = ""
                    if is_arrival and idx + 1 < len(df):
                        try:
                            next_row = df.iloc[idx + 1]
                            if category_col_idx is not None:
                                next_row_category = str(next_row[category_col_idx]).strip() if pd.notna(next_row[category_col_idx]) else ""
                        except:
                            pass
                    
                    # 출근 행이면서 다음 행이 퇴근인 경우, 함께 처리
                    if is_arrival and next_row_category == '퇴근':
                            # 출근 행과 퇴근 행을 함께 처리
                            for day in range(1, days_in_month + 1):
                                if day not in date_cols:
                                    continue
                                
                                col_idx = date_cols[day]
                                arrival_cell_value = row[col_idx]
                                departure_cell_value = next_row[col_idx]
                                
                                # 출근 시간 파싱
                                arrival_time = None
                                leave_type = None
                                
                                if not (pd.isna(arrival_cell_value) or (isinstance(arrival_cell_value, str) and arrival_cell_value.strip() == '')):
                                    arrival_cell_time_value = arrival_cell_value
                                    arrival_cell_str = format_time_input(arrival_cell_value)
                                    
                                    # 1단계: 원본 값 직접 파싱
                                    arrival_time = parse_time(arrival_cell_time_value)
                                    
                                    # 2단계: 포맷된 문자열로 파싱 시도
                                    if arrival_time is None and arrival_cell_str is not None:
                                        if ':' in arrival_cell_str:
                                            arrival_time = parse_time(arrival_cell_str)
                                        elif arrival_cell_str in ['연차', '반차', '반반차', '공휴', '박람회', '민방위', '출장', '교육', '추석', '설날']:
                                            leave_type = arrival_cell_str
                                    
                                    # 3단계: 원본 값의 문자열 표현으로 파싱 시도
                                    if arrival_time is None and leave_type is None:
                                        original_str = str(arrival_cell_value).strip()
                                        if original_str and original_str != 'nan' and original_str != 'None':
                                            arrival_time = parse_time(original_str)
                                    
                                    # 4단계: pandas의 경우 직접 타입 확인
                                    if arrival_time is None and leave_type is None:
                                        try:
                                            if isinstance(arrival_cell_value, pd.Timestamp):
                                                arrival_time = arrival_cell_value.to_pydatetime().time()
                                            elif isinstance(arrival_cell_value, (float, int)):
                                                if 0.0 <= arrival_cell_value < 1.0:
                                                    total_seconds = int(arrival_cell_value * 86400)
                                                    hours = total_seconds // 3600
                                                    minutes = (total_seconds % 3600) // 60
                                                    from datetime import time as dt_time
                                                    arrival_time = dt_time(hours, minutes)
                                        except:
                                            pass
                                
                                # 퇴근 시간 파싱
                                departure_time = None
                                
                                if not (pd.isna(departure_cell_value) or (isinstance(departure_cell_value, str) and departure_cell_value.strip() == '')):
                                    departure_cell_time_value = departure_cell_value
                                    departure_cell_str = format_time_input(departure_cell_value)
                                    
                                    # 1단계: 원본 값 직접 파싱
                                    departure_time = parse_time(departure_cell_time_value)
                                    
                                    # 2단계: 포맷된 문자열로 파싱 시도
                                    if departure_time is None and departure_cell_str is not None:
                                        departure_time = parse_time(departure_cell_str)
                                    
                                    # 3단계: 원본 값의 문자열 표현으로 파싱 시도
                                    if departure_time is None:
                                        original_str = str(departure_cell_value).strip()
                                        if original_str and original_str != 'nan' and original_str != 'None':
                                            departure_time = parse_time(original_str)
                                    
                                    # 4단계: pandas의 경우 직접 타입 확인
                                    if departure_time is None:
                                        try:
                                            if isinstance(departure_cell_value, pd.Timestamp):
                                                departure_time = departure_cell_value.to_pydatetime().time()
                                            elif isinstance(departure_cell_value, (float, int)):
                                                if 0.0 <= departure_cell_value < 1.0:
                                                    total_seconds = int(departure_cell_value * 86400)
                                                    hours = total_seconds // 3600
                                                    minutes = (total_seconds % 3600) // 60
                                                    from datetime import time as dt_time
                                                    departure_time = dt_time(hours, minutes)
                                        except:
                                            pass
                                
                                # 출근 시간이나 휴가 유형이 있는 경우만 처리 (각 날짜별로 처리)
                                if arrival_time is not None or leave_type is not None or departure_time is not None:
                                    try:
                                        work_date = datetime(year, month, day).date()
                                    except ValueError:
                                        continue
                                    
                                    # 기존 데이터 확인 - 기존 데이터와 병합하여 저장
                                    cursor.execute("""
                                        SELECT arrival_time, departure_time, leave_type
                                        FROM attendance_records
                                        WHERE employee_id = ? AND work_date = ?
                                    """, (emp_id, work_date))
                                    existing = cursor.fetchone()
                                    
                                    # 기존(수동 입력) 데이터는 유지하고, 비어 있는 값만 엑셀 값으로 채운다
                                    existing_arrival = None
                                    existing_departure = None
                                    existing_leave_type = None
                                    
                                    if existing:
                                        existing_arrival, existing_departure, existing_leave_type = existing
                                        # DB에는 "HH:MM:SS" 문자열로 저장되어 있으므로 time 객체로 변환
                                        try:
                                            if existing_arrival:
                                                existing_arrival = datetime.strptime(existing_arrival, "%H:%M:%S").time()
                                        except Exception:
                                            existing_arrival = None
                                        try:
                                            if existing_departure:
                                                existing_departure = datetime.strptime(existing_departure, "%H:%M:%S").time()
                                        except Exception:
                                            existing_departure = None
                                    
                                    # 기존 값이 있으면 그대로 두고, 없는 값만 엑셀 값으로 채운다
                                    final_arrival = existing_arrival or arrival_time
                                    final_departure = existing_departure or departure_time
                                    final_leave_type = (
                                        existing_leave_type if existing_leave_type not in (None, "")
                                        else leave_type
                                    )
                                    
                                    # 최종 값이 기존 값과 완전히 같다면 DB를 다시 쓰지 않는다
                                    if (
                                        not existing
                                        or final_arrival != existing_arrival
                                        or final_departure != existing_departure
                                        or final_leave_type != existing_leave_type
                                    ):
                                        if (
                                            final_arrival is not None
                                            or final_departure is not None
                                            or final_leave_type is not None
                                        ):
                                            # 수동 입력을 보존한 상태로 병합 결과 저장
                                            self.calculator.process_attendance_record(
                                                emp_id, work_date, final_arrival, final_departure,
                                                final_leave_type, ""
                                            )
                                        records_added += 1
                    
                    # 출근 행을 처리했으므로 다음 행(퇴근 행)은 건너뛰기
                    continue
                    
                    # 출근 행이지만 다음 행이 퇴근이 아니거나, 퇴근 행인 경우 (기존 로직)
                    for day in range(1, days_in_month + 1):
                        if day not in date_cols:
                            continue
                        
                        col_idx = date_cols[day]
                        cell_value = row[col_idx]
                    
                        # 빈 셀 처리: pd.isna 또는 빈 문자열이면 None
                        if pd.isna(cell_value) or (isinstance(cell_value, str) and cell_value.strip() == ''):
                            cell_str = None
                            cell_time_value = None
                        else:
                            # 시간 값은 원본과 포맷된 문자열 모두 저장
                            cell_time_value = cell_value
                            cell_str = format_time_input(cell_value)
                        
                        try:
                            work_date = datetime(year, month, day).date()
                        except ValueError:
                            continue
                        
                        # 각 날짜별로 출근/퇴근 데이터를 먼저 수집한 후 일괄 처리
                        # 하지만 현재 구조에서는 각 행을 순차 처리하므로,
                        # 출근 행과 퇴근 행을 각각 처리하되, 기존 데이터를 올바르게 병합해야 함
                        
                        if is_arrival:
                            arrival_time = None
                            leave_type = None
                            
                            if cell_time_value is not None:
                                # 1단계: 원본 값 직접 파싱
                                arrival_time = parse_time(cell_time_value)
                                
                                # 2단계: 포맷된 문자열로 파싱 시도
                                if arrival_time is None and cell_str is not None:
                                    # 시간 형식인지 확인
                                    if ':' in cell_str:
                                        arrival_time = parse_time(cell_str)
                                    # 휴가 유형인지 확인
                                    elif cell_str in ['연차', '반차', '반반차', '공휴', '박람회', '민방위', '출장', '교육', '추석', '설날']:
                                        leave_type = cell_str
                                
                                # 3단계: 원본 값의 문자열 표현으로 파싱 시도
                                if arrival_time is None and leave_type is None:
                                    original_str = str(cell_value).strip()
                                    if original_str and original_str != 'nan' and original_str != 'None':
                                        arrival_time = parse_time(original_str)
                                
                                # 4단계: pandas의 경우 직접 타입 확인
                                if arrival_time is None and leave_type is None:
                                    try:
                                        # pandas가 datetime64로 읽은 경우
                                        if isinstance(cell_value, pd.Timestamp):
                                            arrival_time = cell_value.to_pydatetime().time()
                                        # 또는 float/int로 읽은 경우 (엑셀의 시간 셀)
                                        elif isinstance(cell_value, (float, int)):
                                            # 엑셀의 시간 값은 0.0~1.0 사이의 소수로 저장됨
                                            if 0.0 <= cell_value < 1.0:
                                                total_seconds = int(cell_value * 86400)  # 하루 = 86400초
                                                hours = total_seconds // 3600
                                                minutes = (total_seconds % 3600) // 60
                                                from datetime import time as dt_time
                                                arrival_time = dt_time(hours, minutes)
                                    except Exception as e:
                                        pass
                            
                            # 출근 시간이나 휴가 유형이 있는 경우만 처리
                            if arrival_time is not None or leave_type is not None:
                                cursor.execute("""
                                    SELECT arrival_time, departure_time, leave_type
                                    FROM attendance_records
                                    WHERE employee_id = ? AND work_date = ?
                                """, (emp_id, work_date))
                                existing = cursor.fetchone()
                                
                                # 기존(수동 입력) 데이터를 우선으로 유지하고, 비어 있는 값만 엑셀 값으로 채운다
                                existing_arrival = None
                                existing_departure = None
                                existing_leave_type = None
                                
                                if existing:
                                    existing_arrival, existing_departure, existing_leave_type = existing
                                    try:
                                        if existing_arrival:
                                            existing_arrival = datetime.strptime(existing_arrival, "%H:%M:%S").time()
                                    except Exception:
                                        existing_arrival = None
                                    try:
                                        if existing_departure:
                                            existing_departure = datetime.strptime(existing_departure, "%H:%M:%S").time()
                                    except Exception:
                                        existing_departure = None
                                
                                # 출근은 엑셀 값이 우선이 아니라, 기존 값이 있으면 유지하고 없을 때만 채운다
                                final_arrival = existing_arrival or arrival_time
                                final_departure = existing_departure  # 이 분기에서는 새 퇴근 시간 없음
                                final_leave_type = (
                                    existing_leave_type if existing_leave_type not in (None, "")
                                    else leave_type
                                )
                                
                                # 실제로 변경이 발생한 경우에만 저장
                                if (
                                    not existing
                                    or final_arrival != existing_arrival
                                    or final_departure != existing_departure
                                    or final_leave_type != existing_leave_type
                                ):
                                    if final_arrival is not None or final_leave_type is not None or final_departure is not None:
                                        self.calculator.process_attendance_record(
                                            emp_id, work_date, final_arrival, final_departure,
                                            final_leave_type, ""
                                        )
                                    records_added += 1
                            # 빈 셀인 경우는 건너뛰기 (출근 시간이 없으면 기록하지 않음)
                        
                        elif is_departure:
                            departure_time = None
                            
                            # 퇴근 시간 파싱 시도 (엑셀 셀에 값이 있는 경우)
                            if cell_time_value is not None:
                                # 1단계: 원본 값 직접 파싱
                                departure_time = parse_time(cell_time_value)
                                
                                # 2단계: 포맷된 문자열로 파싱 시도
                                if departure_time is None and cell_str is not None:
                                    departure_time = parse_time(cell_str)
                                
                                # 3단계: 원본 값의 문자열 표현으로 파싱 시도
                                if departure_time is None:
                                    original_str = str(cell_value).strip()
                                    if original_str and original_str != 'nan' and original_str != 'None':
                                        departure_time = parse_time(original_str)
                                
                                # 4단계: pandas의 경우 직접 타입 확인
                                if departure_time is None:
                                    try:
                                        # pandas가 datetime64로 읽은 경우
                                        if isinstance(cell_value, pd.Timestamp):
                                            departure_time = cell_value.to_pydatetime().time()
                                        # 또는 float/int로 읽은 경우 (엑셀의 시간 셀)
                                        elif isinstance(cell_value, (float, int)):
                                            # 엑셀의 시간 값은 0.0~1.0 사이의 소수로 저장됨
                                            if 0.0 <= cell_value < 1.0:
                                                total_seconds = int(cell_value * 86400)  # 하루 = 86400초
                                                hours = total_seconds // 3600
                                                minutes = (total_seconds % 3600) // 60
                                                from datetime import time as dt_time
                                                departure_time = dt_time(hours, minutes)
                                    except Exception as e:
                                        pass
                                
                                # 디버깅: 파싱 성공 여부 확인
                                if departure_time is not None:
                                    print(f"✅ 퇴근 시간 파싱 성공: 직원={current_name}, 날짜={work_date}, "
                                          f"시간={departure_time}, 원본값={cell_value}, 타입={type(cell_value).__name__}")
                                else:
                                    print(f"⚠️ 퇴근 시간 파싱 실패: 직원={current_name}, 날짜={work_date}, "
                                          f"원본값={cell_value}, 타입={type(cell_value).__name__}, "
                                          f"포맷된문자열={cell_str}")
                            
                            # 퇴근 행 처리 - 기존 데이터와 병합하여 저장
                            if departure_time is not None:
                                cursor.execute("""
                                    SELECT arrival_time, departure_time, leave_type
                                    FROM attendance_records
                                    WHERE employee_id = ? AND work_date = ?
                                """, (emp_id, work_date))
                                existing = cursor.fetchone()
                                
                                existing_arrival = None
                                existing_departure = None
                                existing_leave_type = None
                                
                                if existing:
                                    existing_arrival, existing_departure, existing_leave_type = existing
                                    try:
                                        if existing_arrival:
                                            existing_arrival = datetime.strptime(existing_arrival, "%H:%M:%S").time()
                                    except Exception:
                                        existing_arrival = None
                                    try:
                                        if existing_departure:
                                            existing_departure = datetime.strptime(existing_departure, "%H:%M:%S").time()
                                    except Exception:
                                        existing_departure = None
                                
                                # 기존 퇴근 시간이 있으면 유지, 없을 때만 엑셀 값으로 채운다
                                final_arrival = existing_arrival
                                final_departure = existing_departure or departure_time
                                final_leave_type = existing_leave_type
                                
                                if (
                                    not existing
                                    or final_departure != existing_departure
                                    or final_arrival != existing_arrival
                                    or final_leave_type != existing_leave_type
                                ):
                                    if final_departure is not None or final_arrival is not None or final_leave_type is not None:
                                        self.calculator.process_attendance_record(
                                            emp_id, work_date, final_arrival, final_departure,
                                            final_leave_type, ""
                                        )
                                    records_added += 1
                
                except Exception as e:
                    print(f"행 {idx + 2} 처리 중 오류: {str(e)}")
                    continue
            
            conn.commit()
            conn.close()
            
            QMessageBox.information(self, "업로드 완료", f"엑셀 업로드 완료!\n\n추가/업데이트된 기록: {records_added}건")
            self.refresh_data()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 파일 업로드 중 오류 발생: {str(e)}")
    
    def download_excel(self):
        """엑셀 파일 다운로드 - 전체 데이터를 월별 시트로 다운로드"""
        return self._download_attendance_excel(file_path_override=None, silent=False, open_after=True)

    def download_combined_excel(self):
        """연월차 + 출퇴근을 한 파일에 시트로 묶어서 다운로드 (출퇴근 탭에서 눌러도 동일 동작)"""
        if hasattr(self, "leave_gui") and self.leave_gui is not None and hasattr(self.leave_gui, "download_combined_excel"):
            return self.leave_gui.download_combined_excel()
        return self.download_excel()

    def _download_attendance_excel(self, file_path_override=None, silent=False, open_after=True):
        """(내부용) 출퇴근 엑셀 생성. file_path_override가 있으면 다이얼로그 없이 저장."""
        if file_path_override:
            file_path = file_path_override
        else:
            file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일 저장", "", "Excel files (*.xlsx);;All files (*.*)")
            if not file_path:
                return
        
        try:
            import os
            # 파일이 이미 존재하는 경우 삭제 (덮어쓰기)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except PermissionError:
                    QMessageBox.warning(self, "경고", f"파일이 다른 프로그램에서 열려있습니다.\n파일을 닫고 다시 시도해주세요.\n\n파일: {file_path}")
                    return
            
            # 조회에서 선택한 년도와 월 가져오기
            year = self.year_combo.currentData()
            month = self.month_combo.currentData()
            
            if year is None:
                year = datetime.now().year
            if month is None:
                month = datetime.now().month
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # 2025년은 월별(1~12월) 시트를 모두 생성
            # 그 외에는 기존처럼 선택한 월만 생성 (동작 변경 최소화)
            if int(year) == 2025:
                year_months = [(f"{year}-{m:02d}",) for m in range(1, 13)]
                # 연도 기준으로 데이터 존재 여부만 확인 (없으면 생성 중단)
                cursor.execute("""
                    SELECT COUNT(*)
                    FROM attendance_records
                    WHERE strftime('%Y', work_date) = ?
                """, (str(year),))
                data_count = cursor.fetchone()[0]
                if data_count == 0:
                    QMessageBox.warning(self, "알림", f"{year}년 출퇴근 데이터가 없습니다.")
                    conn.close()
                    return
            else:
                # 선택한 년도-월만 사용
                year_month = f"{year}-{month:02d}"
                year_months = [(year_month,)]
                
                # 해당 월에 데이터가 있는지 확인
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM attendance_records
                    WHERE strftime('%Y-%m', work_date) = ?
                """, (year_month,))
                data_count = cursor.fetchone()[0]
                
                if data_count == 0:
                    QMessageBox.warning(self, "알림", f"{year}년 {month}월 출퇴근 데이터가 없습니다.")
                    conn.close()
                    return
            
            # 모든 직원 정보 조회 (부서별 정렬, 퇴사자 포함)
            cursor.execute("""
                SELECT e.id, e.department, e.position, e.name, e.hire_date,
                       COALESCE(e.display_order, 0) as display_order
                FROM employees e
                ORDER BY
                    CASE e.department
                        WHEN '경영지원팀' THEN 1
                        WHEN '영업팀' THEN 2
                        WHEN '글로벌비즈니스팀' THEN 3
                        ELSE 999
                    END,
                    e.department,
                    CASE e.position
                        WHEN '이사' THEN 1
                        WHEN '팀장' THEN 2
                        WHEN '파트장' THEN 3
                        WHEN '과장' THEN 4
                        WHEN '대리' THEN 5
                        WHEN '프로' THEN 6
                        ELSE 999
                    END,
                    e.hire_date ASC
            """)
            employees = cursor.fetchall()
            
            from calendar import monthrange
            
            # openpyxl Workbook 직접 사용
            workbook = openpyxl.Workbook()
            # 기본 시트 제거 (나중에 생성할 시트로 대체)
            workbook.remove(workbook.active)
            
            # 각 년도-월별로 시트 생성
            for year_month_tuple in year_months:
                year_month = year_month_tuple[0]
                year, month = map(int, year_month.split('-'))
                days_in_month = monthrange(year, month)[1]
                
                # 프로그램과 동일한 컬럼 구조 (헤더는 별도로 처리)
                date_columns = [str(i) for i in range(1, 32)]
                summary_columns = ["조기출근(8시이전)", "지각(9시이후)", "야근(20시이후)", "연차사용", "평균 출근시간", "평균 퇴근시간"]
                columns = ["직급", "이름", "구분"] + date_columns + summary_columns
                
                # 데이터 준비
                data = []
                merge_info = []
                merge_info_set = set()
                def _add_merge(m_row, m_col, r_span, c_span):
                    """중복 병합 방지용"""
                    t = (m_row, m_col, r_span, c_span)
                    if t in merge_info_set:
                        return
                    merge_info_set.add(t)
                    merge_info.append(t)
                diagonal_cells = []
                current_department = None
                row_idx = 0
                employee_departure_rows = []
                
                for emp_id, dept, pos, name, hire_date_str, display_order in employees:
                    # 입사일 파싱
                    hire_date = None
                    if hire_date_str:
                        try:
                            if isinstance(hire_date_str, str):
                                hire_date = datetime.strptime(hire_date_str, "%Y-%m-%d").date()
                            else: hire_date = hire_date_str
                        except: pass
                    
                    # 부서 변경 체크 (구분자 행은 엑셀에서 제외)
                    if current_department != dept:
                        current_department = dept
                    
                    cursor.execute("""
                        SELECT work_date, arrival_time, departure_time, early_arrival, late_arrival, late_departure, leave_type, remarks
                        FROM attendance_records
                        WHERE employee_id = ? AND strftime('%Y-%m', work_date) = ?
                        ORDER BY work_date
                    """, (emp_id, year_month))
                    
                    records = cursor.fetchall()
                    records_dict = {}
                    for record in records:
                        work_date, arr, dep, early, late_arr, late_dep, l_type, remarks = record
                        day = int(work_date.split('-')[2]) if isinstance(work_date, str) else work_date.day
                        records_dict[day] = {'arrival': arr, 'departure': dep, 'early': early, 'late_arr': late_arr, 'late_dep': late_dep, 'leave_type': l_type, 'remarks': remarks}
                    
                    # 출근 행 추가
                    arrival_row_idx = row_idx
                    arrival_row = {"직급": pos, "이름": name, "구분": "출근"}
                    arrival_times = []
                    
                    for day in range(1, days_in_month + 1):
                        if day in records_dict:
                            record = records_dict[day]
                            rmk, l_type = record.get('remarks', ''), record.get('leave_type')
                            if l_type and rmk == f'{l_type}_출근': arrival_row[str(day)] = l_type
                            elif l_type and rmk == f'{l_type}_퇴근':
                                if record['arrival']:
                                    arr_s = str(record['arrival'])[:5]
                                    arrival_row[str(day)] = arr_s
                                    try:
                                        at = datetime.strptime(arr_s, "%H:%M").time()
                                        if datetime(year, month, day).weekday() < 5: arrival_times.append(at)
                                    except: pass
                                else: arrival_row[str(day)] = ""
                            elif l_type and not rmk: arrival_row[str(day)] = l_type
                            elif record['arrival']:
                                arr_s = str(record['arrival'])[:5]
                                arrival_row[str(day)] = arr_s
                                try:
                                    at = datetime.strptime(arr_s, "%H:%M").time()
                                    if datetime(year, month, day).weekday() < 5: arrival_times.append(at)
                                except: pass
                            else: arrival_row[str(day)] = ""
                        else: arrival_row[str(day)] = ""
                    for day in range(days_in_month + 1, 32): arrival_row[str(day)] = ""
                    
                    early_count = sum(1 for r in records_dict.values() if r.get('early'))
                    late_arr_count = sum(1 for r in records_dict.values() if r.get('late_arr'))
                    leave_amount = 0.0
                    for r in records_dict.values():
                        lt = r.get('leave_type')
                        if lt in ['연차', '휴가']: leave_amount += 1.0
                        elif lt == '반차': leave_amount += 0.5
                    
                    avg_arr = ""
                    if arrival_times:
                        ts = sum(t.hour * 3600 + t.minute * 60 + t.second for t in arrival_times)
                        av = ts // len(arrival_times)
                        avg_arr = f"{av // 3600:02d}:{(av % 3600) // 60:02d}"
                    
                    arrival_row["조기출근(8시이전)"] = str(early_count) if early_count > 0 else ""
                    arrival_row["지각(9시이후)"] = str(late_arr_count) if late_arr_count > 0 else ""
                    arrival_row["야근(20시이후)"] = ""
                    arrival_row["연차사용"] = str(leave_amount) if leave_amount > 0 else ""
                    arrival_row["평균 출근시간"] = avg_arr
                    arrival_row["평균 퇴근시간"] = ""
                    data.append(arrival_row)
                    row_idx += 1
                    
                    # 퇴근 행 추가
                    departure_row_idx = row_idx
                    departure_row = {"직급": "", "이름": "", "구분": "퇴근"}
                    departure_times = []
                    
                    for day in range(1, days_in_month + 1):
                        if day in records_dict:
                            record = records_dict[day]
                            rmk, l_type = record.get('remarks', ''), record.get('leave_type')
                            if l_type and rmk == f'{l_type}_퇴근': departure_row[str(day)] = l_type
                            elif l_type and rmk == f'{l_type}_출근':
                                if record['departure']:
                                    dep_s = str(record['departure'])[:5]
                                    departure_row[str(day)] = dep_s
                                    try:
                                        dt = datetime.strptime(dep_s, "%H:%M").time()
                                        d_obj = datetime(year, month, day).date()
                                        if d_obj.weekday() < 5 and not self.is_third_wednesday_17_00(d_obj, dt):
                                            departure_times.append(dt)
                                    except: pass
                                else: departure_row[str(day)] = ""
                            elif l_type and not rmk: departure_row[str(day)] = l_type
                            elif record['departure']:
                                dep_s = str(record['departure'])[:5]
                                departure_row[str(day)] = dep_s
                                try:
                                    dt = datetime.strptime(dep_s, "%H:%M").time()
                                    d_obj = datetime(year, month, day).date()
                                    if d_obj.weekday() < 5 and not self.is_third_wednesday_17_00(d_obj, dt):
                                        departure_times.append(dt)
                                except: pass
                            else: departure_row[str(day)] = ""
                        else: departure_row[str(day)] = ""
                    for day in range(days_in_month + 1, 32): departure_row[str(day)] = ""
                    
                    late_dep_count = sum(1 for r in records_dict.values() if r.get('late_dep'))
                    avg_dep = ""
                    if departure_times:
                        ts = sum(t.hour * 3600 + t.minute * 60 + t.second for t in departure_times)
                        av = ts // len(departure_times)
                        avg_dep = f"{av // 3600:02d}:{(av % 3600) // 60:02d}"
                    
                    departure_row["조기출근(8시이전)"] = ""
                    departure_row["지각(9시이후)"] = ""
                    departure_row["야근(20시이후)"] = str(late_dep_count) if late_dep_count > 0 else ""
                    departure_row["연차사용"] = ""
                    departure_row["평균 출근시간"] = ""
                    departure_row["평균 퇴근시간"] = avg_dep
                    data.append(departure_row)
                    employee_departure_rows.append(row_idx)
                    row_idx += 1
                    
                    # 병합 정보 저장
                    _add_merge(arrival_row_idx, 0, 2, 1) # 직급
                    _add_merge(arrival_row_idx, 1, 2, 1) # 이름
                    sum_col = 34
                    _add_merge(arrival_row_idx, sum_col, 2, 1)
                    _add_merge(arrival_row_idx, sum_col + 1, 2, 1)
                    if late_dep_count > 0: data[arrival_row_idx]["야근(20시이후)"] = str(late_dep_count)
                    _add_merge(arrival_row_idx, sum_col + 2, 2, 1)
                    _add_merge(arrival_row_idx, sum_col + 3, 2, 1)
                    _add_merge(arrival_row_idx, sum_col + 4, 2, 1)
                    if avg_dep: data[arrival_row_idx]["평균 퇴근시간"] = avg_dep
                    _add_merge(arrival_row_idx, sum_col + 5, 2, 1)
                    
                    for day in range(1, days_in_month + 1):
                        if hire_date:
                            try:
                                if datetime(year, month, day).date() < hire_date:
                                    _add_merge(arrival_row_idx, 2 + day, 2, 1)
                                    diagonal_cells.append((arrival_row_idx, 2 + day))
                                    data[arrival_row_idx][str(day)] = ""
                                    data[departure_row_idx][str(day)] = ""
                            except: pass
                
                # 시트 이름을 월만 표시 (예: "1월", "2월")
                sheet_name = f"{month}월"[:31]
                worksheet = workbook.create_sheet(title=sheet_name)
                excel_row = 4
                excel_row_map = {}
                for i, row_data in enumerate(data):
                    excel_row_map[i] = excel_row
                    for c_idx, col_name in enumerate(columns, 1):
                        worksheet.cell(row=excel_row, column=c_idx).value = row_data.get(col_name, "")
                    excel_row += 1
                
                worksheet.sheet_view.showGridLines = False
                worksheet.sheet_view.zoomScale = 100
                worksheet.row_dimensions[1].height = 25
                worksheet.merge_cells('D1:I1')
                worksheet.cell(row=1, column=4).value = f"{year}년 {month}월"
                worksheet.cell(row=1, column=4).font = Font(name='맑은 고딕', size=12, bold=True)
                worksheet.cell(row=1, column=4).alignment = Alignment(horizontal='center', vertical='center')
                
                worksheet.row_dimensions[2].height = 17.4
                worksheet.row_dimensions[3].height = 17.4
                for col_idx, text in enumerate(['직급', '이름', '구분'], 1):
                    worksheet.cell(row=2, column=col_idx).value = text
                    worksheet.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
                
                wd_names = ['월', '화', '수', '목', '금', '토', '일']
                for day in range(1, days_in_month + 1):
                    col_idx = 3 + day
                    try:
                        wd = datetime(year, month, day).weekday()
                        worksheet.cell(row=2, column=col_idx).value = f"{day}\n{wd_names[wd]}"
                        worksheet.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
                    except: pass
                
                sum_hdrs = ["조기출근(8시이전)", "지각(9시이후)", "야근(20시이후)", "연차사용", "평균 출근시간", "평균 퇴근시간"]
                for idx, text in enumerate(sum_hdrs):
                    col_idx = 35 + idx
                    worksheet.cell(row=2, column=col_idx).value = text
                    worksheet.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
                
                for m_row, m_col, r_span, c_span in merge_info:
                    if m_row in excel_row_map:
                        e_r = excel_row_map[m_row]
                        # 중복 병합/겹침 병합은 openpyxl에서 예외가 날 수 있어 안전하게 처리
                        try:
                            worksheet.merge_cells(start_row=e_r, start_column=m_col + 1, end_row=e_r + r_span - 1, end_column=m_col + c_span)
                        except Exception:
                            pass

                # 3.5 출근행/퇴근행 동일 텍스트(예: 추석/연차)가 연속되는 구간은 "가로"로도 병합
                # - 날짜 영역(D~말일)만 대상으로 하며, 값이 비어있지 않고(공백 제외) 출근/퇴근이 동일할 때만 병합합니다.
                # - 기존 단일 셀 병합(2행×1열)도 이 로직에서 자연스럽게 포함됩니다.
                try:
                    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    start_day_col = 4
                    end_day_col = 3 + days_in_month
                    # data는 항상 [출근, 퇴근] 2행이 한 묶음
                    for i in range(0, len(data), 2):
                        if i + 1 >= len(data):
                            break
                        top_r = excel_row_map.get(i)
                        bot_r = excel_row_map.get(i + 1)
                        if not top_r or not bot_r:
                            continue

                        run_val = None
                        run_start = None

                        def _flush(run_end_col_exclusive: int):
                            nonlocal run_val, run_start
                            if run_val is None or run_start is None:
                                run_val, run_start = None, None
                                return
                            run_end = run_end_col_exclusive - 1
                            if run_end < run_start:
                                run_val, run_start = None, None
                                return

                            # 기존 병합과 겹치면 스킵(겹침 병합 예외 방지)
                            intersects = False
                            for rr in (top_r, bot_r):
                                for cc in range(run_start, run_end + 1):
                                    try:
                                        coord = f"{get_column_letter(cc)}{rr}"
                                        if coord in worksheet.merged_cells:
                                            intersects = True
                                            break
                                    except Exception:
                                        pass
                                if intersects:
                                    break
                            if intersects:
                                run_val, run_start = None, None
                                return

                            # 병합 전에 하위/우측 셀 값을 비움(병합 후 MergedCell value 대입 불가)
                            try:
                                # 하단행(퇴근행) run_start 포함 전체 비움
                                for cc in range(run_start, run_end + 1):
                                    worksheet.cell(row=bot_r, column=cc).value = ""
                                # 상단행은 run_start 제외 비움
                                for cc in range(run_start + 1, run_end + 1):
                                    worksheet.cell(row=top_r, column=cc).value = ""

                                worksheet.merge_cells(start_row=top_r, start_column=run_start, end_row=bot_r, end_column=run_end)
                                master = worksheet.cell(row=top_r, column=run_start)
                                master.value = run_val
                                master.alignment = align_center
                            except Exception:
                                pass
                            finally:
                                run_val, run_start = None, None

                        for col in range(start_day_col, end_day_col + 2):  # +sentinel
                            if col <= end_day_col:
                                v1 = worksheet.cell(row=top_r, column=col).value
                                v2 = worksheet.cell(row=bot_r, column=col).value
                                s1 = str(v1).strip() if v1 is not None else ""
                                s2 = str(v2).strip() if v2 is not None else ""
                                val = s1 if (s1 and s1 == s2) else None
                            else:
                                val = None

                            if val is None:
                                _flush(col)
                            else:
                                if run_val is None:
                                    run_val = val
                                    run_start = col
                                elif val != run_val:
                                    _flush(col)
                                    run_val = val
                                    run_start = col
                except Exception:
                    pass
                
                for day in range(1, 32): worksheet.column_dimensions[get_column_letter(3 + day)].width = 6.1
                
                # 4. 전체 스타일 및 테두리 최종 적용 (A1 ~ AN 마지막행까지 전수 주입)
                border_limit_row = max(excel_row - 1, 45)
                
                # Side 도구 설정
                # 일부 엑셀 환경에서 rgb 색상 지정이 무시되는 케이스가 있어, "indexed color"를 사용합니다.
                from openpyxl.styles.colors import Color
                thin_side = Side(style='thin', color=Color(indexed=8))   # black
                outer_side = Side(style='thick', color=Color(indexed=8)) # black
                diag_side = Side(style='thin', color=Color(indexed=22))  # gray
                # Border 객체는 캐싱해서 재사용 (스타일 수 폭증/표시 불안정 방지)
                border_cache = {}

                # 스타일 객체는 "재사용"해야 엑셀에서 안정적으로 렌더링됩니다. (스타일 수 폭증 방지)
                # Excel 호환성이 가장 좋은 fgColor 방식 사용
                # 토/일 배경색: 요청에 따라 #D0CECE 로 처리
                # - openpyxl에서 rgb는 ARGB(8자리)가 안전합니다. (#D0CECE = FFD0CECE)
                fill_weekend = PatternFill(patternType="solid", fgColor="FFD0CECE")
                fill_diag = PatternFill(patternType="solid", fgColor=Color(indexed=23))     # light gray(variant)
                align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
                align_data = Alignment(horizontal='center', vertical='center', wrap_text=False)
                align_data_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

                font_header = Font(name='맑은 고딕', size=10, bold=True)
                font_header_sat = Font(name='맑은 고딕', size=10, color="FF0000FF", bold=True)
                font_header_sun = Font(name='맑은 고딕', size=10, color="FFFF0000", bold=True)
                font_data = Font(name='맑은 고딕', size=10, bold=False)
                font_data_bold = Font(name='맑은 고딕', size=10, bold=True)

                weekend_days = set()
                try:
                    for d in range(1, days_in_month + 1):
                        wd = datetime(year, month, d).weekday()  # 5=토, 6=일
                        if wd in (5, 6):
                            weekend_days.add(d)
                except Exception:
                    weekend_days = set()

                def _get_border(l, r, t, b, diag: bool) -> Border:
                    key = (l.style, getattr(l.color, "rgb", None), r.style, getattr(r.color, "rgb", None),
                           t.style, getattr(t.color, "rgb", None), b.style, getattr(b.color, "rgb", None), diag)
                    cached = border_cache.get(key)
                    if cached is not None:
                        return cached
                    if diag:
                        cached = Border(left=l, right=r, top=t, bottom=b, diagonal=diag_side, diagonalDown=True)
                    else:
                        cached = Border(left=l, right=r, top=t, bottom=b)
                    border_cache[key] = cached
                    return cached
                
                # 퇴근 행 데이터 준비
                dep_rows_excel = {excel_row_map[i] for i in employee_departure_rows if i in excel_row_map}
                
                # 모든 시트 내 셀을 순회하며 스타일 강제 주입
                for r in range(1, border_limit_row + 1):
                    for c in range(1, 41):  # A(1) ~ AN(40)
                        cell = worksheet.cell(row=r, column=c)
                        
                        # [1] Row 1 (D1:I1 제목 영역) 특별 처리
                        if r == 1:
                            if 4 <= c <= 9:
                                cell.border = Border(
                                    top=outer_side, bottom=outer_side,
                                    left=outer_side if c == 4 else None,
                                    right=outer_side if c == 9 else None
                                )
                            continue

                        # [2] 폰트 및 정렬 설정 (객체 재사용)
                        if r <= 3:
                            # 헤더 (2~3행)
                            if 4 <= c <= 3 + days_in_month:
                                try:
                                    wd = datetime(year, month, c - 3).weekday()
                                    if wd == 5:
                                        cell.font = font_header_sat
                                    elif wd == 6:
                                        cell.font = font_header_sun
                                    else:
                                        cell.font = font_header
                                except: pass
                            if cell.font is None:
                                cell.font = font_header
                            cell.alignment = align_header
                        else:
                            # 데이터 (4행 이상)
                            # 굵게 여부만 유지하고, 폰트 객체는 재사용
                            is_bold = False
                            try:
                                if cell.font and getattr(cell.font, 'bold', False):
                                    is_bold = True
                            except Exception:
                                pass
                            cell.font = font_data_bold if is_bold else font_data
                            cell.alignment = align_data_wrap if (c in [2, 3]) else align_data

                        # 배경색: 기본 흰색은 굳이 셀마다 채우지 않음(스타일 폭증 방지)
                        # 토/일 날짜 컬럼은 연한 회색 음영 처리 (사선/특수 채움은 아래에서 덮어씀)
                        if r >= 2 and 4 <= c <= 3 + days_in_month:
                            if (c - 3) in weekend_days:
                                cell.fill = fill_weekend

                        # [3] 테두리 결정 (캐시된 Border 재사용)
                        l_b = outer_side if c == 1 else thin_side
                        r_b = outer_side if c == 40 else thin_side
                        t_b = outer_side if r == 2 else thin_side
                        b_b = outer_side if (r == border_limit_row or r in dep_rows_excel) else thin_side
                        
                        # [4] 사선 및 최종 테두리 주입
                        is_diag_cell = False
                        for dr, dc in diagonal_cells:
                            if dr in excel_row_map:
                                base_r = excel_row_map[dr]
                                if (r == base_r or r == base_r + 1) and c == dc + 1:
                                    is_diag_cell = True; break
                        
                        if is_diag_cell:
                            cell.border = _get_border(l_b, r_b, t_b, b_b, True)
                            cell.fill = fill_diag
                        else:
                            cell.border = _get_border(l_b, r_b, t_b, b_b, False)

            # 파일 저장
            workbook.save(file_path)

            # --- 저장 후 테두리 "실제 기록 여부"를 항상 표시 (사용자 환경에서 원인 분리용) ---
            # 여기서 읽힌 값이 존재하는데도 엑셀에서 안 보이면, 다른 파일을 열었거나 뷰어/캐시 문제일 가능성이 큼.
            verify_msg = f"\n\n[저장 경로]\n- {file_path}"
            try:
                wb_check = openpyxl.load_workbook(file_path)
                # 사용자가 실제로 확인하는 시트를 우선 체크
                # 시트 이름을 월만 표시하므로 검증 대상도 동일 규칙 사용
                # 2025년처럼 월별 12시트를 생성하는 경우에도 "1월"이 항상 존재
                if int(year) == 2025:
                    expected_sheet = "1월"
                else:
                    expected_sheet = f"{month}월"
                expected_sheet = expected_sheet[:31]
                ws_check = wb_check[expected_sheet] if expected_sheet in wb_check.sheetnames else wb_check.active

                b = ws_check.cell(row=2, column=1).border  # A2
                left = getattr(b, "left", None)
                top = getattr(b, "top", None)
                left_style = getattr(left, "style", None)
                top_style = getattr(top, "style", None)
                left_color = getattr(getattr(left, "color", None), "rgb", None)
                top_color = getattr(getattr(top, "color", None), "rgb", None)
                verify_msg = (
                    f"{verify_msg}\n\n[테두리 검증]\n"
                    f"- 검사 시트: {ws_check.title}\n"
                    f"- A2 left: {left_style} / {left_color}\n"
                    f"- A2 top : {top_style} / {top_color}"
                )

                # 토/일 음영(fill)도 같이 검증
                try:
                    from calendar import monthrange as _mr
                    _days = _mr(int(year), int(month))[1]
                    weekend_day = None
                    for d in range(1, _days + 1):
                        if datetime(int(year), int(month), d).weekday() in (5, 6):
                            weekend_day = d
                            break
                    if weekend_day is not None:
                        # 날짜 컬럼은 D열부터(=4) 시작 => col = 3 + day
                        col_idx = 3 + weekend_day
                        # 헤더행(2행)과 데이터행(4행이 있으면 4행)에서 fill 확인
                        r_data = 4 if ws_check.max_row >= 4 else 2
                        c_header = ws_check.cell(row=2, column=col_idx)
                        c_data = ws_check.cell(row=r_data, column=col_idx)
                        h_fill = getattr(getattr(c_header.fill, "start_color", None), "rgb", None)
                        d_fill = getattr(getattr(c_data.fill, "start_color", None), "rgb", None)
                        verify_msg += (
                            f"\n\n[토/일 음영 검증]\n"
                            f"- 첫 토/일: {weekend_day}일\n"
                            f"- 헤더(2행) fill: {h_fill}\n"
                            f"- 데이터({r_data}행) fill: {d_fill}"
                        )
                except Exception:
                    pass
                if not left_style and not top_style:
                    verify_msg += (
                        "\n\n※ 테두리 정보가 파일에 기록되지 않았습니다.\n"
                        "수정한 .py가 아니라 다른 실행파일(exe)로 실행 중이거나,\n"
                        "다른 파일을 열어 확인 중일 수 있습니다."
                    )
            except Exception:
                verify_msg = f"{verify_msg}\n\n[테두리 검증] 실패(파일 재열기 불가)"
            
            conn.close()
            
            if not silent:
                QMessageBox.information(self, "성공",
                                        f"엑셀 파일이 생성되었습니다.\n총 {len(year_months)}개의 시트가 생성되었습니다."
                                        f"{verify_msg}")
            if open_after:
                try:
                    import os
                    os.startfile(file_path)
                except Exception:
                    pass
        except Exception as e:
            if not silent:
                QMessageBox.critical(self, "오류", f"엑셀 파일 생성 중 오류 발생: {str(e)}")
            else:
                raise
    
    def register_attendance(self):
        """출퇴근 등록"""
        dialog = QDialog(self)
        dialog.setWindowTitle("출퇴근 등록")
        dialog.setModal(True)
        layout = QVBoxLayout(dialog)
        
        employee_layout = QHBoxLayout()
        employee_layout.addWidget(QLabel("직원:"))
        employee_combo = QComboBox()
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM employees ORDER BY name")
        employees = cursor.fetchall()
        employee_dict = {}
        for emp_id, name in employees:
            employee_combo.addItem(name)
            employee_dict[name] = emp_id
        conn.close()
        
        employee_layout.addWidget(employee_combo)
        layout.addLayout(employee_layout)
        
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("날짜:"))
        date_edit = QDateEdit()
        date_edit.setDate(QDate.currentDate())
        date_edit.setCalendarPopup(True)
        date_layout.addWidget(date_edit)
        layout.addLayout(date_layout)
        
        arrival_layout = QHBoxLayout()
        arrival_layout.addWidget(QLabel("출근 시간:"))
        arrival_entry = QLineEdit()
        arrival_entry.setPlaceholderText("HH:MM (예: 09:00)")
        arrival_layout.addWidget(arrival_entry)
        layout.addLayout(arrival_layout)
        
        departure_layout = QHBoxLayout()
        departure_layout.addWidget(QLabel("퇴근 시간:"))
        departure_entry = QLineEdit()
        departure_entry.setPlaceholderText("HH:MM (예: 18:00)")
        departure_layout.addWidget(departure_entry)
        layout.addLayout(departure_layout)
        
        def save_attendance():
            try:
                emp_id = employee_dict[employee_combo.currentText()]
                work_date = date_edit.date().toPython()
                
                arrival_time = None
                if arrival_entry.text().strip():
                    try:
                        arrival_time = datetime.strptime(arrival_entry.text().strip(), "%H:%M").time()
                    except:
                        QMessageBox.critical(self, "오류", "출근 시간 형식이 올바르지 않습니다. (HH:MM 형식)")
                        return
                
                departure_time = None
                if departure_entry.text().strip():
                    try:
                        departure_time = datetime.strptime(departure_entry.text().strip(), "%H:%M").time()
                    except:
                        QMessageBox.critical(self, "오류", "퇴근 시간 형식이 올바르지 않습니다. (HH:MM 형식)")
                        return
                
                self.calculator.process_attendance_record(
                    emp_id, work_date, arrival_time, departure_time, None, ""
                )
                
                QMessageBox.information(self, "성공", "출퇴근이 등록되었습니다.")
                dialog.accept()
                self.refresh_data()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"등록 중 오류 발생: {str(e)}")
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(save_attendance)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        dialog.exec()
    
    def delete_all_data(self):
        """현재 조회 월의 출퇴근 데이터 전체 삭제"""
        year = self.year_combo.currentData()
        month = self.month_combo.currentData()
        if year is None:
            year = datetime.now().year
        if month is None:
            month = datetime.now().month
        
        # 특정 월 삭제
        from calendar import monthrange
        days_in_month = monthrange(year, month)[1]
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, month, days_in_month).date()
        period_text = f"{year}년 {month}월"
        
        # 확인 대화상자
        reply = QMessageBox.question(
            self, 
            "데이터 삭제 확인",
            f"{period_text}의 모든 출퇴근 데이터를 삭제하시겠습니까?\n"
            f"삭제 기간: {start_date} ~ {end_date}\n\n"
            f"이 작업은 되돌릴 수 없습니다.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            try:
                # 해당 기간의 모든 출퇴근 기록 삭제
                cursor.execute("""
                    DELETE FROM attendance_records
                    WHERE work_date >= ? AND work_date <= ?
                """, (start_date, end_date))
                
                deleted_count = cursor.rowcount
                conn.commit()
                conn.close()
                
                QMessageBox.information(
                    self, 
                    "삭제 완료", 
                    f"{year}년 {month}월의 출퇴근 데이터 {deleted_count}건이 삭제되었습니다."
                )
                
                # 데이터 새로고침
                self.refresh_data()
                
            except Exception as e:
                conn.rollback()
                conn.close()
                QMessageBox.critical(self, "오류", f"데이터 삭제 중 오류 발생: {str(e)}")


class MainApplication(QMainWindow):
    """메인 애플리케이션"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1400, 800)
        
        # favicon.ico 아이콘 설정
        icon_path = Path(__file__).parent / "favicon.ico"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        else:
            # favicon.ico가 없으면 기본 아이콘 사용 (선택사항)
            # self.setWindowIcon(QIcon.fromTheme("application-x-executable"))
            pass
        
        self.db = DatabaseManager()
        self.leave_calculator = LeaveCalculator(self.db)
        self.attendance_calculator = AttendanceCalculator(self.db)
        
        tab_widget = QTabWidget()
        
        # 재직인원 탭 (먼저 생성)
        employee_gui = EmployeeManagementGUI(tab_widget, self.db, None, None)
        
        # 연월차 관리대장 탭 (재직인원 탭 참조 전달)
        leave_gui = LeaveManagementGUI(tab_widget, self.db, self.leave_calculator, employee_gui)
        
        # 출퇴근 관리대장 탭 (연월차 관리대장 참조 전달)
        attendance_gui = AttendanceManagementGUI(tab_widget, self.db, self.attendance_calculator, leave_gui, employee_gui)
        
        # 재직인원 탭에 다른 탭 참조 설정
        employee_gui.leave_gui = leave_gui
        employee_gui.attendance_gui = attendance_gui
        # 연월차 탭에서 출퇴근 탭 참조(통합 엑셀 다운로드용)
        leave_gui.attendance_gui = attendance_gui
        
        tab_widget.addTab(employee_gui, "재직인원")
        tab_widget.addTab(leave_gui, "연월차 관리대장")
        tab_widget.addTab(attendance_gui, "출퇴근 관리대장")
        
        self.setCentralWidget(tab_widget)
        
        # 상태바에 저작권 정보 추가
        status_bar = QStatusBar()
        copyright_label = QLabel("© 월드베스트로지스틱스 | 만든이: AQMAN")
        copyright_label.setStyleSheet("color: gray; padding: 2px;")
        status_bar.addPermanentWidget(copyright_label)
        self.setStatusBar(status_bar)


if __name__ == "__main__":
    import sys
    import os
    import subprocess
    
    # PySide6가 설치되어 있는지 확인
    if not PYSIDE6_AVAILABLE:
        python_exe = sys.executable
        print("=" * 60)
        print("오류: PySide6가 설치되지 않았습니다.")
        print("=" * 60)
        print("\n현재 사용 중인 Python 경로:")
        print(f"   {python_exe}")
        print("\nPython 버전:")
        print(f"   {sys.version.split()[0]}")
        print("\n오류 상세:")
        print(f"   {PYSIDE6_ERROR}")
        print("=" * 60)
        print("\n자동 설치를 시도하시겠습니까? (Y/N): ", end="", flush=True)
        
        try:
            user_input = input().strip().upper()
            if user_input == 'Y' or user_input == 'YES' or user_input == '':
                print("\nPySide6 설치 중...")
                try:
                    # 현재 Python 인터프리터로 PySide6 설치 시도
                    result = subprocess.run(
                        [python_exe, "-m", "pip", "install", "PySide6>=6.5.0"],
                        capture_output=True,
                        text=True,
                        timeout=300
                    )
                    
                    if result.returncode == 0:
                        print("\n✅ PySide6 설치가 완료되었습니다!")
                        print("프로그램을 다시 실행해주세요.")
                    else:
                        print("\n❌ 설치 중 오류가 발생했습니다:")
                        print(result.stderr)
                        print("\n수동 설치 명령어:")
                        print(f'   "{python_exe}" -m pip install PySide6>=6.5.0')
                except subprocess.TimeoutExpired:
                    print("\n❌ 설치 시간이 초과되었습니다.")
                    print("수동으로 설치해주세요.")
                except Exception as e:
                    print(f"\n❌ 설치 중 예외 발생: {str(e)}")
                    print("\n수동 설치 명령어:")
                    print(f'   "{python_exe}" -m pip install PySide6>=6.5.0')
            else:
                print("\n수동 설치 방법:")
                print("=" * 60)
                print("VS Code에서:")
                print("1. 터미널(Terminal) 메뉴 > 새 터미널 열기")
                print("2. 다음 명령어 실행:")
                print(f'   "{python_exe}" -m pip install PySide6>=6.5.0')
                print("\n또는 명령 프롬프트(CMD)에서:")
                print(f'   "{python_exe}" -m pip install PySide6>=6.5.0')
                print("\n또는 install_pyside6.bat 파일을 더블클릭")
                print("=" * 60)
        except (EOFError, KeyboardInterrupt):
            print("\n\n수동 설치 명령어:")
            print(f'   "{python_exe}" -m pip install PySide6>=6.5.0')
        
        # IDE에서 실행 중인지 확인
        if 'VSCODE_PID' in os.environ or 'PYCHARM_HOSTED' in os.environ:
            input("\n아무 키나 누르면 종료됩니다...")
        else:
            try:
                input("\n아무 키나 누르면 종료됩니다...")
            except:
                import time
                time.sleep(5)
        sys.exit(1)
    
    try:
        # 처리되지 않은 예외를 잡기 위한 핸들러 설정
        def exception_hook(exctype, value, tb):
            """처리되지 않은 예외를 잡아서 프로그램이 종료되지 않도록 함"""
            error_msg = ''.join(traceback.format_exception(exctype, value, tb))
            print(f"처리되지 않은 예외 발생:\n{error_msg}")
            # QMessageBox를 사용할 수 있는 경우에만 표시
            try:
                if PYSIDE6_AVAILABLE:
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Critical)
                    msg.setWindowTitle("오류")
                    msg.setText(f"처리되지 않은 오류가 발생했습니다.\n\n{str(value)}")
                    msg.setDetailedText(error_msg)
                    msg.exec()
            except:
                pass
        
        sys.excepthook = exception_hook
        
        # QApplication 생성
        app = QApplication(sys.argv)
        
        # favicon.ico 아이콘 설정 (애플리케이션 전체 아이콘)
        icon_path = Path(__file__).parent / "favicon.ico"
        if icon_path.exists():
            app.setWindowIcon(QIcon(str(icon_path)))
        
        # 메인 윈도우 생성
        window = MainApplication()
        window.show()
        
        # 이벤트 루프 시작
        sys.exit(app.exec())
    except Exception as e:
        print(f"\n프로그램 실행 중 오류가 발생했습니다:")
        print(f"오류: {str(e)}")
        print("\n오류 유형:", type(e).__name__)
        import traceback
        traceback.print_exc()
        
        # IDE에서 실행 중인지 확인
        if 'VSCODE_PID' in os.environ or 'PYCHARM_HOSTED' in os.environ:
            input("\n아무 키나 누르면 종료됩니다...")
        else:
            try:
                input("\n아무 키나 누르면 종료됩니다...")
            except:
                import time
                time.sleep(5)
        sys.exit(1)
